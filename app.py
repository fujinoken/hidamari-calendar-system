
# -*- coding: utf-8 -*-
"""
ひだまり帳 Ver1.4.7
PostgreSQL永続化版
Python + Streamlit + PostgreSQL

起動:
    streamlit run app.py

必要ライブラリ:
    pip install streamlit pandas openpyxl psycopg2-binary reportlab requests
"""

import os
import calendar
import re
import hashlib
import mimetypes
import urllib.parse
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from io import BytesIO

import pandas as pd
import streamlit as st

try:
    import requests
except ImportError:
    requests = None


try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import psycopg2
except ImportError:
    psycopg2 = None

try:
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OpenpyxlWorkbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False


APP_TITLE = "ひだまり帳 Ver1.4.7 PostgreSQL版"
AI_SHIFT_RULE_VERSION = "shift_overlap_strict_v6_staff_deduplicate_excel_export"
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
FILE_DIR = Path("attached_files")
FILE_DIR.mkdir(exist_ok=True)

DB_INTEGRITY_ERROR = psycopg2.IntegrityError if psycopg2 else Exception


# -----------------------------
# DB（PostgreSQL / Supabase対応）
# -----------------------------
def get_database_url():
    """
    PostgreSQL接続URLを取得する。
    Streamlit Cloudでは st.secrets["DATABASE_URL"] を推奨。
    ローカルでは環境変数 DATABASE_URL / SUPABASE_DB_URL でも動作する。
    """
    try:
        if "DATABASE_URL" in st.secrets:
            return st.secrets["DATABASE_URL"]
        if "SUPABASE_DB_URL" in st.secrets:
            return st.secrets["SUPABASE_DB_URL"]
        if "postgres" in st.secrets and "url" in st.secrets["postgres"]:
            return st.secrets["postgres"]["url"]
    except Exception:
        pass
    return os.getenv("DATABASE_URL") or os.getenv("SUPABASE_DB_URL")


def to_pg_query(query):
    """既存コードの ? プレースホルダを PostgreSQL/psycopg2 の %s へ変換する。"""
    return str(query).replace("?", "%s")


def get_conn():
    if psycopg2 is None:
        st.error("psycopg2 がインストールされていません。requirements.txt に psycopg2-binary を追加してください。")
        st.stop()

    database_url = get_database_url()
    if not database_url:
        st.error("PostgreSQL接続URLが未設定です。Streamlit secrets に DATABASE_URL を設定してください。")
        st.stop()

    # SupabaseではSSL必須のことが多いため、URLにsslmodeが無い場合はrequireを付ける。
    if "sslmode=" in database_url:
        return psycopg2.connect(database_url)
    return psycopg2.connect(database_url, sslmode=os.getenv("PGSSLMODE", "require"))


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id SERIAL PRIMARY KEY,
        user_id TEXT UNIQUE,
        user_name TEXT NOT NULL UNIQUE,
        kana TEXT,
        room_no TEXT,
        note TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS staff (
        id SERIAL PRIMARY KEY,
        staff_name TEXT NOT NULL UNIQUE,
        role TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS categories (
        id SERIAL PRIMARY KEY,
        category_name TEXT NOT NULL UNIQUE,
        mark TEXT,
        sort_order INTEGER DEFAULT 100,
        is_active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS events (
        id SERIAL PRIMARY KEY,
        event_date TEXT NOT NULL,
        category TEXT NOT NULL,
        title TEXT NOT NULL,
        user_id TEXT,
        user_name TEXT,
        staff_name TEXT,
        start_time TEXT,
        end_time TEXT,
        memo TEXT,
        important INTEGER DEFAULT 0,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS event_photos (
        id SERIAL PRIMARY KEY,
        event_id INTEGER NOT NULL REFERENCES events(id) ON DELETE CASCADE,
        file_name TEXT NOT NULL,
        file_path TEXT NOT NULL,
        photo_memo TEXT,
        created_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS event_files (
        id SERIAL PRIMARY KEY,
        event_id INTEGER NOT NULL REFERENCES events(id) ON DELETE CASCADE,
        file_name TEXT NOT NULL,
        file_path TEXT NOT NULL,
        file_type TEXT,
        file_memo TEXT,
        created_at TEXT NOT NULL
    )
    """)

    # 既存PostgreSQLテーブルからの移行：列がなければ追加する
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS user_id TEXT")
    cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS room_no TEXT")
    cur.execute("ALTER TABLE events ADD COLUMN IF NOT EXISTS user_id TEXT")

    # 既存利用者にIDがない場合、U0001形式で仮IDを付与
    cur.execute("SELECT id, user_id FROM users")
    for row_id, user_id in cur.fetchall():
        if not user_id:
            cur.execute("UPDATE users SET user_id=%s WHERE id=%s", (f"U{int(row_id):04d}", int(row_id)))

    # 既存予定にIDがない場合、利用者名から補完
    cur.execute("""
        UPDATE events
        SET user_id = (
            SELECT users.user_id FROM users
            WHERE users.user_name = events.user_name
            LIMIT 1
        )
        WHERE (user_id IS NULL OR user_id = '') AND user_name IS NOT NULL
    """)

    # カテゴリマスタ初期投入
    cur.execute("SELECT COUNT(*) FROM categories")
    category_count = cur.fetchone()[0]
    if category_count == 0:
        default_marks = {
            "通院": "🏥",
            "面会": "👪",
            "行事": "🎉",
            "外出": "🚶",
            "注意": "⚠️",
            "申し送り": "📝",
            "夜勤": "🌙",
            "その他": "・",
        }
        for i, name in enumerate(DEFAULT_CATEGORIES, start=1):
            cur.execute("""
                INSERT INTO categories
                (category_name, mark, sort_order, is_active, created_at, updated_at)
                VALUES (%s, %s, %s, 1, %s, %s)
                ON CONFLICT (category_name) DO NOTHING
            """, (name, default_marks.get(name, "・"), i * 10, now_text(), now_text()))

    cur.execute("""
    CREATE TABLE IF NOT EXISTS staff_shifts (
        id SERIAL PRIMARY KEY,
        shift_date TEXT NOT NULL,
        staff_name TEXT NOT NULL,
        shift_kind TEXT NOT NULL,
        start_time TEXT,
        end_time TEXT,
        next_day INTEGER DEFAULT 0,
        memo TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS shift_month_status (
        id SERIAL PRIMARY KEY,
        shift_year INTEGER NOT NULL,
        shift_month INTEGER NOT NULL,
        is_confirmed INTEGER DEFAULT 0,
        confirmed_at TEXT,
        confirmed_by TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS staff_shift_limits (
        id SERIAL PRIMARY KEY,
        staff_name TEXT NOT NULL UNIQUE,
        max_day_shifts INTEGER DEFAULT 31,
        max_night_shifts INTEGER DEFAULT 31,
        max_total_shifts INTEGER DEFAULT 31,
        note TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL
    )
    """)

    # よく使う検索用インデックス。既存DBにも安全に追加できる。
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_event_date ON events(event_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_date_start ON events(event_date, start_time)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_events_user_name ON events(user_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_event_photos_event_id ON event_photos(event_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_event_files_event_id ON event_files(event_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_categories_active_sort ON categories(is_active, sort_order)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_active_name ON users(is_active, user_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_staff_active_name ON staff(is_active, staff_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_staff_shifts_date ON staff_shifts(shift_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_staff_shifts_staff ON staff_shifts(staff_name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_staff_shifts_date_kind ON staff_shifts(shift_date, shift_kind)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_shift_month_status_ym ON shift_month_status(shift_year, shift_month)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_staff_shift_limits_staff ON staff_shift_limits(staff_name)")

    conn.commit()
    cur.close()
    conn.close()


JST = ZoneInfo("Asia/Tokyo")


def now_text():
    """
    Streamlit Cloud上でも日本時間で保存する。
    """
    return datetime.now(JST).strftime("%Y-%m-%d %H:%M:%S")


def today_jst():
    return datetime.now(JST).date()


def normalize_query_params(params=()):
    """cache用にSQLパラメータをtupleへ正規化する。"""
    if params is None:
        return ()
    if isinstance(params, tuple):
        return params
    if isinstance(params, list):
        return tuple(params)
    return (params,)


@st.cache_data(ttl=20, show_spinner=False)
def _fetch_df_cached(query, params_tuple=()):
    """
    読み取りSQLを短時間キャッシュする。
    予定表示・マスタ取得・件数確認の体感速度を上げる。
    """
    conn = get_conn()
    try:
        df = pd.read_sql_query(to_pg_query(query), conn, params=params_tuple)
    finally:
        conn.close()
    return df


def fetch_df(query, params=()):
    return _fetch_df_cached(str(query), normalize_query_params(params))


def clear_read_cache():
    """DB更新後に読み取りキャッシュをクリアする。"""
    try:
        st.cache_data.clear()
    except Exception:
        pass


def execute(query, params=()):
    """
    INSERT/UPDATE/DELETEを実行する。
    INSERTの場合は自動で RETURNING id を付け、登録IDを返す。
    """
    conn = get_conn()
    cur = conn.cursor()
    q = to_pg_query(query).strip()
    q_no_semicolon = q[:-1].strip() if q.endswith(";") else q
    is_insert = q_no_semicolon.lower().startswith("insert")
    if is_insert and " returning " not in q_no_semicolon.lower():
        q_exec = q_no_semicolon + " RETURNING id"
    else:
        q_exec = q_no_semicolon

    try:
        cur.execute(q_exec, normalize_query_params(params))
        last_id = None
        if is_insert:
            row = cur.fetchone()
            last_id = row[0] if row else None
        conn.commit()
        clear_read_cache()
        return last_id
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def execute_many(query, params_list):
    """複数行のINSERT/UPDATEをまとめて実行する。"""
    if not params_list:
        return 0
    conn = get_conn()
    cur = conn.cursor()
    q = to_pg_query(query).strip()
    try:
        cur.executemany(q, [normalize_query_params(p) for p in params_list])
        conn.commit()
        clear_read_cache()
        return len(params_list)
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


def init_db_once():
    """
    init_dbはALTER TABLEや初期投入を含むため、毎回実行せずセッション内で1回だけ実行する。
    これだけで画面切替・再描画が1テンポ軽くなる。
    """
    if not st.session_state.get("_hidamari_db_initialized", False):
        init_db()
        st.session_state["_hidamari_db_initialized"] = True
        clear_read_cache()




# -----------------------------
# Supabase Storage（第1段階：新規ファイル保存）
# -----------------------------
STORAGE_PATH_PREFIX = "storage://"


def get_secret_or_env(*keys, default=None):
    """Streamlit secrets または環境変数から設定値を取得する。"""
    for key in keys:
        try:
            if key in st.secrets:
                value = st.secrets[key]
                if value:
                    return str(value).strip()
        except Exception:
            pass
        value = os.getenv(key)
        if value:
            return str(value).strip()
    return default


def get_supabase_url():
    """
    Supabase Project URLを取得する。
    /rest/v1 や /storage/v1 まで入れてしまった場合も、プロジェクトURLへ補正する。
    """
    url = get_secret_or_env("SUPABASE_URL", "SUPABASE_PROJECT_URL", default="")
    if not url:
        return ""
    url = str(url).strip().rstrip("/")
    for suffix in ["/rest/v1", "/storage/v1", "/storage/v1/object"]:
        if url.endswith(suffix):
            url = url[: -len(suffix)]
    return url.rstrip("/")


def get_supabase_storage_key():
    """
    Supabase Storage操作用キーを取得する。
    非公開バケットをサーバー側から扱うため、Streamlit secretsには SERVICE_ROLE_KEY を入れる。
    """
    return get_secret_or_env("SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_SERVICE_KEY", "SUPABASE_STORAGE_KEY", default="")


def get_supabase_storage_bucket():
    """Storageバケット名。未指定時は hidamari-calendar-files を使う。"""
    return get_secret_or_env("SUPABASE_STORAGE_BUCKET", default="hidamari-calendar-files")


def storage_is_configured():
    """Supabase Storage保存に必要な設定が揃っているか確認する。"""
    return bool(get_supabase_url() and get_supabase_storage_key() and get_supabase_storage_bucket())


def require_storage_ready():
    """Storage未設定やrequests未導入の場合に分かりやすいエラーを出す。"""
    if requests is None:
        raise RuntimeError("requests がインストールされていません。requirements.txt に requests を追加してください。")
    if not get_supabase_url():
        raise RuntimeError("SUPABASE_URL が未設定です。Streamlit secrets に設定してください。")
    if not get_supabase_storage_key():
        raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY が未設定です。Streamlit secrets に設定してください。")
    if not get_supabase_storage_bucket():
        raise RuntimeError("SUPABASE_STORAGE_BUCKET が未設定です。")


def guess_content_type(file_name):
    content_type, _ = mimetypes.guess_type(str(file_name or ""))
    return content_type or "application/octet-stream"


def make_storage_object_path(event_id, original_name, kind):
    """
    Storage内の保存パスを作る。
    元ファイル名はDBのfile_nameに保存し、Storage側は衝突しにくい安全な名前にする。
    """
    suffix = Path(str(original_name or "")).suffix.lower()
    if not suffix:
        suffix = ".bin"
    timestamp = datetime.now(JST).strftime("%Y%m%d%H%M%S%f")
    digest = hashlib.md5(f"{event_id}-{original_name}-{timestamp}".encode("utf-8")).hexdigest()[:12]
    safe_kind = "photos" if kind == "photos" else "files"
    return f"events/{int(event_id)}/{safe_kind}/{timestamp}_{digest}{suffix}"


def make_storage_db_path(bucket, object_path):
    """DBには storage://bucket/path の形式で保存する。"""
    return f"{STORAGE_PATH_PREFIX}{bucket}/{object_path}"


def parse_storage_db_path(file_path):
    """storage://bucket/path を (bucket, path) に分解する。"""
    value = str(file_path or "").strip()
    if not value.startswith(STORAGE_PATH_PREFIX):
        return None, None
    rest = value[len(STORAGE_PATH_PREFIX):]
    if "/" not in rest:
        return None, None
    bucket, object_path = rest.split("/", 1)
    return bucket, object_path


def is_storage_file(file_path):
    return str(file_path or "").strip().startswith(STORAGE_PATH_PREFIX)


def storage_url_for_object(bucket, object_path, authenticated=False):
    """
    Supabase Storage REST URLを作る。
    upload/delete は /object、非公開ファイルの取得は /object/authenticated を使う。
    """
    base_url = get_supabase_url()
    quoted_bucket = urllib.parse.quote(str(bucket), safe="")
    quoted_path = urllib.parse.quote(str(object_path), safe="/")
    prefix = "object/authenticated" if authenticated else "object"
    return f"{base_url}/storage/v1/{prefix}/{quoted_bucket}/{quoted_path}"


def storage_headers(content_type=None, upsert=False):
    key = get_supabase_storage_key()
    headers = {
        "Authorization": f"Bearer {key}",
        "apikey": key,
    }
    if content_type:
        headers["Content-Type"] = content_type
    if upsert:
        headers["x-upsert"] = "true"
    return headers


def upload_bytes_to_storage(file_bytes, object_path, content_type):
    """Supabase Storageへファイル本体をアップロードする。"""
    require_storage_ready()
    bucket = get_supabase_storage_bucket()
    url = storage_url_for_object(bucket, object_path)
    response = requests.post(
        url,
        headers=storage_headers(content_type=content_type, upsert=True),
        data=file_bytes,
        timeout=90,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"Supabase Storageへのアップロードに失敗しました: {response.status_code} {response.text[:300]}")
    return make_storage_db_path(bucket, object_path)


def download_bytes_from_storage(file_path):
    """storage:// のDBパスからStorage上のファイル本体を取得する。"""
    require_storage_ready()
    bucket, object_path = parse_storage_db_path(file_path)
    if not bucket or not object_path:
        raise RuntimeError("Storageパスの形式が不正です。")
    url = storage_url_for_object(bucket, object_path, authenticated=True)
    response = requests.get(url, headers=storage_headers(), timeout=90)
    if response.status_code >= 400:
        raise RuntimeError(f"Supabase Storageからの取得に失敗しました: {response.status_code} {response.text[:300]}")
    return response.content


def delete_storage_object(file_path):
    """
    Storage上のファイル削除。
    環境やAPI仕様差に備え、失敗しても画面操作を止めないためFalseを返す。
    """
    if not is_storage_file(file_path):
        return False
    try:
        require_storage_ready()
        bucket, object_path = parse_storage_db_path(file_path)
        if not bucket or not object_path:
            return False
        url = storage_url_for_object(bucket, object_path)
        response = requests.delete(url, headers=storage_headers(), timeout=60)
        return response.status_code < 400
    except Exception:
        return False


@st.cache_data(ttl=300, show_spinner=False)
def _read_saved_file_bytes_cached(file_path):
    """
    Storage保存・旧ローカル保存の両方に対応してファイル本体を読む。
    画像や添付の再表示を速くするため、短時間キャッシュする。
    """
    if is_storage_file(file_path):
        return download_bytes_from_storage(file_path)

    local_path = Path(str(file_path or ""))
    if local_path.exists():
        with open(local_path, "rb") as f:
            return f.read()
    return None


def read_saved_file_bytes(file_path):
    return _read_saved_file_bytes_cached(str(file_path or ""))


def render_saved_image(file_path, caption=None, use_container_width=True):
    """Storage保存・旧ローカル保存の両方に対応して画像を表示する。"""
    try:
        data = read_saved_file_bytes(file_path)
        if data:
            st.image(data, caption=caption, use_container_width=use_container_width)
            return True
    except Exception as e:
        st.warning(f"画像を取得できません：{e}")
        return False

    st.warning("画像ファイルが見つかりません。")
    return False


def render_saved_download_button(label, file_path, file_name, key):
    """Storage保存・旧ローカル保存の両方に対応してダウンロードボタンを表示する。"""
    try:
        data = read_saved_file_bytes(file_path)
        if data:
            st.download_button(
                label,
                data=data,
                file_name=file_name,
                key=key,
            )
            return True
    except Exception as e:
        st.warning(f"ファイルを取得できません：{e}")
        return False

    st.warning("ファイルが見つかりません。")
    return False


def delete_saved_file(file_path):
    """Storage保存・旧ローカル保存の両方に対応してファイル本体を削除する。"""
    if is_storage_file(file_path):
        return delete_storage_object(file_path)

    local_path = Path(str(file_path or ""))
    if local_path.exists():
        try:
            local_path.unlink()
            return True
        except Exception:
            return False
    return False


def storage_object_exists(file_path):
    """
    DBに保存されたfile_pathが実際に読めるか確認する。
    Storage保存はHEAD、失敗時はGETで確認。旧ローカル保存はファイル存在で確認。
    """
    value = str(file_path or "").strip()
    if not value:
        return False, "file_path空欄"

    if is_storage_file(value):
        try:
            require_storage_ready()
            bucket, object_path = parse_storage_db_path(value)
            if not bucket or not object_path:
                return False, "Storageパス形式不正"

            url = storage_url_for_object(bucket, object_path, authenticated=True)
            response = requests.head(url, headers=storage_headers(), timeout=30)
            if response.status_code < 400:
                return True, "OK"

            # 環境によってHEADが許可されない場合の保険
            response = requests.get(url, headers=storage_headers(), timeout=30)
            if response.status_code < 400:
                return True, "OK"

            return False, f"Storage取得不可 {response.status_code}: {response.text[:120]}"
        except Exception as e:
            return False, f"Storage確認エラー: {e}"

    local_path = Path(value)
    if local_path.exists():
        return True, "OK（旧ローカル）"
    return False, "旧ローカルファイルなし"


def check_storage_bucket_access():
    """Storageバケットにアクセスできるか簡易確認する。"""
    if requests is None:
        return False, "requests が未導入です。"
    if not storage_is_configured():
        return False, "Storage設定が未完了です。"

    try:
        bucket = get_supabase_storage_bucket()
        base_url = get_supabase_url()
        quoted_bucket = urllib.parse.quote(str(bucket), safe="")
        url = f"{base_url}/storage/v1/object/list/{quoted_bucket}"
        response = requests.post(
            url,
            headers=storage_headers(content_type="application/json"),
            json={"prefix": "", "limit": 1, "offset": 0},
            timeout=30,
        )
        if response.status_code < 400:
            return True, "Storageバケットへ接続できました。"
        return False, f"Storageバケット確認失敗 {response.status_code}: {response.text[:180]}"
    except Exception as e:
        return False, f"Storageバケット確認エラー: {e}"


def count_query(sql, params=()):
    """COUNT系SQLを安全にintで返す。"""
    try:
        df = fetch_df(sql, params)
        if df is None or df.shape[0] == 0 or df.shape[1] == 0:
            return 0
        value = df.iat[0, 0]
        if pd.isna(value):
            return 0
        return int(value or 0)
    except Exception as e:
        st.warning(f"件数取得に失敗しました：{e}")
        return 0




def save_uploaded_photos(event_id, uploaded_files, photo_memo=""):
    """
    アップロード写真をSupabase Storageへ保存し、DBへ紐づける。
    戻り値: (保存成功数, 失敗数)
    """
    if not uploaded_files:
        return 0, 0

    saved_count = 0
    failed_count = 0

    if not storage_is_configured():
        st.error("Supabase Storage設定が未完了のため、写真メモは保存されませんでした。")
        return 0, len(uploaded_files)

    for uploaded in uploaded_files:
        try:
            object_path = make_storage_object_path(event_id, uploaded.name, kind="photos")
            file_bytes = bytes(uploaded.getbuffer())
            storage_path = upload_bytes_to_storage(
                file_bytes,
                object_path,
                guess_content_type(uploaded.name),
            )

            photo_id = execute("""
                INSERT INTO event_photos
                (event_id, file_name, file_path, photo_memo, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (
                int(event_id),
                uploaded.name,
                storage_path,
                photo_memo.strip() or None,
                now_text(),
            ))
            if photo_id:
                saved_count += 1
        except Exception as e:
            failed_count += 1
            st.error(f"写真メモのStorage保存に失敗しました：{uploaded.name}｜{e}")

    return saved_count, failed_count


@st.cache_data(ttl=20, show_spinner=False)
def get_event_photos(event_id):
    return fetch_df(
        "SELECT * FROM event_photos WHERE event_id=? ORDER BY id",
        (int(event_id),)
    )


def save_uploaded_files(event_id, uploaded_files, file_memo=""):
    """
    Excel等の添付ファイルをSupabase Storageへ保存し、DBへ紐づける。
    戻り値: (保存成功数, 失敗数)
    """
    if not uploaded_files:
        return 0, 0

    saved_count = 0
    failed_count = 0

    if not storage_is_configured():
        st.error("Supabase Storage設定が未完了のため、Excel・書類ファイルは保存されませんでした。")
        return 0, len(uploaded_files)

    for uploaded in uploaded_files:
        try:
            suffix = Path(uploaded.name).suffix.lower()
            object_path = make_storage_object_path(event_id, uploaded.name, kind="files")
            file_bytes = bytes(uploaded.getbuffer())
            storage_path = upload_bytes_to_storage(
                file_bytes,
                object_path,
                guess_content_type(uploaded.name),
            )

            file_id = execute("""
                INSERT INTO event_files
                (event_id, file_name, file_path, file_type, file_memo, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                int(event_id),
                uploaded.name,
                storage_path,
                suffix.replace(".", ""),
                file_memo.strip() or None,
                now_text(),
            ))
            if file_id:
                saved_count += 1
        except Exception as e:
            failed_count += 1
            st.error(f"Excel・書類ファイルのStorage保存に失敗しました：{uploaded.name}｜{e}")

    return saved_count, failed_count


@st.cache_data(ttl=20, show_spinner=False)
def get_event_files(event_id):
    return fetch_df(
        "SELECT * FROM event_files WHERE event_id=? ORDER BY id",
        (int(event_id),)
    )


@st.cache_data(ttl=60, show_spinner=False)
def get_active_users():
    return fetch_df("SELECT user_id, user_name FROM users WHERE is_active=1 ORDER BY user_name")


def user_display_map():
    df = get_active_users()
    mapping = {"": ("", "")}
    if not df.empty:
        for _, r in df.iterrows():
            label = f"{r['user_name']}（ID:{r['user_id']}）"
            mapping[label] = (r["user_id"], r["user_name"])
    return mapping


@st.cache_data(ttl=60, show_spinner=False)
def get_active_staff():
    """
    有効な職員名を取得する。
    「職員 A」と「職員A」のような空白違いは同一職員として扱い、
    シフト表に重複行が出ないようにする。
    """
    df = fetch_df("SELECT staff_name FROM staff WHERE is_active=1 ORDER BY staff_name")
    if df.empty:
        return []

    names = []
    seen = set()
    for raw_name in df["staff_name"].dropna().astype(str).tolist():
        name = normalize_staff_name(raw_name)
        if not name or name in seen:
            continue
        seen.add(name)
        names.append(name)
    return names


# -----------------------------
# UI helpers
# -----------------------------
DEFAULT_CATEGORIES = ["通院", "面会", "行事", "外出", "注意", "申し送り", "夜勤", "その他"]

CATEGORY_MARK = {
    "通院": "🏥",
    "面会": "👪",
    "行事": "🎉",
    "外出": "🚶",
    "注意": "⚠️",
    "申し送り": "📝",
    "夜勤": "🌙",
    "その他": "・",
}


@st.cache_data(ttl=60, show_spinner=False)
def get_categories(active_only=True):
    where = "WHERE is_active=1" if active_only else ""
    df = fetch_df(f"""
        SELECT category_name, mark, sort_order, is_active
        FROM categories
        {where}
        ORDER BY sort_order, category_name
    """)
    if df.empty and active_only:
        return DEFAULT_CATEGORIES
    return df["category_name"].tolist()


@st.cache_data(ttl=60, show_spinner=False)
def get_category_mark_map():
    df = fetch_df("""
        SELECT category_name, mark
        FROM categories
        WHERE is_active=1
    """)
    mapping = {}
    if not df.empty:
        for _, r in df.iterrows():
            name = str(r["category_name"] or "").strip()
            mark = str(r["mark"] or "").strip()
            if name and mark:
                mapping[name] = mark
    return mapping


def get_category_mark(category_name):
    mark_map = get_category_mark_map()
    category_name = str(category_name or "").strip()
    return mark_map.get(category_name, CATEGORY_MARK.get(category_name, "・"))


def add_css():
    st.markdown("""
    <style>
    .main .block-container {
        padding-top: 1.2rem;
        max-width: 1200px;
    }
    .calendar-title {
        font-size: 1.6rem;
        font-weight: 700;
        margin: 12px 0 10px 0;
        color: #3f3a35;
    }
    .calendar-grid {
        display: grid;
        grid-template-columns: repeat(7, minmax(0, 1fr));
        gap: 0;
        width: 100%;
        border-top: 1px solid #d8d0c4;
        border-left: 1px solid #d8d0c4;
        background: #fffdf8;
    }
    .day-head {
        font-weight: 700;
        text-align: center;
        padding: 8px 4px;
        border-right: 1px solid #d8d0c4;
        border-bottom: 1px solid #d8d0c4;
        background: #f3eee6;
        color: #4b4035;
        font-size: 0.82rem;
        letter-spacing: 0.08em;
    }
    .day-cell {
        min-height: 135px;
        border-right: 1px solid #d8d0c4;
        border-bottom: 1px solid #d8d0c4;
        padding: 8px;
        background: #fffdf8;
        box-sizing: border-box;
        overflow: hidden;
    }
    .blank-cell {
        background: #fffdf8;
    }
    .day-cell-muted {
        min-height: 150px;
        border: 1px solid #eee6dc;
        border-radius: 4px;
        padding: 8px;
        background: #fffdf8;
        color: #aaa;
    }
    .day-num {
        font-size: 2.4rem;
        font-weight: 700;
        line-height: 1;
        margin-bottom: 8px;
        letter-spacing: -1px;
    }
    .write-lines {
        height: 34px;
        margin: 4px 0 6px 0;
        background-image: repeating-linear-gradient(
            to bottom,
            transparent 0px,
            transparent 13px,
            #d9d2c8 14px
        );
        opacity: 0.75;
    }
    .sunday { color: #c0392b; }
    .saturday { color: #1f4e79; }
    .event-line {
        font-size: 0.78rem;
        line-height: 1.35;
        margin: 3px 0;
        padding: 3px 5px;
        border-radius: 5px;
        background: #f6efe6;
        overflow-wrap: anywhere;
        border-left: 3px solid #bfae9b;
    }
    .important {
        background: #ffe9e0;
        border-left: 4px solid #d65a31;
    }
    .small-note {
        color: #7a6a5b;
        font-size: 0.9rem;
    }

    .today-board-card {
        border: 2px solid #d8d0c4;
        border-left: 10px solid #bfae9b;
        background: #fffdf8;
        border-radius: 8px;
        padding: 14px 16px;
        margin: 10px 0;
        box-shadow: none;
    }
    .today-board-card-important {
        border-left: 10px solid #d65a31;
        background: #fff3ee;
    }
    .today-board-main {
        font-size: 1.35rem;
        font-weight: 800;
        line-height: 1.35;
        color: #3f3a35;
    }
    .today-board-time {
        display: inline-block;
        min-width: 76px;
        font-size: 1.45rem;
        font-weight: 900;
        color: #2f2a25;
    }
    .today-board-memo {
        font-size: 1.05rem;
        margin-top: 8px;
        padding-left: 82px;
        color: #4f463d;
    }
    .today-board-sub {
        font-size: 0.9rem;
        margin-top: 8px;
        padding-left: 82px;
        color: #7a6a5b;
    }
    .today-board-badge {
        display: inline-block;
        padding: 2px 8px;
        border-radius: 999px;
        background: #f2eadf;
        font-size: 0.82rem;
        margin-right: 4px;
    }
    .today-summary-box {
        border: 2px solid #d8d0c4;
        background: #f9f5ee;
        border-radius: 8px;
        padding: 12px 14px;
        margin: 8px 0 14px 0;
        font-size: 1.05rem;
        font-weight: 700;
    }

    </style>
    """, unsafe_allow_html=True)


@st.cache_data(ttl=30, show_spinner=False)
def monthly_events(year, month):
    start = f"{year}-{month:02d}-01"
    last_day = calendar.monthrange(year, month)[1]
    end = f"{year}-{month:02d}-{last_day:02d}"

    df = fetch_df("""
        SELECT * FROM events
        WHERE event_date BETWEEN ? AND ?
        ORDER BY event_date, start_time, category, id
    """, (start, end))

    events_by_day = {}
    for _, row in df.iterrows():
        events_by_day.setdefault(row["event_date"], []).append(row)
    return events_by_day



@st.cache_data(ttl=20, show_spinner=False)
def events_by_date(target_date):
    """指定日の予定一覧を取得する。"""
    if hasattr(target_date, "strftime"):
        target_date = target_date.strftime("%Y-%m-%d")
    return fetch_df("""
        SELECT *
        FROM events
        WHERE event_date = ?
        ORDER BY
            CASE WHEN start_time IS NULL OR start_time = '' THEN 1 ELSE 0 END,
            start_time,
            category,
            id
    """, (str(target_date),))


def render_event_button_list(df, empty_message="予定はありません。", include_date=True):
    """予定一覧をボタン表示し、押すと同一ページに詳細を出す。"""
    if df is None or df.empty:
        st.info(empty_message)
        return

    cols = st.columns(3)
    for i, (_, ev) in enumerate(df.iterrows()):
        mark = get_category_mark(ev["category"])
        label_date = str(ev["event_date"])[5:] if include_date and ev["event_date"] else ""
        label_time = f" {ev['start_time']}" if ev["start_time"] else ""
        label_user = f"／{ev['user_name']}" if ev["user_name"] else ""
        important = "⚠️ " if int(ev["important"] or 0) == 1 else ""
        label = f"{label_date}{label_time} {important}{mark}{ev['title']}{label_user}".strip()

        with cols[i % 3]:
            if st.button(label, key=f"event_btn_{ev['id']}", use_container_width=True):
                set_selected_event(int(ev["id"]))
                st.rerun()



def first_line_text(text, max_len=42):
    """メモの1行目だけを短く表示する。"""
    if text is None:
        return ""
    value = str(text).strip()
    if not value:
        return ""
    first = value.splitlines()[0].strip()
    if len(first) > max_len:
        return first[:max_len] + "…"
    return first


@st.cache_data(ttl=20, show_spinner=False)
def get_attachment_counts(event_id):
    """写真・添付ファイル数を取得する。"""
    photos = fetch_df("SELECT COUNT(*) AS cnt FROM event_photos WHERE event_id=?", (int(event_id),))
    files = fetch_df("SELECT COUNT(*) AS cnt FROM event_files WHERE event_id=?", (int(event_id),))
    photo_count = int(photos.iloc[0]["cnt"]) if not photos.empty else 0
    file_count = int(files.iloc[0]["cnt"]) if not files.empty else 0
    return photo_count, file_count


def render_today_board(df):
    """
    今日画面専用のホワイトボード風一覧。
    クリックしなくても7割分かるように、時刻・カテゴリ・利用者・メモ1行・担当・添付を表示する。
    """
    if df is None or df.empty:
        st.info("今日の予定はありません。")
        return

    for _, ev in df.iterrows():
        important = int(ev["important"] or 0) == 1
        card_class = "today-board-card today-board-card-important" if important else "today-board-card"
        time_text = ev["start_time"] if ev["start_time"] else "時間未定"
        mark = get_category_mark(ev["category"])
        category = html_escape(ev["category"])
        title = html_escape(ev["title"])
        user_name = html_escape(ev["user_name"] or "")
        staff_name = html_escape(ev["staff_name"] or "")
        memo_line = html_escape(first_line_text(ev["memo"]))
        warning = "⚠️ " if important else ""

        main_line = f'{warning}<span class="today-board-time">{html_escape(time_text)}</span>{mark} {category}｜{title}'
        if user_name:
            main_line += f'｜{user_name}'

        memo_html = f'<div class="today-board-memo">メモ：{memo_line}</div>' if memo_line else ""

        sub_items = []
        if staff_name:
            sub_items.append(f'<span class="today-board-badge">担当：{staff_name}</span>')

        photo_count, file_count = get_attachment_counts(ev["id"])
        if photo_count:
            sub_items.append(f'<span class="today-board-badge">写真 {photo_count}</span>')
        if file_count:
            sub_items.append(f'<span class="today-board-badge">添付 {file_count}</span>')

        sub_html = f'<div class="today-board-sub">{" ".join(sub_items)}</div>' if sub_items else ""

        html = (
            f'<div class="{card_class}">'
            f'<div class="today-board-main">{main_line}</div>'
            f'{memo_html}'
            f'{sub_html}'
            f'</div>'
        )
        st.markdown(html, unsafe_allow_html=True)

        if st.button("詳細を見る", key=f"today_detail_{ev['id']}", use_container_width=True):
            set_selected_event(int(ev["id"]))
            st.rerun()


def html_escape(text):
    """HTML表示用の簡易エスケープ。"""
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#x27;")
    )


def set_selected_event(event_id):
    st.session_state["selected_calendar_event_id"] = int(event_id)


@st.cache_data(ttl=20, show_spinner=False)
def get_event_by_id(event_id):
    if not event_id:
        return pd.DataFrame()
    return fetch_df("SELECT * FROM events WHERE id=? LIMIT 1", (int(event_id),))


def render_event_detail_panel():
    event_id = st.session_state.get("selected_calendar_event_id")
    if not event_id:
        return

    ev_df = get_event_by_id(event_id)
    if ev_df.empty:
        st.warning("選択された予定が見つかりません。")
        st.session_state["selected_calendar_event_id"] = None
        return

    ev = ev_df.iloc[0]
    st.markdown("---")
    st.subheader("予定詳細")
    st.markdown(f"### {ev['event_date']}｜{ev['category']}｜{ev['title']}")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.write(f"**利用者**：{ev['user_name'] or ''}")
        st.caption(f"利用者ID：{ev['user_id'] or ''}")
    with c2:
        st.write(f"**時間**：{ev['start_time'] or ''} 〜 {ev['end_time'] or ''}")
    with c3:
        st.write(f"**担当**：{ev['staff_name'] or ''}")

    if ev["important"]:
        st.warning("重要マークあり")

    st.write("**メモ**")
    st.info(ev["memo"] or "メモはありません。")

    photos = get_event_photos(event_id)
    if not photos.empty:
        st.write("**写真メモ**")
        cols = st.columns(3)
        for i, (_, p) in enumerate(photos.iterrows()):
            with cols[i % 3]:
                render_saved_image(
                    p["file_path"],
                    caption=p["photo_memo"] or p["file_name"],
                    use_container_width=True,
                )

    files = get_event_files(event_id)
    if not files.empty:
        st.write("**Excel・書類ファイル**")
        for _, frow in files.iterrows():
            st.write(f"📎 {frow['file_name']}　{frow['file_memo'] or ''}")
            render_saved_download_button(
                "ダウンロード",
                frow["file_path"],
                frow["file_name"],
                key=f"detail_download_{frow['id']}",
            )

    col_a, col_b = st.columns(2)
    with col_a:
        if st.button("詳細を閉じる", use_container_width=True):
            st.session_state["selected_calendar_event_id"] = None
            st.rerun()
    with col_b:
        st.caption("編集・削除は「予定検索・更新・削除」メニューで行います。")



def clear_last_saved_event_state():
    """保存後確認エリア用の一時情報を消す。"""
    for key in [
        "last_saved_event_id",
        "last_saved_photo_saved",
        "last_saved_photo_failed",
        "last_saved_file_saved",
        "last_saved_file_failed",
        "last_saved_had_photos",
        "last_saved_had_files",
    ]:
        st.session_state.pop(key, None)


def render_saved_event_confirmation(event_id=None):
    """
    予定登録直後に、保存した予定・写真メモ・添付ファイルの紐づきをその場で確認する。
    入力ミスやStorage保存失敗にすぐ気づけるようにするための確認パネル。
    """
    if event_id is None:
        event_id = st.session_state.get("last_saved_event_id")
    if not event_id:
        return

    ev_df = get_event_by_id(event_id)
    if ev_df.empty:
        st.warning("保存後確認：保存した予定が見つかりません。")
        return

    ev = ev_df.iloc[0]
    photos = get_event_photos(event_id)
    files = get_event_files(event_id)

    photo_saved = int(st.session_state.get("last_saved_photo_saved", len(photos)) or 0)
    photo_failed = int(st.session_state.get("last_saved_photo_failed", 0) or 0)
    file_saved = int(st.session_state.get("last_saved_file_saved", len(files)) or 0)
    file_failed = int(st.session_state.get("last_saved_file_failed", 0) or 0)
    had_photos = bool(st.session_state.get("last_saved_had_photos", False))
    had_files = bool(st.session_state.get("last_saved_had_files", False))

    st.markdown("---")
    st.subheader("保存後の確認")
    st.caption("いま保存した予定の内容と、写真・添付ファイルの紐づき状態を確認できます。")

    if photo_failed or file_failed:
        st.warning("予定本体は保存されていますが、写真メモまたは添付ファイルで保存失敗があります。")
    else:
        st.success("予定本体と、保存できた写真・添付ファイルの紐づきを確認できます。")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("予定ID", int(event_id))
    with c2:
        st.metric("写真紐づき", len(photos))
    with c3:
        st.metric("添付紐づき", len(files))
    with c4:
        ng_count = int(photo_failed + file_failed)
        st.metric("保存失敗", ng_count)

    if had_photos:
        if photo_failed:
            st.warning(f"写真メモ：{photo_saved}件保存、{photo_failed}件失敗")
        else:
            st.info(f"写真メモ：{photo_saved}件保存")
    if had_files:
        if file_failed:
            st.warning(f"Excel・書類：{file_saved}件保存、{file_failed}件失敗")
        else:
            st.info(f"Excel・書類：{file_saved}件保存")

    st.markdown("#### 保存した予定の詳細")
    d1, d2, d3 = st.columns(3)
    with d1:
        st.write(f"**日付**：{ev['event_date']}")
        st.write(f"**カテゴリ**：{ev['category']}")
        st.write(f"**タイトル**：{ev['title']}")
    with d2:
        st.write(f"**利用者**：{ev['user_name'] or ''}")
        st.caption(f"利用者ID：{ev['user_id'] or ''}")
        st.write(f"**担当**：{ev['staff_name'] or ''}")
    with d3:
        st.write(f"**時間**：{ev['start_time'] or ''} 〜 {ev['end_time'] or ''}")
        st.write(f"**重要**：{'あり' if int(ev['important'] or 0) == 1 else 'なし'}")
        st.caption(f"登録日時：{ev['created_at']}")

    st.write("**メモ**")
    st.info(ev["memo"] or "メモはありません。")

    st.markdown("#### 写真メモの確認")
    if photos.empty:
        if had_photos:
            st.warning("写真を選択していましたが、DB上の写真メモ紐づきは0件です。Storage保存失敗の可能性があります。")
        else:
            st.info("写真メモは添付されていません。")
    else:
        st.caption(f"DBに紐づいた写真メモ：{len(photos)}件")
        cols = st.columns(3)
        for i, (_, p) in enumerate(photos.iterrows()):
            with cols[i % 3]:
                render_saved_image(
                    p["file_path"],
                    caption=p["photo_memo"] or p["file_name"],
                    use_container_width=True,
                )
                st.caption(f"写真ID：{p['id']}")

    st.markdown("#### Excel・書類ファイルの確認")
    if files.empty:
        if had_files:
            st.warning("ファイルを選択していましたが、DB上の添付ファイル紐づきは0件です。Storage保存失敗の可能性があります。")
        else:
            st.info("Excel・書類ファイルは添付されていません。")
    else:
        st.caption(f"DBに紐づいた添付ファイル：{len(files)}件")
        for _, frow in files.iterrows():
            st.write(f"📎 **{frow['file_name']}**　{frow['file_memo'] or ''}")
            st.caption(f"ファイルID：{frow['id']} / 保存先：{frow['file_path']}")
            render_saved_download_button(
                "ダウンロードして確認",
                frow["file_path"],
                frow["file_name"],
                key=f"saved_confirm_download_{frow['id']}",
            )

    cc1, cc2 = st.columns(2)
    with cc1:
        if st.button("この予定を下の詳細パネルでも表示", use_container_width=True):
            set_selected_event(int(event_id))
            st.rerun()
    with cc2:
        if st.button("保存後確認を閉じる", use_container_width=True):
            clear_last_saved_event_state()
            st.rerun()




def render_calendar(year, month):
    """
    紙の壁カレンダー風レイアウト。
    カレンダーはHTMLで表示し、予定詳細は下部の予定ボタンから同じページ内に表示する。
    """
    events_by_day = monthly_events(year, month)
    first_weekday, last_day = calendar.monthrange(year, month)  # Monday=0, Sunday=6
    start_col = (first_weekday + 1) % 7

    cells = [""] * 42
    for day in range(1, last_day + 1):
        cells[start_col + day - 1] = day

    html = []
    html.append(f'<div class="calendar-title">{year}年 {month}月</div>')
    html.append('<div class="calendar-grid">')

    for h in ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]:
        html.append(f'<div class="day-head">{h}</div>')

    for idx, day in enumerate(cells):
        dow = idx % 7
        dow_cls = "sunday" if dow == 0 else ("saturday" if dow == 6 else "")

        if day == "":
            html.append('<div class="day-cell blank-cell"></div>')
            continue

        key = f"{year}-{month:02d}-{int(day):02d}"
        html.append('<div class="day-cell">')
        html.append(f'<div class="day-num {dow_cls}">{day}</div>')
        html.append('<div class="write-lines"></div>')

        for ev in events_by_day.get(key, []):
            mark = get_category_mark(ev["category"])
            time_part = f'{ev["start_time"]} ' if ev["start_time"] else ""
            user_part = f'／{ev["user_name"]}' if ev["user_name"] else ""
            imp_cls = " important" if int(ev["important"] or 0) == 1 else ""
            text = html_escape(f'{mark}{time_part}{ev["title"]}{user_part}')
            html.append(f'<div class="event-line{imp_cls}">{text}</div>')

        html.append('</div>')

    html.append('</div>')
    st.markdown("".join(html), unsafe_allow_html=True)

    # Streamlitの通常ボタンで詳細表示。HTML内クリックより安定。
    month_events = []
    for key in sorted(events_by_day.keys()):
        for ev in events_by_day[key]:
            month_events.append(ev)

    if month_events:
        st.markdown("### 予定詳細を表示")
        st.caption("下の予定ボタンから選ぶと、詳細が同じページ内に表示されます。")
        month_df = pd.DataFrame(month_events)
        render_event_button_list(month_df, empty_message="この月の予定はありません。", include_date=True)

    render_event_detail_panel()


# -----------------------------
# Pages
# -----------------------------
def page_calendar():
    today = today_jst()
    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        year = st.number_input("", min_value=2020, max_value=2100, value=today.year, step=1)
    with col2:
        month = st.number_input("", min_value=1, max_value=12, value=today.month, step=1)
    with col3:
        st.markdown('<div class="small-note"></div>', unsafe_allow_html=True)

    render_calendar(int(year), int(month))



def page_today():
    st.subheader("今日は何ある")
    st.caption("朝の申し送り前に、今日の予定をホワイトボード感覚で確認できます。")

    target = today_jst()
    target_text = target.strftime("%Y-%m-%d")
    st.markdown(f"## {target_text} の予定")

    df = events_by_date(target)

    if not df.empty:
        summary = df.groupby("category").size().reset_index(name="件数")
        summary_text = "　".join([
            f"{get_category_mark(r['category'])}{r['category']} {r['件数']}件"
            for _, r in summary.iterrows()
        ])

        important_df = df[df["important"].fillna(0).astype(int) == 1]
        important_text = f"　⚠️重要 {len(important_df)}件" if not important_df.empty else ""

        st.markdown(
            f'<div class="today-summary-box">本日の予定：{len(df)}件　{summary_text}{important_text}</div>',
            unsafe_allow_html=True
        )

        st.markdown("### 今日のホワイトボード")
        render_today_board(df)
    else:
        st.info("今日の予定はありません。")

    render_event_detail_panel()


def page_event_register():
    st.subheader("予定登録")
    if storage_is_configured():
        st.caption(f"写真・添付ファイルは新規登録分からSupabase Storageへ保存されます。Storage: {get_supabase_storage_bucket()} / URL: {get_supabase_url()}")
    else:
        st.warning("Supabase Storage設定が未完了です。写真・添付ファイルを保存するには SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY / SUPABASE_STORAGE_BUCKET を設定してください。")

    user_map = user_display_map()
    users = list(user_map.keys())
    staff = [""] + get_active_staff()

    with st.form("event_register_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            event_date = st.date_input("日付", value=today_jst())
            category = st.selectbox("カテゴリ", get_categories())
        with c2:
            start_time = st.text_input("開始時刻", placeholder="例：10:00")
            end_time = st.text_input("終了時刻", placeholder="例：11:00")
        with c3:
            user_name = st.selectbox("利用者", users)
            staff_name = st.selectbox("担当職員", staff)

        title = st.text_input("予定タイトル", placeholder="例：内科受診、家族面会、外出支援")
        memo = st.text_area("メモ", placeholder="必要な注意点、持ち物、申し送りなど")
        important = st.checkbox("重要マークを付ける")
        uploaded_photos = st.file_uploader(
            "写真メモ（複数可）",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            help="紙カレンダー、受診票、家族メモ、持ち物写真などを予定に紐づけできます。"
        )
        photo_memo = st.text_input("写真メモ補足", placeholder="例：受診予定表、家族からのメモ、持ち物確認")

        uploaded_files = st.file_uploader(
            "Excel・書類ファイル（複数可）",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            help="通院予定表、持ち物表、行事予定表、家族共有用Excelなどを予定に紐づけできます。"
        )
        file_memo = st.text_input("Excel・書類メモ補足", placeholder="例：通院予定表、買い物リスト、行事参加者表")

        submitted = st.form_submit_button("保存する")

    if submitted:
        if not title.strip():
            st.error("予定タイトルを入力してください。")
            return

        selected_user_id, selected_user_name = user_map.get(user_name, ("", ""))

        event_id = execute("""
            INSERT INTO events
            (event_date, category, title, user_id, user_name, staff_name, start_time, end_time, memo, important, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            event_date.strftime("%Y-%m-%d"),
            category,
            title.strip(),
            selected_user_id or None,
            selected_user_name or None,
            staff_name or None,
            start_time.strip() or None,
            end_time.strip() or None,
            memo.strip() or None,
            1 if important else 0,
            now_text(),
            now_text(),
        ))
        photo_saved, photo_failed = save_uploaded_photos(event_id, uploaded_photos, photo_memo)
        file_saved, file_failed = save_uploaded_files(event_id, uploaded_files, file_memo)

        st.session_state["selected_calendar_event_id"] = int(event_id)
        st.session_state["last_saved_event_id"] = int(event_id)
        st.session_state["last_saved_photo_saved"] = int(photo_saved)
        st.session_state["last_saved_photo_failed"] = int(photo_failed)
        st.session_state["last_saved_file_saved"] = int(file_saved)
        st.session_state["last_saved_file_failed"] = int(file_failed)
        st.session_state["last_saved_had_photos"] = bool(uploaded_photos)
        st.session_state["last_saved_had_files"] = bool(uploaded_files)

        messages = [f"予定を保存しました（予定ID: {event_id}）。"]
        if uploaded_photos:
            messages.append(f"写真メモ：{photo_saved}件保存")
            if photo_failed:
                messages.append(f"写真メモ：{photo_failed}件失敗")
        if uploaded_files:
            messages.append(f"Excel・書類：{file_saved}件保存")
            if file_failed:
                messages.append(f"Excel・書類：{file_failed}件失敗")

        if photo_failed or file_failed:
            st.warning(" / ".join(messages))
        else:
            st.success(" / ".join(messages))


    render_saved_event_confirmation()


def page_event_manage():
    st.subheader("予定検索・更新・削除")

    c1, c2, c3 = st.columns(3)
    with c1:
        start = st.date_input("開始日", value=today_jst().replace(day=1))
    with c2:
        end = st.date_input("終了日", value=today_jst())
    with c3:
        category_filter = st.selectbox("カテゴリ絞り込み", ["すべて"] + get_categories())

    keyword = st.text_input("キーワード検索", placeholder="タイトル・メモ・利用者名・職員名")

    query = """
        SELECT * FROM events
        WHERE event_date BETWEEN ? AND ?
    """
    params = [start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")]

    if category_filter != "すべて":
        query += " AND category = ?"
        params.append(category_filter)

    if keyword.strip():
        query += " AND (title LIKE ? OR memo LIKE ? OR user_name LIKE ? OR staff_name LIKE ?)"
        kw = f"%{keyword.strip()}%"
        params.extend([kw, kw, kw, kw])

    query += " ORDER BY event_date DESC, start_time, id DESC"
    df = fetch_df(query, params)

    if df.empty:
        st.info("該当する予定はありません。")
        return

    show_cols = ["id", "event_date", "category", "title", "user_id", "user_name", "staff_name", "start_time", "end_time", "important", "memo"]
    st.dataframe(df[show_cols], use_container_width=True, hide_index=True)

    selected_id = st.selectbox("編集・削除する予定ID", df["id"].tolist())
    target = df[df["id"] == selected_id].iloc[0]

    st.markdown("---")
    st.write("選択中の予定の写真メモ")

    photos = get_event_photos(selected_id)
    if photos.empty:
        st.info("この予定に紐づく写真メモはありません。")
    else:
        for _, p in photos.iterrows():
            c_img, c_info = st.columns([1, 2])
            with c_img:
                render_saved_image(
                    p["file_path"],
                    caption=p["file_name"],
                    use_container_width=True,
                )
            with c_info:
                st.write(f"メモ：{p['photo_memo'] or ''}")
                st.caption(f"登録日時：{p['created_at']}")
                if st.button(f"この写真を削除 ID:{p['id']}", key=f"delete_photo_{p['id']}"):
                    delete_saved_file(p["file_path"])
                    execute("DELETE FROM event_photos WHERE id=?", (int(p["id"]),))
                    st.success("写真メモを削除しました。画面を再読み込みしてください。")

    with st.expander("この予定に写真メモを追加する"):
        add_photos = st.file_uploader(
            "追加する写真",
            type=["png", "jpg", "jpeg", "webp"],
            accept_multiple_files=True,
            key=f"add_photos_{selected_id}"
        )
        add_photo_memo = st.text_input("追加写真メモ", key=f"add_photo_memo_{selected_id}")
        if st.button("写真メモを追加", key=f"add_photo_btn_{selected_id}"):
            saved, failed = save_uploaded_photos(selected_id, add_photos, add_photo_memo)
            if failed:
                st.warning(f"写真メモ：{saved}件保存、{failed}件失敗しました。")
            else:
                st.success(f"写真メモを{saved}件追加しました。画面を再読み込みしてください。")

    st.markdown("---")
    st.write("選択中の予定のExcel・書類ファイル")

    files = get_event_files(selected_id)
    if files.empty:
        st.info("この予定に紐づくExcel・書類ファイルはありません。")
    else:
        for _, frow in files.iterrows():
            c_file, c_action = st.columns([3, 1])
            with c_file:
                st.write(f"📎 **{frow['file_name']}**")
                st.write(f"メモ：{frow['file_memo'] or ''}")
                st.caption(f"登録日時：{frow['created_at']}")
                render_saved_download_button(
                    "ダウンロード",
                    frow["file_path"],
                    frow["file_name"],
                    key=f"download_file_{frow['id']}",
                )
            with c_action:
                if st.button(f"削除 ID:{frow['id']}", key=f"delete_file_{frow['id']}"):
                    delete_saved_file(frow["file_path"])
                    execute("DELETE FROM event_files WHERE id=?", (int(frow["id"]),))
                    st.success("ファイルを削除しました。画面を再読み込みしてください。")

    with st.expander("この予定にExcel・書類ファイルを追加する"):
        add_files = st.file_uploader(
            "追加するExcel・書類",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            key=f"add_files_{selected_id}"
        )
        add_file_memo = st.text_input("追加ファイルメモ", key=f"add_file_memo_{selected_id}")
        if st.button("Excel・書類ファイルを追加", key=f"add_file_btn_{selected_id}"):
            saved, failed = save_uploaded_files(selected_id, add_files, add_file_memo)
            if failed:
                st.warning(f"Excel・書類ファイル：{saved}件保存、{failed}件失敗しました。")
            else:
                st.success(f"Excel・書類ファイルを{saved}件追加しました。画面を再読み込みしてください。")

    st.markdown("---")
    st.write("選択中の予定を編集")

    user_map = user_display_map()
    users = list(user_map.keys())
    staff = [""] + get_active_staff()

    with st.form("event_update_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            new_date = st.date_input("日付", value=datetime.strptime(target["event_date"], "%Y-%m-%d").date())
            category_options = get_categories()
            if target["category"] and target["category"] not in category_options:
                category_options = [target["category"]] + category_options
            new_category = st.selectbox(
                "カテゴリ",
                category_options,
                index=category_options.index(target["category"]) if target["category"] in category_options else 0
            )
        with c2:
            new_start_time = st.text_input("開始時刻", value=target["start_time"] or "")
            new_end_time = st.text_input("終了時刻", value=target["end_time"] or "")
        with c3:
            current_user_label = ""
            if target["user_name"] and target["user_id"]:
                current_user_label = f"{target['user_name']}（ID:{target['user_id']}）"
            elif target["user_name"]:
                for label, pair in user_map.items():
                    if pair[1] == target["user_name"]:
                        current_user_label = label
                        break
            user_index = users.index(current_user_label) if current_user_label in users else 0
            staff_index = staff.index(target["staff_name"]) if target["staff_name"] in staff else 0
            new_user = st.selectbox("利用者", users, index=user_index)
            new_staff = st.selectbox("担当職員", staff, index=staff_index)

        new_title = st.text_input("予定タイトル", value=target["title"])
        new_memo = st.text_area("メモ", value=target["memo"] or "")
        new_important = st.checkbox("重要マーク", value=bool(target["important"]))

        c_update, c_delete = st.columns(2)
        with c_update:
            update_btn = st.form_submit_button("更新する")
        with c_delete:
            delete_btn = st.form_submit_button("削除する")

    if update_btn:
        if not new_title.strip():
            st.error("予定タイトルを入力してください。")
            return

        new_user_id, new_user_name = user_map.get(new_user, ("", ""))

        execute("""
            UPDATE events
            SET event_date=?, category=?, title=?, user_id=?, user_name=?, staff_name=?,
                start_time=?, end_time=?, memo=?, important=?, updated_at=?
            WHERE id=?
        """, (
            new_date.strftime("%Y-%m-%d"),
            new_category,
            new_title.strip(),
            new_user_id or None,
            new_user_name or None,
            new_staff or None,
            new_start_time.strip() or None,
            new_end_time.strip() or None,
            new_memo.strip() or None,
            1 if new_important else 0,
            now_text(),
            int(selected_id),
        ))
        st.success("予定を更新しました。画面を再読み込みしてください。")

    if delete_btn:
        photos = get_event_photos(selected_id)
        for _, p in photos.iterrows():
            delete_saved_file(p["file_path"])
        files = get_event_files(selected_id)
        for _, frow in files.iterrows():
            delete_saved_file(frow["file_path"])
        execute("DELETE FROM event_photos WHERE event_id=?", (int(selected_id),))
        execute("DELETE FROM event_files WHERE event_id=?", (int(selected_id),))
        execute("DELETE FROM events WHERE id=?", (int(selected_id),))
        st.warning("予定と紐づく写真メモ・Excelファイルを削除しました。画面を再読み込みしてください。")



def page_category_master():
    st.subheader("予定カテゴリ設定")
    st.caption("予定登録で使うカテゴリを追加・編集・非表示にできます。")

    with st.form("category_add_form", clear_on_submit=True):
        c1, c2, c3 = st.columns([2, 1, 1])
        with c1:
            category_name = st.text_input("カテゴリ名", placeholder="例：訪問診療、往診、買い物、家族連絡")
        with c2:
            mark = st.text_input("マーク", placeholder="例：🏥", value="・")
        with c3:
            sort_order = st.number_input("並び順", min_value=1, max_value=999, value=100, step=10)

        add = st.form_submit_button("カテゴリを追加")

    if add:
        if not category_name.strip():
            st.error("カテゴリ名を入力してください。")
        else:
            try:
                execute("""
                    INSERT INTO categories
                    (category_name, mark, sort_order, is_active, created_at, updated_at)
                    VALUES (?, ?, ?, 1, ?, ?)
                """, (
                    category_name.strip(),
                    mark.strip() or "・",
                    int(sort_order),
                    now_text(),
                    now_text(),
                ))
                st.success("カテゴリを追加しました。")
            except Exception as e:
                st.error(f"追加できませんでした。同じカテゴリ名がある可能性があります：{e}")

    st.markdown("---")
    df = fetch_df("""
        SELECT id, category_name, mark, sort_order, is_active, created_at, updated_at
        FROM categories
        ORDER BY sort_order, category_name
    """)

    if df.empty:
        st.info("カテゴリが登録されていません。")
        return

    st.dataframe(df, use_container_width=True, hide_index=True)

    selected_id = st.selectbox("編集するカテゴリID", df["id"].tolist())
    target = df[df["id"] == selected_id].iloc[0]

    with st.form("category_edit_form"):
        c1, c2, c3, c4 = st.columns([2, 1, 1, 1])
        with c1:
            new_name = st.text_input("カテゴリ名", value=target["category_name"])
        with c2:
            new_mark = st.text_input("マーク", value=target["mark"] or "・")
        with c3:
            new_sort = st.number_input("並び順", min_value=1, max_value=999, value=int(target["sort_order"] or 100), step=10)
        with c4:
            new_active = st.selectbox(
                "状態",
                [1, 0],
                index=0 if int(target["is_active"] or 1) == 1 else 1,
                format_func=lambda x: "表示" if x == 1 else "非表示"
            )

        c_update, c_delete = st.columns(2)
        with c_update:
            update = st.form_submit_button("更新する")
        with c_delete:
            delete = st.form_submit_button("削除する")

    if update:
        if not new_name.strip():
            st.error("カテゴリ名を入力してください。")
        else:
            try:
                execute("""
                    UPDATE categories
                    SET category_name=?, mark=?, sort_order=?, is_active=?, updated_at=?
                    WHERE id=?
                """, (
                    new_name.strip(),
                    new_mark.strip() or "・",
                    int(new_sort),
                    int(new_active),
                    now_text(),
                    int(selected_id),
                ))
                st.success("カテゴリを更新しました。画面を再読み込みしてください。")
            except Exception as e:
                st.error(f"更新できませんでした：{e}")

    if delete:
        # 既存予定で使われているカテゴリは削除ではなく非表示推奨
        used = fetch_df("SELECT COUNT(*) AS cnt FROM events WHERE category=?", (target["category_name"],))
        count = int(used.iloc[0]["cnt"]) if not used.empty else 0
        if count > 0:
            execute("UPDATE categories SET is_active=0, updated_at=? WHERE id=?", (now_text(), int(selected_id)))
            st.warning(f"このカテゴリは既存予定で {count} 件使われているため、削除せず非表示にしました。")
        else:
            execute("DELETE FROM categories WHERE id=?", (int(selected_id),))
            st.warning("カテゴリを削除しました。画面を再読み込みしてください。")


def page_master_users():
    st.subheader("利用者マスタ")

    with st.form("user_add_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            user_id = st.text_input("利用者ID", placeholder="例：U0001 / 健康管理アプリ側のID")
        with c2:
            user_name = st.text_input("利用者名")
        with c3:
            kana = st.text_input("ふりがな")
        room_no = st.text_input("居室番号", placeholder="例：101")
        note = st.text_area("メモ")
        add = st.form_submit_button("利用者を追加")

    if add:
        if not user_name.strip():
            st.error("利用者名を入力してください。")
        else:
            try:
                if not user_id.strip():
                    max_df = fetch_df("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM users")
                    user_id_value = f"U{int(max_df.iloc[0]['next_id']):04d}"
                else:
                    user_id_value = user_id.strip()

                execute("""
                    INSERT INTO users (user_id, user_name, kana, room_no, note, is_active, created_at)
                    VALUES (?, ?, ?, ?, ?, 1, ?)
                """, (user_id_value, user_name.strip(), kana.strip(), room_no.strip(), note.strip(), now_text()))
                st.success("利用者を追加しました。")
            except DB_INTEGRITY_ERROR:
                st.error("同じ利用者名がすでに登録されています。")

    df = fetch_df("SELECT * FROM users ORDER BY is_active DESC, user_name")
    st.dataframe(df, use_container_width=True, hide_index=True)

    if not df.empty:
        selected = st.selectbox("有効／無効を変更する利用者ID", df["id"].tolist())
        target = df[df["id"] == selected].iloc[0]
        new_active = st.radio("状態", [1, 0], index=0 if target["is_active"] == 1 else 1, format_func=lambda x: "有効" if x == 1 else "無効")
        if st.button("利用者状態を更新"):
            execute("UPDATE users SET is_active=? WHERE id=?", (int(new_active), int(selected)))
            st.success("状態を更新しました。")


def page_master_staff():
    st.subheader("職員マスタ")

    with st.form("staff_add_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        with c1:
            staff_name = st.text_input("職員名")
        with c2:
            role = st.text_input("役割", placeholder="例：管理者、日勤、夜勤")
        add = st.form_submit_button("職員を追加")

    if add:
        clean_staff_name = normalize_staff_name(staff_name)
        if not clean_staff_name:
            st.error("職員名を入力してください。")
        else:
            existing_staff = fetch_df("SELECT staff_name FROM staff")
            existing_norms = set()
            if existing_staff is not None and not existing_staff.empty:
                existing_norms = set(existing_staff["staff_name"].dropna().astype(str).apply(normalize_staff_name).tolist())
            if clean_staff_name in existing_norms:
                st.error("同じ職員名がすでに登録されています。（空白違いも同一職員として扱います）")
            else:
                try:
                    execute("""
                        INSERT INTO staff (staff_name, role, is_active, created_at)
                        VALUES (?, ?, 1, ?)
                    """, (clean_staff_name, role.strip(), now_text()))
                    st.success("職員を追加しました。")
                except DB_INTEGRITY_ERROR:
                    st.error("同じ職員名がすでに登録されています。")

    df = fetch_df("SELECT * FROM staff ORDER BY is_active DESC, staff_name")
    st.dataframe(df, use_container_width=True, hide_index=True)

    duplicate_df = get_duplicate_staff_name_groups()
    if not duplicate_df.empty:
        st.warning("職員マスタに空白違いの重複があります。シフト表では1人として表示します。不要な重複行は無効にしてください。")
        st.dataframe(duplicate_df, use_container_width=True, hide_index=True)

    if not df.empty:
        selected = st.selectbox("有効／無効を変更する職員ID", df["id"].tolist())
        target = df[df["id"] == selected].iloc[0]
        new_active = st.radio("状態", [1, 0], index=0 if target["is_active"] == 1 else 1, format_func=lambda x: "有効" if x == 1 else "無効")
        if st.button("職員状態を更新"):
            execute("UPDATE staff SET is_active=? WHERE id=?", (int(new_active), int(selected)))
            st.success("状態を更新しました。")



def page_photo_notes():
    st.subheader("写真メモ一覧")

    df = fetch_df("""
        SELECT
            p.id AS photo_id,
            p.file_name,
            p.file_path,
            p.photo_memo,
            p.created_at AS photo_created_at,
            e.id AS event_id,
            e.event_date,
            e.category,
            e.title,
            e.user_id,
            e.user_name,
            e.staff_name,
            e.memo
        FROM event_photos p
        JOIN events e ON p.event_id = e.id
        ORDER BY e.event_date DESC, p.id DESC
    """)

    if df.empty:
        st.info("写真メモはまだ登録されていません。写真アップロード時にStorage保存が失敗している場合、予定本体だけが保存され、写真メモは登録されません。")
        return

    keyword = st.text_input("写真メモ検索", placeholder="予定タイトル・利用者名・写真メモなど")
    if keyword.strip():
        kw = keyword.strip()
        mask = (
            df["title"].fillna("").str.contains(kw, case=False, na=False) |
            df["user_id"].fillna("").str.contains(kw, case=False, na=False) |
            df["user_name"].fillna("").str.contains(kw, case=False, na=False) |
            df["staff_name"].fillna("").str.contains(kw, case=False, na=False) |
            df["photo_memo"].fillna("").str.contains(kw, case=False, na=False) |
            df["memo"].fillna("").str.contains(kw, case=False, na=False)
        )
        df = df[mask]

    if df.empty:
        st.info("該当する写真メモはありません。")
        return

    for _, row in df.iterrows():
        st.markdown("---")
        c1, c2 = st.columns([1, 2])
        with c1:
            render_saved_image(
                row["file_path"],
                use_container_width=True,
            )
        with c2:
            st.write(f"**{row['event_date']}｜{row['category']}｜{row['title']}**")
            if row["user_name"]:
                st.write(f"利用者：{row['user_name']}（ID:{row['user_id'] or ''}）")
            if row["staff_name"]:
                st.write(f"担当：{row['staff_name']}")
            st.write(f"写真メモ：{row['photo_memo'] or ''}")
            if row["memo"]:
                st.caption(f"予定メモ：{row['memo']}")
            st.caption(f"写真登録：{row['photo_created_at']}")



def page_attached_files():
    st.subheader("Excel・書類ファイル一覧")

    df = fetch_df("""
        SELECT
            f.id AS file_id,
            f.file_name,
            f.file_path,
            f.file_type,
            f.file_memo,
            f.created_at AS file_created_at,
            e.id AS event_id,
            e.event_date,
            e.category,
            e.title,
            e.user_id,
            e.user_name,
            e.staff_name,
            e.memo
        FROM event_files f
        JOIN events e ON f.event_id = e.id
        ORDER BY e.event_date DESC, f.id DESC
    """)

    if df.empty:
        st.info("Excel・書類ファイルはまだ登録されていません。")
        return

    keyword = st.text_input("ファイル検索", placeholder="予定タイトル・利用者名・ファイル名・メモなど")
    if keyword.strip():
        kw = keyword.strip()
        mask = (
            df["title"].fillna("").str.contains(kw, case=False, na=False) |
            df["user_id"].fillna("").str.contains(kw, case=False, na=False) |
            df["user_name"].fillna("").str.contains(kw, case=False, na=False) |
            df["staff_name"].fillna("").str.contains(kw, case=False, na=False) |
            df["file_name"].fillna("").str.contains(kw, case=False, na=False) |
            df["file_memo"].fillna("").str.contains(kw, case=False, na=False) |
            df["memo"].fillna("").str.contains(kw, case=False, na=False)
        )
        df = df[mask]

    if df.empty:
        st.info("該当するファイルはありません。")
        return

    for _, row in df.iterrows():
        st.markdown("---")
        st.write(f"📎 **{row['file_name']}**")
        st.write(f"{row['event_date']}｜{row['category']}｜{row['title']}")
        if row["user_name"]:
            st.write(f"利用者：{row['user_name']}")
        if row["staff_name"]:
            st.write(f"担当：{row['staff_name']}")
        st.write(f"ファイルメモ：{row['file_memo'] or ''}")
        st.caption(f"登録：{row['file_created_at']}")

        render_saved_download_button(
            "ダウンロード",
            row["file_path"],
            row["file_name"],
            key=f"file_list_download_{row['file_id']}",
        )



# -----------------------------
# 予定データ取込（健康管理アプリ → ひだまり帳）
# -----------------------------
def normalize_import_bool(value, default=True):
    """Excel/CSV上のチェック値をTrue/Falseへ寄せる。"""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    s = str(value).strip().lower()
    if s in ["true", "1", "yes", "y", "登録", "登録する", "する", "〇", "○", "✓", "☑"]:
        return True
    if s in ["false", "0", "no", "n", "登録しない", "しない", "×", "✕", "☐"]:
        return False
    return default


def normalize_import_date(value):
    """日付をYYYY-MM-DD文字列へ正規化。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return ""
    # 2026-05-27 11:00 のような値は日付だけ使う
    dt = pd.to_datetime(text, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d")
    return text[:10]


def normalize_import_time(value):
    """時刻をHH:MMへ正規化。空なら空文字。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if not text or text.lower() in ["none", "nan", "nat"]:
        return ""
    # Excel由来で 11:00:00 になる場合
    if re.match(r"^\d{1,2}:\d{2}(:\d{2})?$", text):
        parts = text.split(":")
        return f"{int(parts[0]):02d}:{parts[1]}"
    # 0.5 のようなExcel時刻シリアル対策
    try:
        f = float(text)
        if 0 <= f < 1:
            minutes = int(round(f * 24 * 60))
            return f"{minutes//60:02d}:{minutes%60:02d}"
    except Exception:
        pass
    return text


def pick_import_col(df, candidates):
    """候補列名のうち、存在する最初の列名を返す。"""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def read_schedule_import_file(uploaded_file):
    """予定候補Excel/CSVを読み込む。複数シートExcelは先頭シートを読む。"""
    name = str(getattr(uploaded_file, "name", "")).lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file, encoding="utf-8-sig")
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding="cp932")
    return pd.read_excel(uploaded_file)


def clean_import_value(value):
    """nan / NaT / Noneを空文字に寄せる。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in ["nan", "nat", "none"]:
        return ""
    return text


def make_upload_key(uploaded_file):
    """ファイルが変わった時にdata_editorの古い状態が残らないようにする。"""
    name = str(getattr(uploaded_file, "name", "uploaded"))
    size = int(getattr(uploaded_file, "size", 0) or 0)
    return hashlib.md5(f"{name}-{size}".encode("utf-8")).hexdigest()[:10]


def normalize_schedule_import_df(raw_df):
    """
    健康管理アプリ側の出力列を、ひだまり帳events登録用に正規化する。
    元データとの対応が分かるように、元行番号を保持する。
    """
    base_cols = [
        "元行", "登録する", "event_date", "start_time", "end_time", "category", "title",
        "user_id", "user_name", "staff_name", "memo", "important", "取込状態"
    ]
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=base_cols)

    df = raw_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # 健康チェックアプリ側の列名ゆれを広めに吸収
    col_register = pick_import_col(df, ["登録する", "取込対象", "取込", "登録", "import", "selected", "All Day Event"])
    col_date = pick_import_col(df, ["event_date", "予定日", "日付", "開始日", "日時", "開始日時", "予定日時", "実施日", "Start Date", "start date", "DATE", "Date"])
    col_start = pick_import_col(df, ["start_time", "開始時刻", "開始時間", "時刻", "予定時刻", "時間", "Start Time", "start time", "START TIME"])
    col_end = pick_import_col(df, ["end_time", "終了時刻", "終了時間", "End Time", "end time", "END TIME"])
    col_category = pick_import_col(df, ["category", "分類", "カテゴリ", "予定分類", "種別", "キーワード", "Category", "CATEGORIES"])
    col_title = pick_import_col(df, ["title", "タイトル", "件名", "予定タイトル", "予定", "候補", "予定候補", "Subject", "subject", "SUMMARY", "Summary"])
    col_user_id = pick_import_col(df, ["user_id", "利用者ID", "利用者id", "入居者ID", "入居者id"])
    col_user_name = pick_import_col(df, ["user_name", "利用者名", "利用者", "入居者名", "入居者", "対象者"])
    col_staff = pick_import_col(df, ["staff_name", "担当", "担当者", "職員", "記入者", "作成者"])
    col_memo = pick_import_col(df, ["memo", "詳細", "メモ", "内容", "備考", "申し送り", "申し送り内容", "本文", "元の申し送り", "Description", "description", "Location", "location"])
    col_important = pick_import_col(df, ["important", "重要", "重要マーク", "注意", "要注意"])

    rows = []
    for original_index, r in df.iterrows():
        raw_date_value = r.get(col_date, "") if col_date else ""
        event_date = normalize_import_date(raw_date_value)

        start_time = normalize_import_time(r.get(col_start, "")) if col_start else ""
        end_time = normalize_import_time(r.get(col_end, "")) if col_end else ""

        # 日時列に時刻が含まれる場合は、開始時刻にも反映
        if not start_time and col_date:
            dt = pd.to_datetime(raw_date_value, errors="coerce")
            if pd.notna(dt) and (dt.hour != 0 or dt.minute != 0):
                start_time = dt.strftime("%H:%M")

        category = clean_import_value(r.get(col_category, "")) if col_category else ""
        title = clean_import_value(r.get(col_title, "")) if col_title else ""
        user_id = clean_import_value(r.get(col_user_id, "")) if col_user_id else ""
        user_name = clean_import_value(r.get(col_user_name, "")) if col_user_name else ""
        staff_name = clean_import_value(r.get(col_staff, "")) if col_staff else ""
        memo = clean_import_value(r.get(col_memo, "")) if col_memo else ""

        # タイトルが空なら、申し送り本文やカテゴリから予定名を補う
        if not category:
            category = "その他"
        if not title:
            if memo:
                title = memo.splitlines()[0][:30]
            else:
                title = category or "予定"

        important = 1 if normalize_import_bool(r.get(col_important, False), default=False) else 0
        register = normalize_import_bool(r.get(col_register, True), default=True) if col_register else True

        status = "OK"
        if not event_date:
            status = "日付なし"
        elif not title:
            status = "タイトルなし"

        rows.append({
            "元行": int(original_index) + 2,  # Excel上の見た目に近い行番号。1行目は見出し想定
            "登録する": bool(register),
            "event_date": event_date,
            "start_time": start_time,
            "end_time": end_time,
            "category": category,
            "title": title,
            "user_id": user_id,
            "user_name": user_name,
            "staff_name": staff_name,
            "memo": memo,
            "important": int(important),
            "取込状態": status,
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out

    keep_cols = ["event_date", "title", "memo", "user_id", "user_name"]
    out = out[out[keep_cols].astype(str).agg("".join, axis=1).str.strip() != ""].copy()
    return out.reset_index(drop=True)


def make_import_compare_df(raw_df, import_df):
    """元データと変換後予定候補を横並びで確認する表を作る。"""
    if raw_df is None or raw_df.empty or import_df is None or import_df.empty:
        return pd.DataFrame()
    raw = raw_df.copy().reset_index(drop=True)
    raw.insert(0, "元行", raw.index + 2)
    preview_cols = ["元行"] + [c for c in raw.columns if c != "元行"][:8]
    converted_cols = [
        "元行", "event_date", "start_time", "category", "title",
        "user_id", "user_name", "memo", "取込状態", "既存重複"
    ]
    left = raw[preview_cols]
    right = import_df[[c for c in converted_cols if c in import_df.columns]]
    return pd.merge(left, right, on="元行", how="outer", suffixes=("_元", "_変換後"))


def event_exists(row):
    """同一予定らしいものが既にあるか簡易チェック。"""
    df = fetch_df("""
        SELECT id FROM events
        WHERE event_date=?
          AND COALESCE(start_time, '')=?
          AND title=?
          AND COALESCE(user_id, '')=?
          AND COALESCE(user_name, '')=?
        LIMIT 1
    """, (
        row.get("event_date", ""),
        row.get("start_time", "") or "",
        row.get("title", ""),
        row.get("user_id", "") or "",
        row.get("user_name", "") or "",
    ))
    return not df.empty


def ensure_import_category(category_name):
    """取込データのカテゴリがマスタになければ追加する。"""
    category_name = str(category_name or "その他").strip() or "その他"
    hit = fetch_df("SELECT id FROM categories WHERE category_name=? LIMIT 1", (category_name,))
    if not hit.empty:
        return
    mark = CATEGORY_MARK.get(category_name, "・")
    try:
        execute("""
            INSERT INTO categories
            (category_name, mark, sort_order, is_active, created_at, updated_at)
            VALUES (?, ?, 900, 1, ?, ?)
        """, (category_name, mark, now_text(), now_text()))
    except Exception:
        pass


def upsert_import_user(user_id, user_name):
    """user_id付き利用者が未登録なら、ひだまり帳の利用者マスタにも補助登録する。"""
    user_id = str(user_id or "").strip()
    user_name = str(user_name or "").strip()
    if not user_id or not user_name:
        return
    hit = fetch_df("SELECT id FROM users WHERE user_id=? OR user_name=? LIMIT 1", (user_id, user_name))
    if not hit.empty:
        return
    try:
        execute("""
            INSERT INTO users (user_id, user_name, kana, room_no, note, is_active, created_at)
            VALUES (?, ?, '', '', '予定データ取込から自動追加', 1, ?)
        """, (user_id, user_name, now_text()))
    except Exception:
        pass


def insert_import_event(row):
    """正規化済み1行をeventsへ登録。"""
    ensure_import_category(row.get("category", "その他"))
    upsert_import_user(row.get("user_id", ""), row.get("user_name", ""))
    return execute("""
        INSERT INTO events
        (event_date, category, title, user_id, user_name, staff_name, start_time, end_time, memo, important, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row.get("event_date", ""),
        row.get("category", "その他") or "その他",
        row.get("title", "予定") or "予定",
        row.get("user_id", "") or None,
        row.get("user_name", "") or None,
        row.get("staff_name", "") or None,
        row.get("start_time", "") or None,
        row.get("end_time", "") or None,
        row.get("memo", "") or None,
        int(row.get("important", 0) or 0),
        now_text(),
        now_text(),
    ))


def page_schedule_import():
    st.subheader("予定データ取込")
    st.caption("健康管理アプリで出力した予定候補Excel/CSVを読み込み、内容確認後にひだまり帳の予定へ登録します。")

    uploaded = st.file_uploader(
        "予定候補Excel/CSVを選択",
        type=["xlsx", "xls", "csv"],
        help="健康管理アプリの『予定候補一覧をExcelでダウンロード』『CSV出力』で作成したファイルを読み込めます。"
    )

    if not uploaded:
        st.info("ファイルを選択すると、予定候補を一覧表示します。")
        st.markdown("""
        **推奨列名**  
        `event_date / start_time / end_time / category / title / user_id / user_name / memo / important`  
        日本語列名（予定日、開始時刻、終了時刻、分類、タイトル、利用者ID、利用者名、詳細、重要）でも読み込めます。
        """)
        return

    try:
        raw_df = read_schedule_import_file(uploaded)
    except Exception as e:
        st.error(f"ファイルを読み込めませんでした：{e}")
        return

    import_df = normalize_schedule_import_df(raw_df)
    if import_df.empty:
        st.warning("取込できる予定候補がありません。")
        with st.expander("読み込んだ元データを確認"):
            st.dataframe(raw_df, use_container_width=True, hide_index=True)
        return

    # 重複チェック列を追加
    import_df["既存重複"] = import_df.apply(lambda r: "あり" if event_exists(r) else "", axis=1)
    # 重複ありは初期状態で登録対象から外す
    import_df.loc[import_df["既存重複"] == "あり", "登録する"] = False

    st.success(f"予定候補を {len(import_df)} 件読み込みました。内容を確認してから登録してください。")
    st.caption("上の件数は、元データを予定登録用に変換した後の件数です。空行や日付なし行は除外されます。")

    upload_key = make_upload_key(uploaded)
    editor_key = f"schedule_import_editor_{upload_key}"

    edited = st.data_editor(
        import_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "元行": st.column_config.NumberColumn("元行", disabled=True),
            "登録する": st.column_config.CheckboxColumn("登録する"),
            "event_date": st.column_config.TextColumn("予定日", help="YYYY-MM-DD"),
            "start_time": st.column_config.TextColumn("開始時刻", help="例：10:00"),
            "end_time": st.column_config.TextColumn("終了時刻", help="例：11:00"),
            "category": st.column_config.TextColumn("分類"),
            "title": st.column_config.TextColumn("タイトル"),
            "user_id": st.column_config.TextColumn("利用者ID"),
            "user_name": st.column_config.TextColumn("利用者名"),
            "staff_name": st.column_config.TextColumn("担当"),
            "memo": st.column_config.TextColumn("詳細"),
            "important": st.column_config.CheckboxColumn("重要"),
            "取込状態": st.column_config.TextColumn("状態", disabled=True),
            "既存重複": st.column_config.TextColumn("既存重複", disabled=True),
        },
        key=editor_key,
    )

    selected = edited[edited["登録する"].astype(bool)].copy()
    # 空文字だけでなく、nan / NaT / None 文字列も登録不可として扱う
    date_ok = ~selected["event_date"].astype(str).str.strip().str.lower().isin(["", "nan", "nat", "none"])
    title_ok = ~selected["title"].astype(str).str.strip().str.lower().isin(["", "nan", "nat", "none"])
    valid = selected[date_ok & title_ok].copy()

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("登録対象", len(selected))
    with c2:
        st.metric("登録可能", len(valid))
    with c3:
        st.metric("重複候補", int((edited["既存重複"].astype(str) == "あり").sum()))

    with st.expander("元データと予定候補の対応を確認"):
        compare_df = make_import_compare_df(raw_df, import_df)
        if compare_df.empty:
            st.info("比較できるデータがありません。")
        else:
            st.dataframe(compare_df, use_container_width=True, hide_index=True)
        st.caption("左側が読み込んだ元データ、右側が予定候補へ変換した内容です。元行番号で対応を確認できます。")

    with st.expander("読み込んだ元データだけを確認"):
        st.dataframe(raw_df, use_container_width=True, hide_index=True)

    skip_duplicates = st.checkbox("既存重複ありの行は登録しない", value=True)
    confirm = st.checkbox("内容を確認しました。eventsテーブルへ登録します。")

    if st.button("選択した予定をひだまり帳へ登録", type="primary", use_container_width=True):
        if not confirm:
            st.error("確認チェックを入れてください。")
            return
        if valid.empty:
            st.error("登録可能な行がありません。予定日とタイトルを確認してください。")
            return

        inserted = 0
        skipped = 0
        errors = []
        for idx, row in valid.iterrows():
            try:
                if skip_duplicates and event_exists(row):
                    skipped += 1
                    continue
                insert_import_event(row)
                inserted += 1
            except Exception as e:
                row_no = row.get("元行", idx + 1)
                errors.append(f"元行{row_no}: {e}")

        if inserted:
            st.success(f"ひだまり帳へ {inserted} 件登録しました。")
        if skipped:
            st.info(f"重複候補のため {skipped} 件スキップしました。")
        if errors:
            st.error("一部登録できませんでした。")
            for err in errors[:10]:
                st.write(err)

# -----------------------------
# PDFカレンダー出力
# -----------------------------
PDF_FONT_GOTHIC = "HeiseiKakuGo-W5"
PDF_FONT_MINCHO = "HeiseiMin-W3"


def init_pdf_fonts():
    """ReportLabの日本語CIDフォントを登録する。"""
    if not REPORTLAB_AVAILABLE:
        return False
    try:
        if PDF_FONT_GOTHIC not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT_GOTHIC))
        if PDF_FONT_MINCHO not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT_MINCHO))
        return True
    except Exception:
        return False


def pdf_text(value):
    """PDF描画用にNone/nanを空文字へ寄せる。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def text_width(text, font_name, font_size):
    """ReportLab上の文字幅を取得する。"""
    try:
        return pdfmetrics.stringWidth(str(text), font_name, font_size)
    except Exception:
        return len(str(text)) * font_size


def fit_pdf_text(text, max_width, font_name=PDF_FONT_GOTHIC, font_size=8, ellipsis="…"):
    """
    指定幅に収まるよう、1行表示用の文字列へ整える。
    改行・連続空白を1つの空白へ寄せ、長い場合は末尾を…で省略する。
    """
    s = re.sub(r"\s+", " ", pdf_text(text)).strip()
    if not s:
        return ""
    if text_width(s, font_name, font_size) <= max_width:
        return s

    # 省略記号だけでも幅を超える場合
    if text_width(ellipsis, font_name, font_size) > max_width:
        return ""

    low, high = 0, len(s)
    best = ""
    while low <= high:
        mid = (low + high) // 2
        candidate = s[:mid] + ellipsis
        if text_width(candidate, font_name, font_size) <= max_width:
            best = candidate
            low = mid + 1
        else:
            high = mid - 1
    return best or ellipsis


def wrap_pdf_text(text, max_width, font_name, font_size, max_lines=3):
    """
    日本語を含む文字列を、文字単位で指定幅に折り返す。
    長すぎる場合は最終行を…で省略する。
    """
    text = pdf_text(text).replace("\r", "\n")
    if not text:
        return []

    lines = []
    for raw in text.splitlines():
        raw = raw.strip()
        if not raw:
            continue
        current = ""
        for ch in raw:
            candidate = current + ch
            if text_width(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = ch
                if len(lines) >= max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
        if len(lines) >= max_lines:
            break

    if len(lines) > max_lines:
        lines = lines[:max_lines]

    if lines:
        original_joined = "".join([ln for ln in text.splitlines() if ln.strip()])
        wrapped_joined = "".join(lines)
        if len(wrapped_joined) < len(original_joined):
            last = lines[-1]
            while last and text_width(last + "…", font_name, font_size) > max_width:
                last = last[:-1]
            lines[-1] = (last + "…") if last else "…"

    return lines


def draw_wrapped_text(c, text, x, y, max_width, font_name, font_size, leading, max_lines=3, color=None):
    """折り返しテキストを描画し、描画後のy座標を返す。"""
    if color is not None:
        c.setFillColor(color)
    c.setFont(font_name, font_size)
    lines = wrap_pdf_text(text, max_width, font_name, font_size, max_lines=max_lines)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    c.setFillColor(colors.black)
    return y


def draw_single_line(c, text, x, y, max_width, font_name=PDF_FONT_GOTHIC, font_size=8, color=None):
    """
    PDF上に1行だけ描画する。
    幅を超える文字列は自動で…省略するため、次の項目と重ならない。
    """
    if color is not None:
        c.setFillColor(color)
    c.setFont(font_name, font_size)
    c.drawString(x, y, fit_pdf_text(text, max_width, font_name, font_size))
    c.setFillColor(colors.black)


def calendar_events_df(year, month):
    """指定月の予定をDataFrameで取得する。"""
    start = f"{int(year)}-{int(month):02d}-01"
    last_day = calendar.monthrange(int(year), int(month))[1]
    end = f"{int(year)}-{int(month):02d}-{last_day:02d}"
    return fetch_df("""
        SELECT *
        FROM events
        WHERE event_date BETWEEN ? AND ?
        ORDER BY event_date,
            CASE WHEN start_time IS NULL OR start_time = '' THEN 1 ELSE 0 END,
            start_time,
            category,
            id
    """, (start, end))


def make_event_summary_for_pdf(ev, compact=True):
    """カレンダー枠内に表示する予定の短い説明を作る。"""
    category_name = re.sub(r"\s+", " ", pdf_text(ev.get("category", ""))).strip()
    time_part = re.sub(r"\s+", " ", pdf_text(ev.get("start_time", ""))).strip()
    title = re.sub(r"\s+", " ", pdf_text(ev.get("title", ""))).strip()
    user_name = re.sub(r"\s+", " ", pdf_text(ev.get("user_name", ""))).strip()
    staff_name = re.sub(r"\s+", " ", pdf_text(ev.get("staff_name", ""))).strip()
    important = "★" if int(ev.get("important", 0) or 0) == 1 else ""

    if compact:
        # 月間カレンダー枠内は重なり防止を最優先し、情報量を絞る。
        # 利用者・担当・長い詳細は次ページの予定詳細一覧で確認する。
        parts = []
        if time_part:
            parts.append(time_part)

        label_core = f"【{category_name}】" if category_name else ""
        if title:
            label_core += title
        elif user_name:
            label_core += user_name

        if important:
            label_core = f"{important}{label_core}"

        if label_core:
            parts.append(label_core)

        return " ".join([p for p in parts if p]).strip()

    parts = []
    if time_part:
        parts.append(time_part)
    label_core = f"【{category_name}】"
    if important:
        label_core += important
    if title:
        label_core += title
    parts.append(label_core)
    if user_name:
        parts.append(f"利用者:{user_name}")
    if staff_name:
        parts.append(f"担当:{staff_name}")
    return " ".join([p for p in parts if p]).strip()


def draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=False):
    """予定詳細一覧ページのヘッダーを描画する。"""
    suffix = "（続き）" if continuation else ""
    c.setFont(PDF_FONT_GOTHIC, 16)
    c.setFillColor(colors.HexColor("#333333"))
    c.drawString(margin_x, height - 34, f"予定詳細一覧　{year}年 {month}月{suffix}")
    c.setFont(PDF_FONT_GOTHIC, 9)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawRightString(width - margin_x, height - 32, f"合計 {total_count} 件")
    c.setFillColor(colors.black)


def draw_pdf_detail_table_header(c, x, y, widths, row_h):
    """予定詳細一覧の列見出しを描画する。"""
    headers = ["時間", "分類", "予定", "利用者", "担当", "詳細・メモ"]
    c.setFillColor(colors.HexColor("#eee7dc"))
    c.rect(x, y - 4, sum(widths), row_h, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#333333"))
    c.setFont(PDF_FONT_GOTHIC, 8)
    cur_x = x
    for header, w in zip(headers, widths):
        c.drawString(cur_x + 4, y + 1, header)
        cur_x += w
    c.setFillColor(colors.black)
    return y - row_h


def make_calendar_pdf(year, month, include_detail=True):
    """
    A4横の月間カレンダーPDFを作る。
    1ページ目: 月間カレンダー
    2ページ目以降: 予定詳細一覧

    Ver1.3.3相当の修正:
    ・予定詳細一覧を1予定1行の表形式へ変更
    ・詳細・メモも1行に収め、長い場合は…で省略
    ・カレンダー枠内の予定文字が重ならないよう、表示行数・文字量・余白を再調整
    ・月間カレンダー枠内は要約表示に絞り、あふれる予定は「ほか◯件」で表示
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab がインストールされていません。requirements.txt に reportlab を追加してください。")

    init_pdf_fonts()

    year = int(year)
    month = int(month)
    df = calendar_events_df(year, month)

    events_by_day = {}
    if not df.empty:
        for _, row in df.iterrows():
            events_by_day.setdefault(pdf_text(row["event_date"]), []).append(row.to_dict())

    buffer = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    margin_x = 24
    margin_top = 28
    title_h = 40
    grid_x = margin_x
    grid_y_top = height - margin_top - title_h
    grid_w = width - margin_x * 2
    grid_h = height - margin_top - title_h - 24
    day_header_h = 22
    cell_w = grid_w / 7
    cell_h = (grid_h - day_header_h) / 6

    # タイトル
    c.setFont(PDF_FONT_GOTHIC, 18)
    c.setFillColor(colors.HexColor("#333333"))
    c.drawString(margin_x, height - 34, f"ひだまり帳 月間カレンダー　{year}年 {month}月")
    c.setFont(PDF_FONT_GOTHIC, 9)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawRightString(width - margin_x, height - 32, f"出力日: {today_jst().strftime('%Y-%m-%d')}")
    c.setFillColor(colors.black)

    # 曜日ヘッダ
    week_labels = ["日", "月", "火", "水", "木", "金", "土"]
    header_y = grid_y_top - day_header_h
    for i, label in enumerate(week_labels):
        x = grid_x + i * cell_w
        c.setFillColor(colors.HexColor("#f3eee6"))
        c.rect(x, header_y, cell_w, day_header_h, fill=1, stroke=1)
        c.setFillColor(colors.HexColor("#333333"))
        if i == 0:
            c.setFillColor(colors.HexColor("#c0392b"))
        elif i == 6:
            c.setFillColor(colors.HexColor("#1f4e79"))
        c.setFont(PDF_FONT_GOTHIC, 10)
        c.drawCentredString(x + cell_w / 2, header_y + 7, label)

    # 月間グリッド
    weeks = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
    while len(weeks) < 6:
        weeks.append([0, 0, 0, 0, 0, 0, 0])

    for r, week in enumerate(weeks[:6]):
        for col, day in enumerate(week):
            x = grid_x + col * cell_w
            y = header_y - (r + 1) * cell_h

            c.setFillColor(colors.HexColor("#fffdf8") if day else colors.HexColor("#fafafa"))
            c.rect(x, y, cell_w, cell_h, fill=1, stroke=1)

            if not day:
                continue

            # 日付を少し小さくし、予定表示エリアを広げる
            date_color = colors.HexColor("#333333")
            if col == 0:
                date_color = colors.HexColor("#c0392b")
            elif col == 6:
                date_color = colors.HexColor("#1f4e79")
            c.setFillColor(date_color)
            c.setFont(PDF_FONT_GOTHIC, 15)
            c.drawString(x + 6, y + cell_h - 18, str(day))

            key = f"{year}-{month:02d}-{day:02d}"
            evs = events_by_day.get(key, [])
            if not evs:
                continue

            # 枠内で文字が重ならないよう、予定表示エリアを固定管理する
            left_pad = 5
            right_pad = 5
            top_after_date = 28
            bottom_reserved = 10
            band_h = 10
            row_gap = 2
            row_pitch = band_h + row_gap
            event_area_top = y + cell_h - top_after_date
            event_area_bottom = y + bottom_reserved
            usable_h = max(0, event_area_top - event_area_bottom)
            max_lines_total = max(1, int(usable_h // row_pitch))
            max_lines_total = min(max_lines_total, 4)

            visible_events = evs[:max_lines_total]
            for idx, ev in enumerate(visible_events):
                summary = make_event_summary_for_pdf(ev, compact=True)
                important = int(ev.get("important", 0) or 0) == 1

                band_top = event_area_top - idx * row_pitch
                band_y = band_top - band_h
                band_w = cell_w - left_pad - right_pad

                c.setFillColor(colors.HexColor("#ffe9e0") if important else colors.HexColor("#f6efe6"))
                c.roundRect(x + left_pad, band_y, band_w, band_h, 2.5, fill=1, stroke=0)

                text_color = colors.HexColor("#8a2d18") if important else colors.HexColor("#3f3a35")
                draw_single_line(
                    c,
                    summary,
                    x + left_pad + 2,
                    band_y + 2.1,
                    band_w - 4,
                    PDF_FONT_GOTHIC,
                    6.0,
                    color=text_color,
                )

            remaining = len(evs) - len(visible_events)
            if remaining > 0:
                c.setFont(PDF_FONT_GOTHIC, 6.3)
                c.setFillColor(colors.HexColor("#555555"))
                c.drawString(x + 7, y + 5, f"ほか {remaining} 件")
                c.setFillColor(colors.black)

    # 凡例
    c.setFont(PDF_FONT_GOTHIC, 8)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawString(margin_x, 12, "※枠内は要約表示です。詳細は次ページ以降の一覧で確認できます。重要予定は薄赤で表示します。")
    c.setFillColor(colors.black)

    # 詳細一覧
    if include_detail:
        c.showPage()
        c.setPageSize(page_size)

        total_count = 0 if df.empty else len(df)
        draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=False)

        detail_x = margin_x
        detail_w = width - margin_x * 2
        row_h = 17
        date_h = 18
        bottom_y = 34

        # 横幅に収まる固定列幅。最後の「詳細・メモ」は残り幅を使う。
        col_widths = [58, 54, 160, 78, 78]
        memo_w = detail_w - sum(col_widths)
        col_widths.append(memo_w)

        y = height - 62
        y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)

        if df.empty:
            c.setFont(PDF_FONT_GOTHIC, 12)
            c.setFillColor(colors.HexColor("#333333"))
            c.drawString(detail_x, y - 8, "この月の予定はありません。")
        else:
            current_date = ""
            row_index = 0

            for _, row in df.iterrows():
                event_date = pdf_text(row["event_date"])

                # 日付見出し＋1行分が入らなければ改ページ
                if y < bottom_y + date_h + row_h + 8:
                    c.showPage()
                    c.setPageSize(page_size)
                    draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=True)
                    y = height - 62
                    y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)
                    current_date = ""

                if event_date != current_date:
                    current_date = event_date
                    try:
                        d = datetime.strptime(event_date, "%Y-%m-%d").date()
                        dow = ["月", "火", "水", "木", "金", "土", "日"][d.weekday()]
                        date_label = f"{d.month}/{d.day}（{dow}）"
                    except Exception:
                        date_label = event_date

                    c.setFillColor(colors.HexColor("#f7f1e8"))
                    c.rect(detail_x, y - 3, detail_w, date_h, fill=1, stroke=0)
                    c.setFillColor(colors.HexColor("#333333"))
                    c.setFont(PDF_FONT_GOTHIC, 9)
                    c.drawString(detail_x + 6, y + 2, date_label)
                    c.setFillColor(colors.black)
                    y -= date_h

                # 1行分が入らなければ改ページ
                if y < bottom_y + row_h:
                    c.showPage()
                    c.setPageSize(page_size)
                    draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=True)
                    y = height - 62
                    y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)

                important = int(row.get("important", 0) or 0) == 1
                category_name = pdf_text(row.get("category", ""))

                start_time = pdf_text(row.get("start_time", ""))
                end_time = pdf_text(row.get("end_time", ""))
                if start_time and end_time:
                    time_text = f"{start_time}〜{end_time}"
                elif start_time:
                    time_text = start_time
                elif end_time:
                    time_text = f"〜{end_time}"
                else:
                    time_text = "時間未定"

                title_text = pdf_text(row.get("title", ""))
                user_text = pdf_text(row.get("user_name", ""))
                staff_text = pdf_text(row.get("staff_name", ""))
                memo_text = pdf_text(row.get("memo", ""))
                if important:
                    title_text = "重要 " + title_text

                # 行背景
                bg = "#fffdf8" if row_index % 2 == 0 else "#faf6ef"
                if important:
                    bg = "#fff0e8"
                c.setFillColor(colors.HexColor(bg))
                c.rect(detail_x, y - 3, detail_w, row_h, fill=1, stroke=0)

                # 下線
                c.setStrokeColor(colors.HexColor("#e2d8cc"))
                c.line(detail_x, y - 3, detail_x + detail_w, y - 3)
                c.setStrokeColor(colors.black)

                # 各列を1行で描画。幅を超えたら…で省略する。
                values = [time_text, category_name, title_text, user_text, staff_text, memo_text]
                cur_x = detail_x
                text_color = colors.HexColor("#8a2d18") if important else colors.HexColor("#222222")
                for value, w in zip(values, col_widths):
                    draw_single_line(
                        c, value, cur_x + 4, y + 1, max(w - 8, 10),
                        PDF_FONT_GOTHIC, 8, color=text_color
                    )
                    cur_x += w

                y -= row_h
                row_index += 1

    c.save()
    buffer.seek(0)
    return buffer.getvalue()





# -----------------------------
# シフト管理・AI担当割当
# -----------------------------
SHIFT_KINDS = ["日勤", "夜勤", "夜勤明け", "休み", "希望休", "有休", "その他"]
SHIFT_EDITOR_OPTIONS = ["", "日", "夜", "明", "希", "有", "他"]


def default_shift_times(shift_kind):
    """基本勤務時間。日勤2名、夜勤1名の運用を想定。"""
    if shift_kind == "日勤":
        return "08:30", "17:30", 0
    if shift_kind == "夜勤":
        return "16:30", "09:30", 1
    if shift_kind == "夜勤明け":
        return "", "", 0
    return "", "", 0


def shift_short_label(shift_kind):
    return {
        "日勤": "日",
        "夜勤": "夜",
        "夜勤明け": "明",
        "休み": "",
        "希望休": "希",
        "有休": "有",
        "その他": "他",
    }.get(str(shift_kind), str(shift_kind)[:1])


def shift_kind_from_editor_label(label):
    """月間シフト直接入力用の短縮ラベルを勤務区分へ変換する。"""
    value = str(label or "").strip()
    mapping = {
        "日": ["日勤"],
        "夜": ["夜勤"],
        "明": ["夜勤明け"],
        "希": ["希望休"],
        "有": ["有休"],
        "他": ["その他"],
    }
    return mapping.get(value, [])


def get_shift_month_status(year, month):
    """指定月のシフト確定状態を取得する。"""
    try:
        df = fetch_df("""
            SELECT *
            FROM shift_month_status
            WHERE shift_year=? AND shift_month=?
            LIMIT 1
        """, (int(year), int(month)))
        if df.empty:
            return {"is_confirmed": 0, "confirmed_at": "", "confirmed_by": ""}
        r = df.iloc[0]
        return {
            "is_confirmed": int(r.get("is_confirmed", 0) or 0),
            "confirmed_at": str(r.get("confirmed_at") or ""),
            "confirmed_by": str(r.get("confirmed_by") or ""),
        }
    except Exception:
        return {"is_confirmed": 0, "confirmed_at": "", "confirmed_by": ""}


def set_shift_month_status(year, month, is_confirmed, confirmed_by=""):
    """指定月のシフト確定／未確定を保存する。"""
    try:
        execute("""
            DELETE FROM shift_month_status
            WHERE shift_year=? AND shift_month=?
        """, (int(year), int(month)))
        execute("""
            INSERT INTO shift_month_status
            (shift_year, shift_month, is_confirmed, confirmed_at, confirmed_by, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            int(year),
            int(month),
            1 if is_confirmed else 0,
            now_text() if is_confirmed else None,
            confirmed_by or "",
            now_text(),
            now_text(),
        ))
    except Exception as e:
        st.warning(f"シフト確定状態の保存に失敗しました：{e}")

def get_staff_shift_limits():
    """
    職員別の月間勤務回数上限を取得する。
    未設定職員は、日勤31回・夜勤31回・合計31回として扱う。
    """
    staff_names = get_active_staff()
    try:
        df = fetch_df("""
            SELECT staff_name, max_day_shifts, max_night_shifts, max_total_shifts, note
            FROM staff_shift_limits
            ORDER BY staff_name
        """)
    except Exception:
        df = pd.DataFrame(columns=["staff_name", "max_day_shifts", "max_night_shifts", "max_total_shifts", "note"])

    rows = []
    for staff_name in staff_names:
        hit = df[df["staff_name"].astype(str) == str(staff_name)] if not df.empty else pd.DataFrame()
        if hit.empty:
            rows.append({
                "職員名": staff_name,
                "日勤上限": 31,
                "夜勤上限": 31,
                "合計上限": 31,
                "メモ": "",
            })
        else:
            r = hit.iloc[0]
            rows.append({
                "職員名": staff_name,
                "日勤上限": safe_shift_limit_value(r.get("max_day_shifts", 31), 31),
                "夜勤上限": safe_shift_limit_value(r.get("max_night_shifts", 31), 31),
                "合計上限": safe_shift_limit_value(r.get("max_total_shifts", 31), 31),
                "メモ": str(r.get("note") or ""),
            })
    return pd.DataFrame(rows)


def save_staff_shift_limits_from_editor(limits_df):
    """職員別勤務上限を保存する。"""
    if limits_df is None or limits_df.empty:
        return 0
    execute("DELETE FROM staff_shift_limits")
    params = []
    seen_staff_names = set()
    for _, r in limits_df.iterrows():
        staff_name = normalize_staff_name(r.get("職員名", ""))
        if not staff_name or staff_name in seen_staff_names:
            continue
        seen_staff_names.add(staff_name)
        params.append((
            staff_name,
            safe_shift_limit_value(r.get("日勤上限", 31), 31),
            safe_shift_limit_value(r.get("夜勤上限", 31), 31),
            safe_shift_limit_value(r.get("合計上限", 31), 31),
            str(r.get("メモ", "") or "").strip() or None,
            now_text(),
            now_text(),
        ))
    if not params:
        return 0
    return execute_many("""
        INSERT INTO staff_shift_limits
        (staff_name, max_day_shifts, max_night_shifts, max_total_shifts, note, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, params)




def normalize_staff_name(value):
    """
    職員名を比較・保存用に正規化する。
    「職員 A」「職員　A」「職員A」のような空白違いを同一職員として扱う。
    シフト表・上限表・AI案ではこの正規化名を使い、重複行を防ぐ。
    """
    if value is None:
        return ""
    v = str(value).replace("\u3000", " ").strip()
    v = re.sub(r"\s+", "", v)
    return v


def get_duplicate_staff_name_groups():
    """職員マスタ内の空白違い重複を確認するための一覧を作る。"""
    try:
        df = fetch_df("SELECT id, staff_name, role, is_active, created_at FROM staff ORDER BY is_active DESC, staff_name")
    except Exception:
        return pd.DataFrame(columns=["正規化名", "登録ID", "登録名", "状態"])
    if df is None or df.empty:
        return pd.DataFrame(columns=["正規化名", "登録ID", "登録名", "状態"])

    tmp = df.copy()
    tmp["正規化名"] = tmp["staff_name"].apply(normalize_staff_name)
    dup_names = tmp[tmp["正規化名"].duplicated(keep=False)]["正規化名"].unique().tolist()
    if not dup_names:
        return pd.DataFrame(columns=["正規化名", "登録ID", "登録名", "状態"])

    rows = []
    for norm in dup_names:
        part = tmp[tmp["正規化名"] == norm]
        for _, r in part.iterrows():
            rows.append({
                "正規化名": norm,
                "登録ID": int(r.get("id")),
                "登録名": str(r.get("staff_name") or ""),
                "状態": "有効" if int(r.get("is_active") or 0) == 1 else "無効",
            })
    return pd.DataFrame(rows, columns=["正規化名", "登録ID", "登録名", "状態"])


def safe_shift_limit_value(value, default=31):
    """
    職員別勤務回数上限を安全に数値化する。
    重要：0回指定は「勤務不可」として有効な設定なので、0を31へ戻さない。
    空欄・NaN・不正値だけを初期値31として扱う。
    """
    try:
        if value is None:
            return int(default)
        if isinstance(value, str) and value.strip() == "":
            return int(default)
        if pd.isna(value):
            return int(default)
        v = int(float(value))
        return max(0, min(31, v))
    except Exception:
        return int(default)


def normalize_shift_limits_for_staff(limit_map, staff_name):
    """limit_mapから職員の上限を取り出し、day/night/totalを必ず安全なintにそろえる。"""
    limit_map = limit_map or {}
    raw = limit_map.get(normalize_staff_name(staff_name), {"day": 31, "night": 31, "total": 31})
    day_limit = safe_shift_limit_value(raw.get("day", 31), 31)
    night_limit = safe_shift_limit_value(raw.get("night", 31), 31)
    total_limit = safe_shift_limit_value(raw.get("total", 31), 31)
    return {"day": day_limit, "night": night_limit, "total": total_limit}


def staff_month_work_counts(df, staff_name, year, month):
    """
    指定職員の指定月内の日勤・夜勤・合計勤務数を数える。
    職員名は正規化して比較するため、空白違いでも上限判定から漏れない。
    """
    if df is None or df.empty:
        return 0, 0, 0
    target_staff = normalize_staff_name(staff_name)
    work = df.copy()
    work["_staff_norm"] = work["staff_name"].apply(normalize_staff_name)
    work = work[work["_staff_norm"] == target_staff]
    if work.empty:
        return 0, 0, 0
    ym = f"{int(year)}-{int(month):02d}-"
    work = work[work["shift_date"].astype(str).str.startswith(ym)]
    if work.empty:
        return 0, 0, 0
    day_count = int(len(work[work["shift_kind"].astype(str) == "日勤"]))
    night_count = int(len(work[work["shift_kind"].astype(str) == "夜勤"]))
    total_count = int(len(work[work["shift_kind"].astype(str).isin(["日勤", "夜勤"])]))
    return day_count, night_count, total_count

def would_exceed_staff_shift_limit(df, staff_name, target_date, shift_kind, limit_map=None):
    """
    その勤務を1件追加した場合に、職員別の日勤・夜勤・合計上限を超えるか判定する。
    AIシフト案作成時の最終ガードとして使う。
    """
    if shift_kind not in ["日勤", "夜勤"]:
        return False, ""
    target = pd.to_datetime(str(target_date)).date()
    limits = normalize_shift_limits_for_staff(limit_map or get_staff_shift_limit_map(), staff_name)
    day_count, night_count, total_count = staff_month_work_counts(df, staff_name, target.year, target.month)

    add_day = 1 if shift_kind == "日勤" else 0
    add_night = 1 if shift_kind == "夜勤" else 0

    if day_count + add_day > limits["day"]:
        return True, f"日勤上限{limits['day']}回を超えるため候補外"
    if night_count + add_night > limits["night"]:
        return True, f"夜勤上限{limits['night']}回を超えるため候補外"
    if total_count + 1 > limits["total"]:
        return True, f"合計上限{limits['total']}回を超えるため候補外"
    return False, ""


def get_staff_shift_limit_map():
    """
    職員別勤務回数上限をDBから直接取得し、職員名を正規化して辞書化する。
    画面表示用の結合結果ではなく保存値そのものを見ることで、
    日勤上限0・夜勤上限0がAI判定で31扱いになる事故を防ぐ。
    """
    mapping = {}
    try:
        df = fetch_df("""
            SELECT staff_name, max_day_shifts, max_night_shifts, max_total_shifts
            FROM staff_shift_limits
        """)
    except Exception:
        df = pd.DataFrame(columns=["staff_name", "max_day_shifts", "max_night_shifts", "max_total_shifts"])

    if df is not None and not df.empty:
        for _, r in df.iterrows():
            staff_name = normalize_staff_name(r.get("staff_name", ""))
            if not staff_name:
                continue
            mapping[staff_name] = {
                "day": safe_shift_limit_value(r.get("max_day_shifts", 31), 31),
                "night": safe_shift_limit_value(r.get("max_night_shifts", 31), 31),
                "total": safe_shift_limit_value(r.get("max_total_shifts", 31), 31),
            }

    # 未設定職員だけ31回にする。設定済みの0回は絶対に上書きしない。
    for staff_name in get_active_staff():
        s = normalize_staff_name(staff_name)
        if s and s not in mapping:
            mapping[s] = {"day": 31, "night": 31, "total": 31}

    return mapping


def debug_shift_limit_summary_for_ai(limit_map=None):
    """AIが実際に参照している上限表を確認するための表示用データ。"""
    limit_map = limit_map or get_staff_shift_limit_map()
    rows = []
    for staff_name in sorted(limit_map.keys()):
        limits = normalize_shift_limits_for_staff(limit_map, staff_name)
        rows.append({
            "職員名": staff_name,
            "AI参照_日勤上限": limits["day"],
            "AI参照_夜勤上限": limits["night"],
            "AI参照_合計上限": limits["total"],
        })
    return pd.DataFrame(rows)

def create_shift_limit_check_table(matrix, limit_map=None):
    """月間勤務回数が職員別上限を超えていないか確認する。"""
    if matrix is None or matrix.empty:
        return pd.DataFrame(columns=["重要度", "種類", "職員名", "内容"])
    limit_map = limit_map or get_staff_shift_limit_map()
    rows = []
    for _, r in matrix.iterrows():
        staff_name = normalize_staff_name(r.get("職員名", ""))
        limits = normalize_shift_limits_for_staff(limit_map, staff_name)
        day_count = int(r.get("日勤", 0) or 0)
        night_count = int(r.get("夜勤", 0) or 0)
        total_count = int(r.get("合計", 0) or 0)
        if day_count > limits["day"]:
            rows.append({"重要度": "高", "種類": "日勤上限超過", "職員名": staff_name, "内容": f"日勤 {day_count}回 / 上限 {limits['day']}回"})
        if night_count > limits["night"]:
            rows.append({"重要度": "高", "種類": "夜勤上限超過", "職員名": staff_name, "内容": f"夜勤 {night_count}回 / 上限 {limits['night']}回"})
        if total_count > limits["total"]:
            rows.append({"重要度": "中", "種類": "合計上限超過", "職員名": staff_name, "内容": f"合計 {total_count}回 / 上限 {limits['total']}回"})
    return pd.DataFrame(rows, columns=["重要度", "種類", "職員名", "内容"])




def clear_month_staff_shifts(year, month):
    """
    指定月のシフトを全削除する。
    職員別勤務回数上限は残し、月間シフト表だけを空にして再入力できるようにする。
    """
    last_day = calendar.monthrange(int(year), int(month))[1]
    start = f"{int(year)}-{int(month):02d}-01"
    end = f"{int(year)}-{int(month):02d}-{last_day:02d}"
    execute("DELETE FROM staff_shifts WHERE shift_date BETWEEN ? AND ?", (start, end))
    set_shift_month_status(int(year), int(month), False, current_login_user_for_shift())
    try:
        st.session_state.pop("ai_shift_draft", None)
        st.session_state["shift_editor_reset_counter"] = int(st.session_state.get("shift_editor_reset_counter", 0) or 0) + 1
    except Exception:
        pass
    return True


def shift_month_is_confirmed(year, month):
    return bool(get_shift_month_status(year, month).get("is_confirmed"))


def get_staff_shifts(start_date, end_date, keyword=""):
    query = """
        SELECT *
        FROM staff_shifts
        WHERE shift_date BETWEEN ? AND ?
    """
    params = [str(start_date), str(end_date)]
    if keyword.strip():
        query += " AND (staff_name LIKE ? OR shift_kind LIKE ? OR memo LIKE ?)"
        kw = f"%{keyword.strip()}%"
        params.extend([kw, kw, kw])
    query += " ORDER BY shift_date, staff_name, shift_kind, id"
    return fetch_df(query, params)


def get_staff_shifts_month(year, month, include_prev_day=False):
    last_day = calendar.monthrange(int(year), int(month))[1]
    start_date = date(int(year), int(month), 1)
    if include_prev_day:
        start_date = start_date - pd.Timedelta(days=1)
        start = start_date.strftime("%Y-%m-%d")
    else:
        start = f"{int(year)}-{int(month):02d}-01"
    end = f"{int(year)}-{int(month):02d}-{last_day:02d}"
    return get_staff_shifts(start, end)


def save_single_shift(shift_date, staff_name, shift_kind, start_time=None, end_time=None, next_day=0, memo=""):
    """1件のシフトを追加保存する。"""
    staff_name = normalize_staff_name(staff_name)
    if not staff_name or not shift_kind:
        return None
    if start_time is None or end_time is None:
        start_time, end_time, default_next = default_shift_times(shift_kind)
        next_day = default_next
    return execute("""
        INSERT INTO staff_shifts
        (shift_date, staff_name, shift_kind, start_time, end_time, next_day, memo, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        shift_date,
        staff_name,
        shift_kind,
        start_time or None,
        end_time or None,
        int(next_day or 0),
        memo.strip() or None,
        now_text(),
        now_text(),
    ))


def save_basic_day_shift(shift_date, day_staff_1, day_staff_2, night_staff, memo=""):
    """
    1日分の基本シフトを保存する。
    同日の既存日勤・夜勤は一度削除し、日勤2名・夜勤1名として登録する。
    """
    execute("DELETE FROM staff_shifts WHERE shift_date=? AND shift_kind IN ('日勤', '夜勤')", (shift_date,))
    params = []
    for staff_name in [day_staff_1, day_staff_2]:
        staff_name = normalize_staff_name(staff_name)
        if staff_name:
            stime, etime, nd = default_shift_times("日勤")
            params.append((shift_date, staff_name, "日勤", stime, etime, nd, memo.strip() or None, now_text(), now_text()))
    night_staff = normalize_staff_name(night_staff)
    if night_staff:
        stime, etime, nd = default_shift_times("夜勤")
        params.append((shift_date, night_staff, "夜勤", stime, etime, nd, memo.strip() or None, now_text(), now_text()))
    return execute_many("""
        INSERT INTO staff_shifts
        (shift_date, staff_name, shift_kind, start_time, end_time, next_day, memo, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, params)


def shift_day_labels_for_staff(df, staff_name, target_date):
    """
    指定職員・指定日のラベル一覧。
    前日夜勤があれば明を補完するが、当日に希望休・有休・休み・その他がある場合は明を表示しない。
    """
    labels = []
    if df is not None and not df.empty:
        target_staff = normalize_staff_name(staff_name)
        staff_df = df.copy()
        staff_df["_staff_norm"] = staff_df["staff_name"].apply(normalize_staff_name)
        staff_df = staff_df[staff_df["_staff_norm"] == target_staff]
        day_rows = staff_df[staff_df["shift_date"].astype(str) == str(target_date)]

        day_shift_kinds = []
        for _, r in day_rows.iterrows():
            kind = str(r["shift_kind"])
            day_shift_kinds.append(kind)
            label = shift_short_label(kind)
            if label and label not in labels:
                labels.append(label)

        # 当日に休み系がある場合、明け表示は重ねない。
        has_off_label = any(x in labels for x in ["希", "有", "他"]) or any(k in OFF_OR_BLOCKING_SHIFT_KINDS for k in day_shift_kinds)

        # 前日夜勤の翌日は「明」を自動表示する。ただし希望休・有休・休み・その他がある日は表示しない。
        try:
            prev_date = (pd.to_datetime(str(target_date)).date() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
            prev_rows = staff_df[
                (staff_df["shift_date"].astype(str) == str(prev_date)) &
                (staff_df["shift_kind"].astype(str) == "夜勤")
            ]
            if not prev_rows.empty and "明" not in labels and not has_off_label:
                labels.insert(0, "明")
        except Exception:
            pass

        # 既存データで明けと希望休等が既に重なっている場合も、表示上は休み系を優先して明を消す。
        if has_off_label and "明" in labels:
            labels = [x for x in labels if x != "明"]
    return labels

def labels_to_cell_value(labels):
    labels = [x for x in labels if x]
    if not labels:
        return ""

    # 希望休・有休・その他がある日は、明けや勤務より休み系表示を優先する。
    for off_label in ["希", "有", "他"]:
        if off_label in labels:
            return off_label

    # 日/夜の同日入力は禁止。既存データで重なっている場合は日を優先表示し、チェックで警告する。
    if "日" in labels and "夜" in labels:
        return "日"

    priority = ["日", "夜", "明", "休", "希", "有", "他"]
    ordered = [x for x in priority if x in labels]
    if len(ordered) == 1:
        return ordered[0]
    return "/".join(ordered[:2])

def create_shift_matrix(df, year, month):
    """写真の勤務表イメージに近い、職員×日付の横長表を作る。"""
    last_day = calendar.monthrange(int(year), int(month))[1]
    staff_names = []
    if df is not None and not df.empty:
        staff_names = sorted(set([normalize_staff_name(x) for x in df["staff_name"].dropna().astype(str).tolist() if normalize_staff_name(x)]))
    columns = ["職員名"] + [str(d) for d in range(1, last_day + 1)] + ["日勤", "夜勤", "明", "休み", "希望休", "有休", "合計", "最大連勤"]
    rows = []

    for staff_name in staff_names:
        row = {"職員名": staff_name}
        day_count = night_count = ake_count = rest_count = hope_count = paid_count = total_count = 0
        work_flags = []
        for d in range(1, last_day + 1):
            target_date = f"{int(year)}-{int(month):02d}-{d:02d}"
            labels = shift_day_labels_for_staff(df, staff_name, target_date)
            row[str(d)] = labels_to_cell_value(labels)
            if "日" in labels:
                day_count += 1
                total_count += 1
            if "夜" in labels:
                night_count += 1
                total_count += 1
            if "明" in labels:
                ake_count += 1
            if "休" in labels:
                rest_count += 1
            if "希" in labels:
                hope_count += 1
            if "有" in labels:
                paid_count += 1
            work_flags.append(1 if ("日" in labels or "夜" in labels) else 0)

        row["日勤"] = day_count
        row["夜勤"] = night_count
        row["明"] = ake_count
        row["休み"] = rest_count
        row["希望休"] = hope_count
        row["有休"] = paid_count
        row["合計"] = total_count
        row["最大連勤"] = max_consecutive_ones(work_flags)
        rows.append(row)

    return pd.DataFrame(rows, columns=columns)


def max_consecutive_ones(values):
    best = cur = 0
    for v in values:
        if int(v or 0) == 1:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def create_editable_shift_matrix(staff_names, df, year, month):
    """
    st.data_editorで直接入力しやすい月間シフト表を作る。
    各セルは「」「日」「夜」「明」「休」「希」「有」「日/夜」からプルダウン入力する。
    """
    last_day = calendar.monthrange(int(year), int(month))[1]
    # 職員名は正規化し、空白違いの重複行を1行にまとめる。
    unique_staff_names = []
    seen_staff_names = set()
    for s in staff_names:
        ns = normalize_staff_name(s)
        if ns and ns not in seen_staff_names:
            seen_staff_names.add(ns)
            unique_staff_names.append(ns)
    staff_names = unique_staff_names

    if df is not None and not df.empty:
        for s in sorted(df["staff_name"].dropna().astype(str).unique().tolist()):
            ns = normalize_staff_name(s)
            if ns and ns not in seen_staff_names:
                seen_staff_names.add(ns)
                staff_names.append(ns)

    columns = ["職員名"] + [str(d) for d in range(1, last_day + 1)]
    rows = []

    for staff_name in staff_names:
        row = {"職員名": staff_name}
        for d in range(1, last_day + 1):
            target_date = f"{int(year)}-{int(month):02d}-{d:02d}"
            labels = shift_day_labels_for_staff(df, staff_name, target_date)
            row[str(d)] = labels_to_cell_value(labels)
        rows.append(row)

    return pd.DataFrame(rows, columns=columns)


def save_shift_matrix_from_editor(year, month, edited_df):
    """
    月間シフト表の直接編集内容をDBへ反映する。
    既存の該当月・該当職員分はいったん削除し、セルの内容から再作成する。
    """
    if edited_df is None or edited_df.empty:
        return 0

    last_day = calendar.monthrange(int(year), int(month))[1]
    staff_names = []
    seen_staff_names = set()
    for v in edited_df["職員名"].tolist():
        ns = normalize_staff_name(v)
        if ns and ns not in seen_staff_names:
            seen_staff_names.add(ns)
            staff_names.append(ns)
    if not staff_names:
        return 0

    start = f"{int(year)}-{int(month):02d}-01"
    end = f"{int(year)}-{int(month):02d}-{last_day:02d}"

    # 既存データに「職員 A」「職員A」のような空白違いがあっても、
    # 同じ正規化名としてまとめて削除してから保存する。
    existing_month_df = get_staff_shifts(start, end)
    for staff_name in staff_names:
        aliases = {staff_name}
        if existing_month_df is not None and not existing_month_df.empty:
            for old_name in existing_month_df["staff_name"].dropna().astype(str).unique().tolist():
                if normalize_staff_name(old_name) == staff_name:
                    aliases.add(old_name)
        for alias in aliases:
            execute(
                "DELETE FROM staff_shifts WHERE shift_date BETWEEN ? AND ? AND staff_name=?",
                (start, end, alias),
            )

    params = []
    for _, row in edited_df.iterrows():
        staff_name = normalize_staff_name(row.get("職員名", ""))
        if not staff_name:
            continue
        for d in range(1, last_day + 1):
            label = str(row.get(str(d), "") or "").strip()
            # 明/休のような表示が残った場合でも分解して保存する。
            labels = [label] if "/" not in label else [x.strip() for x in label.split("/") if x.strip()]
            kinds = []
            for lab in labels:
                kinds.extend(shift_kind_from_editor_label(lab))
            if not kinds:
                continue
            shift_date = f"{int(year)}-{int(month):02d}-{d:02d}"
            for kind in kinds:
                stime, etime, next_day = default_shift_times(kind)
                params.append((
                    shift_date,
                    staff_name,
                    kind,
                    stime or None,
                    etime or None,
                    int(next_day or 0),
                    "月間表から直接入力",
                    now_text(),
                    now_text(),
                ))

    if not params:
        return 0

    return execute_many("""
        INSERT INTO staff_shifts
        (shift_date, staff_name, shift_kind, start_time, end_time, next_day, memo, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, params)


def create_shift_shortage_table(df, year, month):
    """日勤2名・夜勤1名を基準に、不足日を確認する。"""
    last_day = calendar.monthrange(int(year), int(month))[1]
    rows = []
    for d in range(1, last_day + 1):
        target_date = f"{int(year)}-{int(month):02d}-{d:02d}"
        day_df = df[df["shift_date"].astype(str) == target_date] if df is not None and not df.empty else pd.DataFrame()
        day_count = night_count = 0
        if not day_df.empty:
            day_count = len(day_df[day_df["shift_kind"] == "日勤"])
            night_count = len(day_df[day_df["shift_kind"] == "夜勤"])
        status = "OK" if day_count >= 2 and night_count >= 1 else "要確認"
        rows.append({
            "日付": target_date,
            "日勤人数": day_count,
            "夜勤人数": night_count,
            "状態": status,
            "不足": f"日勤あと{max(0, 2-day_count)} / 夜勤あと{max(0, 1-night_count)}" if status != "OK" else "",
        })
    return pd.DataFrame(rows)


def create_shift_quality_check_table(df, year, month):
    """夜勤明け・連勤・希望休・勤務偏りを確認する。"""
    rows = []
    last_day = calendar.monthrange(int(year), int(month))[1]
    if df is None or df.empty:
        return pd.DataFrame(columns=["重要度", "種類", "日付", "職員名", "内容"])

    staff_names = sorted(df["staff_name"].dropna().astype(str).unique().tolist())
    for staff_name in staff_names:
        work_flags = []
        for d in range(1, last_day + 1):
            target_date = f"{int(year)}-{int(month):02d}-{d:02d}"
            day_df = df[(df["staff_name"].astype(str) == staff_name) & (df["shift_date"].astype(str) == target_date)]
            kinds = day_df["shift_kind"].astype(str).tolist() if not day_df.empty else []

            # 希望休・有休に勤務が重なっている
            if any(k in kinds for k in OFF_OR_BLOCKING_SHIFT_KINDS) and any(k in kinds for k in ACTIVE_SHIFT_KINDS):
                rows.append({
                    "重要度": "高",
                    "種類": "休み希望と勤務の重複",
                    "日付": target_date,
                    "職員名": staff_name,
                    "内容": "休み・希望休・有休・その他と、日勤・夜勤・明けが同日に入っています。",
                })

            # 同じ日に日勤と夜勤が重なっている
            if "日勤" in kinds and "夜勤" in kinds:
                rows.append({
                    "重要度": "高",
                    "種類": "日勤夜勤重複",
                    "日付": target_date,
                    "職員名": staff_name,
                    "内容": "同じ日に日勤と夜勤が入っています。日/夜入力は禁止です。",
                })

            # 夜勤翌日は「明」、明け翌日は「休み」を基本にする
            try:
                prev_date = (date(int(year), int(month), d) - timedelta(days=1)).strftime("%Y-%m-%d")
                prev_df = df[(df["staff_name"].astype(str) == staff_name) & (df["shift_date"].astype(str) == prev_date)]
                prev_kinds = prev_df["shift_kind"].astype(str).tolist() if not prev_df.empty else []

                if "夜勤" in prev_kinds:
                    if "日勤" in kinds or "夜勤" in kinds:
                        rows.append({
                            "重要度": "高",
                            "種類": "夜勤翌日勤務",
                            "日付": target_date,
                            "職員名": staff_name,
                            "内容": "夜勤の翌日に勤務が入っています。翌日は明け扱いにしてください。",
                        })
                    if "夜勤明け" not in kinds:
                        rows.append({
                            "重要度": "中",
                            "種類": "夜勤翌日明け未登録",
                            "日付": target_date,
                            "職員名": staff_name,
                            "内容": "夜勤の翌日に「明」が登録されていません。",
                        })

                if "夜勤明け" in prev_kinds:
                    if "日勤" in kinds or "夜勤" in kinds:
                        rows.append({
                            "重要度": "高",
                            "種類": "明け翌日勤務",
                            "日付": target_date,
                            "職員名": staff_name,
                            "内容": "明けの翌日に勤務が入っています。原則として休みにしてください。",
                        })
                    if not any(k in kinds for k in ["休み", "有休", "希望休"]):
                        rows.append({
                            "重要度": "中",
                            "種類": "明け翌日休み未登録",
                            "日付": target_date,
                            "職員名": staff_name,
                            "内容": "明けの翌日に休みが登録されていません。",
                        })
            except Exception:
                pass

            work_flags.append(1 if ("日勤" in kinds or "夜勤" in kinds) else 0)

        # 5連勤・6連勤以上
        cur = 0
        for idx, v in enumerate(work_flags, start=1):
            if v:
                cur += 1
                if cur == 5:
                    rows.append({
                        "重要度": "中",
                        "種類": "5連勤",
                        "日付": f"{int(year)}-{int(month):02d}-{idx:02d}",
                        "職員名": staff_name,
                        "内容": "5連勤になっています。疲労・調整を確認してください。",
                    })
                elif cur >= 6:
                    rows.append({
                        "重要度": "高",
                        "種類": f"{cur}連勤",
                        "日付": f"{int(year)}-{int(month):02d}-{idx:02d}",
                        "職員名": staff_name,
                        "内容": "6連勤以上になっています。要調整です。",
                    })
            else:
                cur = 0

    # 夜勤回数の偏り
    matrix = create_shift_matrix(df, year, month)
    if not matrix.empty and "夜勤" in matrix.columns:
        night_vals = matrix["夜勤"].fillna(0).astype(int)
        if len(night_vals) >= 2:
            avg = night_vals.mean()
            for _, r in matrix.iterrows():
                if int(r["夜勤"]) >= avg + 3 and int(r["夜勤"]) >= 4:
                    rows.append({
                        "重要度": "中",
                        "種類": "夜勤回数偏り",
                        "日付": f"{int(year)}-{int(month):02d}",
                        "職員名": r["職員名"],
                        "内容": f"夜勤が{int(r['夜勤'])}回で、平均との差が大きい可能性があります。",
                    })

    return pd.DataFrame(rows, columns=["重要度", "種類", "日付", "職員名", "内容"])


def parse_hour_from_time_text(value):
    text = str(value or "").strip()
    if not text:
        return None
    m = re.match(r"^(\d{1,2})[:：](\d{2})", text)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def get_staff_event_counts_for_date(target_date):
    df = fetch_df("""
        SELECT staff_name, COUNT(*) AS cnt
        FROM events
        WHERE event_date = ? AND staff_name IS NOT NULL AND staff_name <> ''
        GROUP BY staff_name
    """, (str(target_date),))
    if df.empty:
        return {}
    return {str(r["staff_name"]): int(r["cnt"]) for _, r in df.iterrows()}


def get_shift_candidates_for_event(event_row, shift_df):
    """予定日時に対して担当候補を出す。日中は日勤、夕方以降・早朝は夜勤を優先する。"""
    if shift_df is None or shift_df.empty:
        return pd.DataFrame()

    event_date = str(event_row["event_date"])
    start_hour = parse_hour_from_time_text(event_row.get("start_time", ""))
    night_event = start_hour is not None and (start_hour >= 16 or start_hour <= 9)

    day_df = shift_df[shift_df["shift_date"].astype(str) == event_date].copy()
    if day_df.empty:
        return pd.DataFrame()

    # 休み系・明けは候補外
    day_df = day_df[~day_df["shift_kind"].isin(["休み", "希望休", "有休", "夜勤明け"])].copy()
    if day_df.empty:
        return pd.DataFrame()

    counts = get_staff_event_counts_for_date(event_date)
    rows = []
    for _, r in day_df.iterrows():
        staff_name = str(r["staff_name"])
        shift_kind = str(r["shift_kind"])
        score = 50
        reason = ["勤務表に登録あり"]

        if night_event:
            if shift_kind == "夜勤":
                score += 40
                reason.append("夜間予定と夜勤が一致")
            elif shift_kind == "日勤":
                score -= 10
                reason.append("夜間予定だが日勤")
        else:
            if shift_kind == "日勤":
                score += 35
                reason.append("日中予定と日勤が一致")
            elif shift_kind == "夜勤":
                score -= 5
                reason.append("日中予定だが夜勤")

        assigned = counts.get(staff_name, 0)
        if assigned:
            score -= assigned * 8
            reason.append(f"同日すでに{assigned}件担当")

        rows.append({
            "職員名": staff_name,
            "シフト": shift_kind,
            "時間": f"{r['start_time'] or ''}〜{r['end_time'] or ''}{'翌' if int(r['next_day'] or 0) else ''}",
            "スコア": int(score),
            "理由": " / ".join(reason),
        })

    if not rows:
        return pd.DataFrame()
    result = pd.DataFrame(rows)
    result = result.sort_values(["職員名", "スコア"], ascending=[True, False]).groupby("職員名", as_index=False).first()
    return result.sort_values(["スコア", "職員名"], ascending=[False, True]).reset_index(drop=True)


def build_event_assignment_preview(events_df, shift_df, only_unassigned=True):
    rows = []
    if events_df is None or events_df.empty:
        return pd.DataFrame()
    for _, ev in events_df.iterrows():
        current = str(ev["staff_name"] or "")
        if only_unassigned and current:
            continue
        cand = get_shift_candidates_for_event(ev, shift_df)
        if cand.empty:
            rows.append({
                "予定ID": int(ev["id"]),
                "日付": ev["event_date"],
                "予定": ev["title"],
                "現在担当": current,
                "AI候補": "",
                "スコア": 0,
                "理由": "候補なし",
            })
        else:
            top = cand.iloc[0]
            rows.append({
                "予定ID": int(ev["id"]),
                "日付": ev["event_date"],
                "予定": ev["title"],
                "現在担当": current,
                "AI候補": top["職員名"],
                "スコア": int(top["スコア"]),
                "理由": top["理由"],
            })
    return pd.DataFrame(rows)



OFF_OR_BLOCKING_SHIFT_KINDS = ["休み", "希望休", "有休", "その他"]
ACTIVE_SHIFT_KINDS = ["日勤", "夜勤", "夜勤明け"]


def get_staff_day_shift_kinds(df, staff_name, target_date):
    """指定職員・指定日の勤務区分一覧を、職員名正規化込みで取得する。"""
    if df is None or df.empty:
        return []
    target_staff = normalize_staff_name(staff_name)
    tmp = df.copy()
    tmp["_staff_norm"] = tmp["staff_name"].apply(normalize_staff_name)
    day_df = tmp[
        (tmp["_staff_norm"] == target_staff) &
        (tmp["shift_date"].astype(str) == str(target_date))
    ]
    if day_df.empty:
        return []
    return [str(x) for x in day_df["shift_kind"].dropna().astype(str).tolist()]


def can_add_shift_without_overlap(existing_kinds, new_kind):
    """
    同じ日に入れてよい勤務かを判定する。
    希望休・有休・休み・その他がある日は、日勤・夜勤・明けを入れない。
    日勤・夜勤・明けがある日は、希望休・有休・休み・その他を重ねない。
    """
    existing_kinds = [str(k) for k in (existing_kinds or []) if str(k)]
    new_kind = str(new_kind or "").strip()
    if not new_kind:
        return False, "勤務区分が空です"

    if new_kind in existing_kinds:
        return False, f"{new_kind}が既に登録済み"

    existing_off = [k for k in existing_kinds if k in OFF_OR_BLOCKING_SHIFT_KINDS]
    existing_active = [k for k in existing_kinds if k in ACTIVE_SHIFT_KINDS]

    if new_kind in ACTIVE_SHIFT_KINDS and existing_off:
        return False, f"休み系（{', '.join(existing_off)}）が既にあるため{new_kind}は入れません"
    if new_kind in OFF_OR_BLOCKING_SHIFT_KINDS and existing_active:
        return False, f"勤務系（{', '.join(existing_active)}）が既にあるため{new_kind}は入れません"

    # 日勤・夜勤・明けは同日で重ねない
    if new_kind in ACTIVE_SHIFT_KINDS and existing_active:
        return False, f"勤務系（{', '.join(existing_active)}）が既にあるため{new_kind}は入れません"

    # 休み系同士も重ねない。希望休や有休が入っていれば休み扱いとして十分。
    if new_kind in OFF_OR_BLOCKING_SHIFT_KINDS and existing_off:
        return False, f"休み系（{', '.join(existing_off)}）が既にあるため重ねません"

    return True, "追加可能"


def can_add_shift_to_working_df(df, staff_name, target_date, new_kind):
    existing_kinds = get_staff_day_shift_kinds(df, staff_name, target_date)
    return can_add_shift_without_overlap(existing_kinds, new_kind)



def has_off_shift_on_day(df, staff_name, target_date):
    """希望休・有休・休み・その他が同日にあるか。"""
    kinds = get_staff_day_shift_kinds(df, staff_name, target_date)
    return any(k in OFF_OR_BLOCKING_SHIFT_KINDS for k in kinds), [k for k in kinds if k in OFF_OR_BLOCKING_SHIFT_KINDS]


def has_active_shift_on_day(df, staff_name, target_date):
    """日勤・夜勤・夜勤明けが同日にあるか。"""
    kinds = get_staff_day_shift_kinds(df, staff_name, target_date)
    return any(k in ACTIVE_SHIFT_KINDS for k in kinds), [k for k in kinds if k in ACTIVE_SHIFT_KINDS]


def staff_available_for_kind(df, staff_name, target_date, shift_kind, limit_map=None):
    """
    AIシフト案用。
    希望休・有休・その他・休みと、日勤・夜勤・夜勤明けが同日に重ならないよう厳密に見る。
    特に夜勤は、翌日が明けになるため、翌日に希望休等がある場合は候補外にする。
    """
    if df is None:
        df = pd.DataFrame()
    date_text = str(target_date)
    target_staff = normalize_staff_name(staff_name)

    existing_kinds = get_staff_day_shift_kinds(df, target_staff, date_text)
    overlap_ok, overlap_reason = can_add_shift_without_overlap(existing_kinds, shift_kind)
    if not overlap_ok:
        return False, overlap_reason

    try:
        prev_date = (pd.to_datetime(date_text).date() - timedelta(days=1)).strftime("%Y-%m-%d")
        prev_kinds = get_staff_day_shift_kinds(df, target_staff, prev_date)
        if "夜勤" in prev_kinds:
            return False, "前日夜勤のため翌日は明け"
        if "夜勤明け" in prev_kinds:
            return False, "前日明けのため翌日は休み"
    except Exception:
        pass

    # 夜勤を入れると翌日は必ず明け扱いになるため、翌日に希望休・有休・休み・その他がある場合は夜勤候補から外す。
    if shift_kind == "夜勤":
        try:
            next_date = (pd.to_datetime(date_text).date() + timedelta(days=1)).strftime("%Y-%m-%d")
            next_kinds = get_staff_day_shift_kinds(df, target_staff, next_date)
            next_off = [k for k in next_kinds if k in OFF_OR_BLOCKING_SHIFT_KINDS]
            next_active = [k for k in next_kinds if k in ACTIVE_SHIFT_KINDS]
            if next_off:
                return False, f"翌日に休み系（{', '.join(next_off)}）があるため夜勤不可"
            if next_active:
                return False, f"翌日に勤務系（{', '.join(next_active)}）があるため夜勤不可"
        except Exception:
            pass

    exceed, reason = would_exceed_staff_shift_limit(df, target_staff, date_text, shift_kind, limit_map=limit_map)
    if exceed:
        return False, reason

    return True, "候補"

def create_ai_shift_draft(df, staff_names, year, month):
    """
    ルール型AIで不足分のシフト案を作る。
    既存シフトとAI追加分を都度数え直し、日勤上限・夜勤上限・合計上限を絶対に超えない範囲だけで候補化する。
    """
    last_day = calendar.monthrange(int(year), int(month))[1]
    # 職員名の空白違いを統一。上限表と候補職員の名前がズレても判定漏れしないようにする。
    staff_names = sorted(set([normalize_staff_name(s) for s in staff_names if normalize_staff_name(s)]))
    if not staff_names:
        return pd.DataFrame()

    limit_map = get_staff_shift_limit_map()

    working_df = df.copy() if df is not None else pd.DataFrame()
    if working_df is None or working_df.empty:
        working_df = pd.DataFrame(columns=["shift_date", "staff_name", "shift_kind", "start_time", "end_time", "next_day", "memo", "created_at", "updated_at"])
    else:
        working_df["staff_name"] = working_df["staff_name"].apply(normalize_staff_name)

    rows = []

    # まず、既存シフトが既に上限超過している職員を明示する。
    for s in staff_names:
        limits = normalize_shift_limits_for_staff(limit_map, s)
        d_count, n_count, t_count = staff_month_work_counts(working_df, s, int(year), int(month))
        if d_count > limits["day"] or n_count > limits["night"] or t_count > limits["total"]:
            rows.append({
                "日付": f"{int(year)}-{int(month):02d}-01",
                "勤務": "注意",
                "候補職員": s,
                "理由": f"既存シフトが上限超過しています（日勤 {d_count}/{limits['day']}、夜勤 {n_count}/{limits['night']}、合計 {t_count}/{limits['total']}）。AIはこの職員へ追加しません。",
                "保存対象": False,
            })

    for d in range(1, last_day + 1):
        target_date = f"{int(year)}-{int(month):02d}-{d:02d}"
        day_df = working_df[working_df["shift_date"].astype(str) == target_date] if not working_df.empty else pd.DataFrame()
        current_day = int(len(day_df[day_df["shift_kind"].astype(str) == "日勤"])) if not day_df.empty else 0
        current_night = int(len(day_df[day_df["shift_kind"].astype(str) == "夜勤"])) if not day_df.empty else 0

        need_list = ["日勤"] * max(0, 2 - current_day) + ["夜勤"] * max(0, 1 - current_night)

        for need_kind in need_list:
            candidates = []
            reject_reasons = []

            for s in staff_names:
                limits = normalize_shift_limits_for_staff(limit_map, s)
                d_count, n_count, t_count = staff_month_work_counts(working_df, s, int(year), int(month))

                # 上限を先に見る。ここで弾くので、夜勤0回の職員に夜勤が入ることはない。
                if need_kind == "日勤" and d_count >= limits["day"]:
                    reject_reasons.append(f"{s}: 日勤上限{limits['day']}回")
                    continue
                if need_kind == "夜勤" and n_count >= limits["night"]:
                    reject_reasons.append(f"{s}: 夜勤上限{limits['night']}回")
                    continue
                if t_count >= limits["total"]:
                    reject_reasons.append(f"{s}: 合計上限{limits['total']}回")
                    continue

                ok, why = staff_available_for_kind(working_df, s, target_date, need_kind, limit_map=limit_map)
                if not ok:
                    reject_reasons.append(f"{s}: {why}")
                    continue

                # 上限内の職員だけを、少ない人優先で点数化
                score = 100
                score -= t_count * 3
                if need_kind == "日勤":
                    score -= d_count * 2
                    # 日勤残り枠が少ない人は少し下げる
                    score -= max(0, 3 - (limits["day"] - d_count))
                if need_kind == "夜勤":
                    score -= n_count * 5
                    score -= max(0, 3 - (limits["night"] - n_count)) * 2

                try:
                    consecutive_before = 0
                    for back in range(1, 6):
                        prev_date = (pd.to_datetime(target_date).date() - timedelta(days=back)).strftime("%Y-%m-%d")
                        p_df = working_df[
                            (working_df["staff_name"].apply(normalize_staff_name) == s) &
                            (working_df["shift_date"].astype(str) == prev_date)
                        ] if not working_df.empty else pd.DataFrame()
                        if not p_df.empty and any(k in p_df["shift_kind"].astype(str).tolist() for k in ["日勤", "夜勤"]):
                            consecutive_before += 1
                        else:
                            break
                    if consecutive_before >= 4:
                        score -= 60
                except Exception:
                    pass

                candidates.append((score, s))

            if not candidates:
                reason_text = "日勤・夜勤・合計上限、希望休、有休、休み、その他、夜勤翌日の明け重複を考慮した結果、候補なし"
                if reject_reasons:
                    reason_text += "（" + " / ".join(reject_reasons[:6]) + (" ほか" if len(reject_reasons) > 6 else "") + "）"
                rows.append({
                    "日付": target_date,
                    "勤務": need_kind,
                    "候補職員": "",
                    "理由": reason_text,
                    "保存対象": False,
                })
                continue

            candidates.sort(reverse=True)
            selected = candidates[0][1]

            # 最終ガード：追加直前にもう一度、現在のworking_dfで上限超過を確認する
            exceed, limit_reason = would_exceed_staff_shift_limit(
                working_df, selected, target_date, need_kind, limit_map=limit_map
            )
            if exceed:
                rows.append({
                    "日付": target_date,
                    "勤務": need_kind,
                    "候補職員": "",
                    "理由": limit_reason,
                    "保存対象": False,
                })
                continue

            rows.append({
                "日付": target_date,
                "勤務": need_kind,
                "候補職員": selected,
                "理由": "日勤・夜勤・合計上限内で候補化",
                "保存対象": True,
            })

            stime, etime, nd = default_shift_times(need_kind)
            add_row = pd.DataFrame([{
                "shift_date": target_date,
                "staff_name": selected,
                "shift_kind": need_kind,
                "start_time": stime,
                "end_time": etime,
                "next_day": nd,
                "memo": "AIシフト案",
                "created_at": now_text(),
                "updated_at": now_text(),
            }])
            working_df = pd.concat([working_df, add_row], ignore_index=True)

            if need_kind == "夜勤":
                # 夜勤の翌日は「明」、明けの翌日は「休み」を自動で入れる
                try:
                    next_date_obj = pd.to_datetime(target_date).date() + timedelta(days=1)
                    rest_date_obj = pd.to_datetime(target_date).date() + timedelta(days=2)

                    if next_date_obj.month == int(month):
                        next_date = next_date_obj.strftime("%Y-%m-%d")
                        next_kinds = get_staff_day_shift_kinds(working_df, selected, next_date)
                        can_add_ake, ake_reason = can_add_shift_without_overlap(next_kinds, "夜勤明け")
                        if not can_add_ake:
                            # 既に夜勤明けがある場合も重複追加しない。休み・希望休・有休・その他がある場合は明けを重ねない。
                            rows.append({
                                "日付": next_date,
                                "勤務": "注意",
                                "候補職員": selected,
                                "理由": f"夜勤翌日の明けは追加しません：{ake_reason}",
                                "保存対象": False,
                            })
                        else:
                            rows.append({
                                "日付": next_date,
                                "勤務": "夜勤明け",
                                "候補職員": selected,
                                "理由": "夜勤翌日の明けを自動付与",
                                "保存対象": True,
                            })
                            ake_row = pd.DataFrame([{
                                "shift_date": next_date,
                                "staff_name": selected,
                                "shift_kind": "夜勤明け",
                                "start_time": "",
                                "end_time": "",
                                "next_day": 0,
                                "memo": "夜勤翌日の明け自動付与",
                                "created_at": now_text(),
                                "updated_at": now_text(),
                            }])
                            working_df = pd.concat([working_df, ake_row], ignore_index=True)
                    else:
                        rows.append({
                            "日付": next_date_obj.strftime("%Y-%m-%d"),
                            "勤務": "注意",
                            "候補職員": selected,
                            "理由": "夜勤翌日の明けが翌月になるため、翌月シフトで確認してください。",
                            "保存対象": False,
                        })

                    if rest_date_obj.month == int(month):
                        rest_date = rest_date_obj.strftime("%Y-%m-%d")
                        rest_kinds = get_staff_day_shift_kinds(working_df, selected, rest_date)
                        can_add_rest, rest_reason = can_add_shift_without_overlap(rest_kinds, "休み")
                        if not can_add_rest:
                            # 希望休・有休・その他が既にある日は、それを休み扱いとして休みを重ねない。
                            rows.append({
                                "日付": rest_date,
                                "勤務": "注意",
                                "候補職員": selected,
                                "理由": f"明け翌日の休みは追加しません：{rest_reason}",
                                "保存対象": False,
                            })
                        else:
                            rows.append({
                                "日付": rest_date,
                                "勤務": "休み",
                                "候補職員": selected,
                                "理由": "明け翌日の休みを自動付与",
                                "保存対象": True,
                            })
                            rest_row = pd.DataFrame([{
                                "shift_date": rest_date,
                                "staff_name": selected,
                                "shift_kind": "休み",
                                "start_time": "",
                                "end_time": "",
                                "next_day": 0,
                                "memo": "明け翌日の休み自動付与",
                                "created_at": now_text(),
                                "updated_at": now_text(),
                            }])
                            working_df = pd.concat([working_df, rest_row], ignore_index=True)
                    else:
                        rows.append({
                            "日付": rest_date_obj.strftime("%Y-%m-%d"),
                            "勤務": "注意",
                            "候補職員": selected,
                            "理由": "明け翌日の休みが翌月になるため、翌月シフトで確認してください。",
                            "保存対象": False,
                        })
                except Exception as e:
                    rows.append({
                        "日付": target_date,
                        "勤務": "注意",
                        "候補職員": selected,
                        "理由": f"夜勤後の明け・休み自動付与に失敗しました：{e}",
                        "保存対象": False,
                    })

    return pd.DataFrame(rows)

def apply_ai_shift_draft_to_df(base_df, draft_df):
    """AIシフト案をDB保存前に画面上で仮反映したDataFrameを作る。重複する案は仮反映しない。"""
    work = base_df.copy() if base_df is not None else pd.DataFrame()
    if work is not None and not work.empty and "staff_name" in work.columns:
        work["staff_name"] = work["staff_name"].apply(normalize_staff_name)
    if draft_df is None or draft_df.empty:
        return work
    if work is None or work.empty:
        work = pd.DataFrame(columns=["id", "shift_date", "staff_name", "shift_kind", "start_time", "end_time", "next_day", "memo", "created_at", "updated_at"])

    rows = []
    for _, r in draft_df.iterrows():
        if not bool(r.get("保存対象", False)):
            continue
        staff_name = normalize_staff_name(r.get("候補職員", ""))
        shift_kind = str(r.get("勤務", "")).strip()
        shift_date = str(r.get("日付", "")).strip()
        if not staff_name or not shift_kind or not shift_date:
            continue
        ok, _ = can_add_shift_to_working_df(work, staff_name, shift_date, shift_kind)
        if not ok:
            continue
        stime, etime, nd = default_shift_times(shift_kind)
        new_row = {
            "id": -1,
            "shift_date": shift_date,
            "staff_name": staff_name,
            "shift_kind": shift_kind,
            "start_time": stime,
            "end_time": etime,
            "next_day": nd,
            "memo": "AIシフト案（未保存）",
            "created_at": now_text(),
            "updated_at": now_text(),
        }
        rows.append(new_row)
        work = pd.concat([work, pd.DataFrame([new_row])], ignore_index=True)
    return work



def sanitize_ai_shift_draft_by_limits(draft_df, base_df, year, month):
    """
    session_stateに残った古いAI案も、現在の上限・重複条件で再検査する。
    希望休・有休・休み・その他がある日は、日勤・夜勤・明けを重ねない。
    """
    if draft_df is None or draft_df.empty:
        return draft_df

    working_df = base_df.copy() if base_df is not None else pd.DataFrame()
    if working_df is None or working_df.empty:
        working_df = pd.DataFrame(columns=["shift_date", "staff_name", "shift_kind", "start_time", "end_time", "next_day", "memo", "created_at", "updated_at"])
    else:
        working_df["staff_name"] = working_df["staff_name"].apply(normalize_staff_name)

    limit_map = get_staff_shift_limit_map()
    cleaned_rows = []

    for _, r in draft_df.iterrows():
        shift_date = str(r.get("日付", "")).strip()
        shift_kind = str(r.get("勤務", "")).strip()
        staff_name = normalize_staff_name(r.get("候補職員", ""))
        save_target = bool(r.get("保存対象", False))

        if save_target and staff_name and shift_kind:
            # 全勤務区分について同日重複・休み系との重なりを再判定
            overlap_ok, overlap_reason = can_add_shift_to_working_df(working_df, staff_name, shift_date, shift_kind)
            if not overlap_ok:
                cleaned_rows.append({
                    "日付": shift_date,
                    "勤務": "注意",
                    "候補職員": staff_name,
                    "理由": f"現在の重複条件で除外しました：{shift_kind} / {overlap_reason}",
                    "保存対象": False,
                })
                continue

            # 日勤・夜勤は上限も再判定
            if shift_kind in ["日勤", "夜勤"]:
                ok, why = staff_available_for_kind(working_df, staff_name, shift_date, shift_kind, limit_map=limit_map)
                if not ok:
                    cleaned_rows.append({
                        "日付": shift_date,
                        "勤務": "注意",
                        "候補職員": staff_name,
                        "理由": f"現在の上限条件で除外しました：{shift_kind} / {why}",
                        "保存対象": False,
                    })
                    continue

            cleaned_rows.append({
                "日付": shift_date,
                "勤務": shift_kind,
                "候補職員": staff_name,
                "理由": str(r.get("理由", "現在条件で確認済み")),
                "保存対象": True,
            })
            stime, etime, nd = default_shift_times(shift_kind)
            add_row = pd.DataFrame([{
                "shift_date": shift_date,
                "staff_name": staff_name,
                "shift_kind": shift_kind,
                "start_time": stime,
                "end_time": etime,
                "next_day": nd,
                "memo": "AIシフト案",
                "created_at": now_text(),
                "updated_at": now_text(),
            }])
            working_df = pd.concat([working_df, add_row], ignore_index=True)
            continue

        row_dict = dict(r)
        if "候補職員" in row_dict:
            row_dict["候補職員"] = staff_name
        cleaned_rows.append(row_dict)

    return pd.DataFrame(cleaned_rows, columns=["日付", "勤務", "候補職員", "理由", "保存対象"])



def save_ai_shift_draft_rows(draft_df):
    """
    AIシフト案を保存する直前にも再検査する。
    画面表示後に上限やシフトが変更されても、DB保存時に日勤・夜勤・合計上限を超えないようにする。
    """
    if draft_df is None or draft_df.empty:
        return 0

    save_targets = draft_df[draft_df.get("保存対象", False).astype(bool)].copy()
    if save_targets.empty:
        return 0

    # 複数月が混在する可能性は低いが、安全のため日付順に処理する
    save_targets["_sort_date"] = save_targets["日付"].astype(str)
    save_targets = save_targets.sort_values(["_sort_date", "勤務", "候補職員"])

    first_date = pd.to_datetime(str(save_targets.iloc[0]["日付"])).date()
    working_df = get_staff_shifts_month(first_date.year, first_date.month, include_prev_day=True)
    if working_df is not None and not working_df.empty and "staff_name" in working_df.columns:
        working_df["staff_name"] = working_df["staff_name"].apply(normalize_staff_name)
    limit_map = get_staff_shift_limit_map()

    params = []
    for _, r in save_targets.iterrows():
        staff_name = normalize_staff_name(r.get("候補職員", ""))
        shift_kind = str(r.get("勤務", "")).strip()
        shift_date = str(r.get("日付", "")).strip()
        if not staff_name or not shift_kind or not shift_date:
            continue

        # すべての勤務区分で、同日重複・休み系との重なりを再確認
        overlap_ok, overlap_reason = can_add_shift_to_working_df(working_df, staff_name, shift_date, shift_kind)
        if not overlap_ok:
            continue

        # 日勤・夜勤は上限、夜勤翌日等も再確認
        if shift_kind in ["日勤", "夜勤"]:
            ok, why = staff_available_for_kind(working_df, staff_name, shift_date, shift_kind, limit_map=limit_map)
            if not ok:
                continue

        stime, etime, nd = default_shift_times(shift_kind)
        params.append((shift_date, staff_name, shift_kind, stime or None, etime or None, int(nd or 0), "AIシフト案から保存", now_text(), now_text()))

        add_row = pd.DataFrame([{
            "shift_date": shift_date,
            "staff_name": staff_name,
            "shift_kind": shift_kind,
            "start_time": stime,
            "end_time": etime,
            "next_day": nd,
            "memo": "AIシフト案から保存",
            "created_at": now_text(),
            "updated_at": now_text(),
        }])
        working_df = pd.concat([working_df, add_row], ignore_index=True)

    if not params:
        return 0
    return execute_many("""
        INSERT INTO staff_shifts
        (shift_date, staff_name, shift_kind, start_time, end_time, next_day, memo, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, params)

def make_staff_shift_pdf(year, month):
    """掲示・印刷しやすい横長のシフトPDFを作成する。"""
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab がインストールされていません。")

    init_pdf_fonts()

    df = get_staff_shifts_month(int(year), int(month), include_prev_day=True)
    matrix = create_shift_matrix(df, int(year), int(month))
    shortage = create_shift_shortage_table(df, int(year), int(month))
    checks = create_shift_quality_check_table(df, int(year), int(month))
    limit_checks = create_shift_limit_check_table(matrix)
    status = get_shift_month_status(int(year), int(month))

    buffer = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    margin = 16
    title_y = height - 26
    c.setFont(PDF_FONT_GOTHIC, 14)
    c.drawString(margin, title_y, f"従業員勤務表　{int(year)}年{int(month)}月")
    c.setFont(PDF_FONT_GOTHIC, 8)
    status_text = "確定" if status.get("is_confirmed") else "作成中"
    c.drawRightString(width - margin, title_y, f"{status_text}　出力日: {today_jst().strftime('%Y-%m-%d')}")

    last_day = calendar.monthrange(int(year), int(month))[1]
    table_x = margin
    table_y_top = height - 46
    staff_w = 54
    summary_w = 22
    day_w = (width - margin * 2 - staff_w - summary_w * 7) / last_day
    row_h = 15

    def draw_shift_table_vertical_lines(y0, y1):
        """日付ごとの縦罫線を細く引く。"""
        old_stroke = getattr(c, "_strokeColorObj", colors.black)
        old_width = getattr(c, "_lineWidth", 1)
        c.setStrokeColor(colors.HexColor("#b8b8b8"))
        c.setLineWidth(0.25)

        # 氏名欄と日付欄の境界
        x_line = table_x + staff_w
        c.line(x_line, y0, x_line, y1)

        # 日付欄の縦罫線
        for d in range(1, last_day + 1):
            x_line = table_x + staff_w + day_w * d
            c.line(x_line, y0, x_line, y1)

        # 集計欄の縦罫線
        summary_start = table_x + staff_w + day_w * last_day
        for i in range(1, 7):
            x_line = summary_start + summary_w * i
            c.line(x_line, y0, x_line, y1)

        c.setStrokeColor(old_stroke)
        c.setLineWidth(old_width)

    # 凡例
    c.setFont(PDF_FONT_GOTHIC, 7)
    c.drawString(margin, height - 38, "凡例：日=日勤 8:30〜17:30　夜=夜勤 16:30〜翌9:30　明=夜勤明け　希=希望休　有=有休　※休みは日別セルには表示しません")

    # ヘッダ
    c.setFillColor(colors.HexColor("#f3eee6"))
    c.rect(table_x, table_y_top - row_h, width - margin * 2, row_h, fill=1, stroke=1)
    draw_shift_table_vertical_lines(table_y_top - row_h, table_y_top)
    c.setFillColor(colors.black)
    c.setFont(PDF_FONT_GOTHIC, 6.2)
    c.drawString(table_x + 3, table_y_top - 10, "氏名")
    x = table_x + staff_w
    for d in range(1, last_day + 1):
        c.drawCentredString(x + day_w / 2, table_y_top - 10, str(d))
        x += day_w
    for label in ["日", "夜", "明", "休", "希", "有", "計"]:
        c.drawCentredString(x + summary_w / 2, table_y_top - 10, label)
        x += summary_w

    y = table_y_top - row_h
    if matrix.empty:
        c.setFont(PDF_FONT_GOTHIC, 10)
        c.drawString(margin, y - 24, "この月のシフトは登録されていません。")
    else:
        c.setFont(PDF_FONT_GOTHIC, 5.8)
        for _, row in matrix.iterrows():
            if y < 58:
                c.showPage()
                c.setPageSize(page_size)
                y = height - 30
            y -= row_h
            c.setFillColor(colors.white)
            c.rect(table_x, y, width - margin * 2, row_h, fill=1, stroke=1)
            draw_shift_table_vertical_lines(y, y + row_h)
            c.setFillColor(colors.black)
            c.drawString(table_x + 3, y + 4, str(row["職員名"])[:8])
            x = table_x + staff_w
            for d in range(1, last_day + 1):
                val = str(row[str(d)] or "")
                c.drawCentredString(x + day_w / 2, y + 4, val)
                x += day_w
            for key in ["日勤", "夜勤", "明", "休み", "希望休", "有休", "合計"]:
                c.drawCentredString(x + summary_w / 2, y + 4, str(row.get(key, "")))
                x += summary_w

    # 不足・警告を下部表示
    y -= 18
    c.setFont(PDF_FONT_GOTHIC, 7)
    ng = shortage[shortage["状態"] != "OK"]
    if not ng.empty:
        c.setFillColor(colors.HexColor("#8a2d18"))
        c.drawString(margin, y, "不足日: " + " / ".join([f"{str(r['日付'])[-2:]}日 {r['不足']}" for _, r in ng.head(10).iterrows()]))
        y -= 10
    if checks is not None and not checks.empty:
        c.setFillColor(colors.HexColor("#8a2d18"))
        c.drawString(margin, y, "確認事項: " + " / ".join([f"{r['日付']} {r['職員名']} {r['種類']}" for _, r in checks.head(6).iterrows()]))
        y -= 10
    if limit_checks is not None and not limit_checks.empty:
        c.setFillColor(colors.HexColor("#8a2d18"))
        c.drawString(margin, y, "上限確認: " + " / ".join([f"{r['職員名']} {r['種類']}" for _, r in limit_checks.head(6).iterrows()]))
    c.setFillColor(colors.black)

    c.save()
    buffer.seek(0)
    return buffer.getvalue()



def make_staff_shift_excel(year, month):
    """
    確定済みシフトをExcel（xlsx）で出力する。
    勤務表シートに月間表、別シートに不足・確認事項・上限確認を出す。
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl がインストールされていません。requirements.txt に openpyxl を追加してください。")

    year = int(year)
    month = int(month)
    df = get_staff_shifts_month(year, month, include_prev_day=True)
    matrix = create_shift_matrix(df, year, month)
    shortage = create_shift_shortage_table(df, year, month)
    checks = create_shift_quality_check_table(df, year, month)
    limit_checks = create_shift_limit_check_table(matrix)
    status = get_shift_month_status(year, month)
    last_day = calendar.monthrange(year, month)[1]

    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = "勤務表"

    title_fill = PatternFill("solid", fgColor="F3EEE6")
    header_fill = PatternFill("solid", fgColor="E7EEF8")
    weekend_sun_fill = PatternFill("solid", fgColor="FCE4D6")
    weekend_sat_fill = PatternFill("solid", fgColor="DDEBF7")
    summary_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(matrix.columns) if not matrix.empty else 8))
    ws.cell(1, 1).value = f"従業員勤務表　{year}年{month}月"
    ws.cell(1, 1).font = Font(bold=True, size=16)
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).alignment = left

    status_text = "確定" if status.get("is_confirmed") else "作成中"
    ws.cell(2, 1).value = "状態"
    ws.cell(2, 2).value = status_text
    ws.cell(2, 3).value = "確定日時"
    ws.cell(2, 4).value = status.get("confirmed_at", "") or ""
    ws.cell(2, 5).value = "出力日"
    ws.cell(2, 6).value = today_jst().strftime("%Y-%m-%d")

    ws.cell(3, 1).value = "凡例"
    ws.cell(3, 2).value = "日=日勤 8:30〜17:30 / 夜=夜勤 16:30〜翌9:30 / 明=夜勤明け / 希=希望休 / 有=有休 / 休みは日別セルには表示しません"
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=max(8, len(matrix.columns) if not matrix.empty else 8))

    start_row = 5
    if matrix is None or matrix.empty:
        ws.cell(start_row, 1).value = "この月のシフトはまだ登録されていません。"
    else:
        columns = list(matrix.columns)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(start_row, col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill
            if str(col_name).isdigit():
                d = int(col_name)
                weekday = date(year, month, d).weekday()  # 月=0 日=6
                if weekday == 6:
                    cell.fill = weekend_sun_fill
                elif weekday == 5:
                    cell.fill = weekend_sat_fill
            elif col_name in ["日勤", "夜勤", "明", "休み", "希望休", "有休", "合計", "最大連勤"]:
                cell.fill = summary_fill

        for row_idx, (_, row) in enumerate(matrix.iterrows(), start=start_row + 1):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row_idx, col_idx)
                cell.value = value
                cell.alignment = center if col_name != "職員名" else left
                cell.border = border
                if str(col_name).isdigit():
                    d = int(col_name)
                    weekday = date(year, month, d).weekday()
                    if weekday == 6:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    elif weekday == 5:
                        cell.fill = PatternFill("solid", fgColor="EAF3F8")
                elif col_name in ["日勤", "夜勤", "明", "休み", "希望休", "有休", "合計", "最大連勤"]:
                    cell.fill = PatternFill("solid", fgColor="F2F8EE")

        ws.freeze_panes = "B6"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(matrix)}"

        # 列幅調整
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            if col_name == "職員名":
                ws.column_dimensions[letter].width = 14
            elif str(col_name).isdigit():
                ws.column_dimensions[letter].width = 4
            else:
                ws.column_dimensions[letter].width = 8

    for row in range(1, 4):
        for col in range(1, max(8, (len(matrix.columns) if not matrix.empty else 8)) + 1):
            ws.cell(row, col).alignment = left if col <= 2 else center

    def add_df_sheet(sheet_name, data_df, empty_message):
        safe_name = sheet_name[:31]
        s = wb.create_sheet(safe_name)
        if data_df is None or data_df.empty:
            s.cell(1, 1).value = empty_message
            s.cell(1, 1).font = Font(bold=True)
            s.column_dimensions["A"].width = 60
            return s
        cols = list(data_df.columns)
        for cidx, col in enumerate(cols, start=1):
            cell = s.cell(1, cidx)
            cell.value = col
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            s.column_dimensions[get_column_letter(cidx)].width = 18 if col not in ["内容", "理由"] else 55
        for ridx, (_, r) in enumerate(data_df.iterrows(), start=2):
            for cidx, col in enumerate(cols, start=1):
                value = r.get(col, "")
                if pd.isna(value):
                    value = ""
                cell = s.cell(ridx, cidx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left" if col in ["内容", "理由"] else "center", vertical="center", wrap_text=True)
                cell.border = border
        s.freeze_panes = "A2"
        s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data_df)+1}"
        return s

    ng = shortage[shortage["状態"] != "OK"] if shortage is not None and not shortage.empty else pd.DataFrame()
    add_df_sheet("不足確認", ng, "日勤2名・夜勤1名の不足日はありません。")
    add_df_sheet("確認事項", checks, "夜勤翌日勤務・明け翌日勤務・5連勤以上・希望休重複などの確認事項はありません。")
    add_df_sheet("上限確認", limit_checks, "職員別の日勤・夜勤・合計勤務回数は上限内です。")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def page_shift_manager():
    st.subheader("職員シフト管理・AI担当割当")
    st.caption("希望休 → AIシフト案 → 不足・連勤チェック → 人が修正 → 確定 → PDF出力の流れで使えます。")

    today = today_jst()
    c1, c2 = st.columns(2)
    with c1:
        shift_year = st.number_input("表示年", min_value=2020, max_value=2100, value=today.year, step=1)
    with c2:
        shift_month = st.number_input("表示月", min_value=1, max_value=12, value=today.month, step=1)

    month_status = get_shift_month_status(int(shift_year), int(shift_month))
    if month_status.get("is_confirmed"):
        st.success(f"この月のシフトは確定済みです。確定日時：{month_status.get('confirmed_at', '')}")
    else:
        st.info("この月のシフトは作成中です。")

    staff_options = [""] + get_active_staff()

    st.markdown("### 1. 希望休・基本シフト入力")
    with st.form("hope_shift_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            hope_date = st.date_input("希望休・休み日", value=today, key="hope_date")
        with c2:
            hope_staff = st.selectbox("職員", staff_options, key="hope_staff")
        with c3:
            hope_kind = st.selectbox("区分", ["希望休", "有休", "休み"], key="hope_kind")
        hope_memo = st.text_input("希望休メモ", placeholder="例：本人希望、通院、家庭都合など")
        submit_hope = st.form_submit_button("希望休・休みを保存")
    if submit_hope:
        if not hope_staff:
            st.error("職員を選択してください。")
        else:
            save_single_shift(hope_date.strftime("%Y-%m-%d"), hope_staff, hope_kind, None, None, 0, hope_memo)
            st.success(f"{hope_staff} さんの {hope_kind} を保存しました。")
            st.rerun()

    with st.expander("1日分の基本シフト入力"):
        with st.form("basic_shift_form", clear_on_submit=True):
            shift_date = st.date_input("シフト日", value=today)
            c1, c2, c3 = st.columns(3)
            with c1:
                day_staff_1 = st.selectbox("日勤1（8:30〜17:30）", staff_options, key="day_staff_1")
            with c2:
                day_staff_2 = st.selectbox("日勤2（8:30〜17:30）", staff_options, key="day_staff_2")
            with c3:
                night_staff = st.selectbox("夜勤（16:30〜翌9:30）", staff_options, key="night_staff")
            shift_memo = st.text_input("シフトメモ", placeholder="例：研修、応援、希望休調整など")
            submit_basic = st.form_submit_button("この日の基本シフトを保存")
        if submit_basic:
            if not day_staff_1 and not day_staff_2 and not night_staff:
                st.error("少なくとも1名は選択してください。")
            else:
                saved = save_basic_day_shift(shift_date.strftime("%Y-%m-%d"), day_staff_1, day_staff_2, night_staff, shift_memo)
                st.success(f"{shift_date.strftime('%Y-%m-%d')} の基本シフトを {saved} 件保存しました。")
                st.rerun()

    with st.expander("個別シフトを追加・調整する"):
        with st.form("single_shift_form", clear_on_submit=True):
            s_date = st.date_input("日付", value=today, key="single_shift_date")
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                s_staff = st.selectbox("職員", staff_options, key="single_shift_staff")
            with c2:
                s_kind = st.selectbox("勤務区分", SHIFT_KINDS, key="single_shift_kind")
            default_start, default_end, default_next = default_shift_times("日勤")
            with c3:
                s_start = st.text_input("開始", value=default_start, key="single_shift_start")
            with c4:
                s_end = st.text_input("終了", value=default_end, key="single_shift_end")
            s_next = st.checkbox("終了は翌日", value=False)
            s_memo = st.text_input("メモ", key="single_shift_memo")
            add_single = st.form_submit_button("個別シフトを追加")
        if add_single:
            if not s_staff:
                st.error("職員を選択してください。")
            else:
                save_single_shift(s_date.strftime("%Y-%m-%d"), s_staff, s_kind, s_start, s_end, 1 if s_next else 0, s_memo)
                st.success("個別シフトを追加しました。")
                st.rerun()

    st.markdown("---")
    st.markdown("### 2. 月間シフト直接入力・表示")

    shift_df = get_staff_shifts_month(int(shift_year), int(shift_month), include_prev_day=True)
    matrix = create_shift_matrix(shift_df, int(shift_year), int(shift_month))
    shortage = create_shift_shortage_table(shift_df, int(shift_year), int(shift_month))
    checks = create_shift_quality_check_table(shift_df, int(shift_year), int(shift_month))

    staff_names_for_editor = get_active_staff()
    editable_matrix = create_editable_shift_matrix(
        staff_names_for_editor,
        shift_df,
        int(shift_year),
        int(shift_month),
    )

    if editable_matrix.empty:
        st.info("職員マスタに職員が登録されていません。先に職員マスタで職員を登録してください。")
    else:
        st.caption("各セルをクリックして、プルダウンから「日」「夜」「明」「希」「有」を直接入力できます。休みはシフト表には表示せず、空欄として扱います。空欄にすると削除扱いになります。")

        with st.expander("月間シフトを全部クリアして再入力する"):
            st.warning("この操作を行うと、表示中の月のシフト入力内容（日勤・夜勤・休み・希望休・有休など）をすべて削除します。職員別勤務回数上限は残ります。")
            clear_confirm = st.checkbox(
                f"{int(shift_year)}年{int(shift_month)}月のシフトを全クリアすることを確認しました",
                key=f"clear_month_shift_confirm_{int(shift_year)}_{int(shift_month)}",
            )
            if st.button("この月のシフトを全クリアする", use_container_width=True, disabled=not clear_confirm):
                clear_month_staff_shifts(int(shift_year), int(shift_month))
                st.success("この月のシフトを全クリアしました。空の月間表から再入力できます。")
                st.rerun()

        last_day_for_editor = calendar.monthrange(int(shift_year), int(shift_month))[1]
        column_config = {
            "職員名": st.column_config.TextColumn("職員名", disabled=True, width="medium")
        }
        for d in range(1, last_day_for_editor + 1):
            column_config[str(d)] = st.column_config.SelectboxColumn(
                str(d),
                options=SHIFT_EDITOR_OPTIONS,
                required=False,
                width="small",
            )

        edited_matrix = st.data_editor(
            editable_matrix,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=column_config,
            key=f"shift_matrix_editor_{int(shift_year)}_{int(shift_month)}_{st.session_state.get('shift_editor_reset_counter', 0)}",
        )

        if month_status.get("is_confirmed"):
            st.warning("確定済みの月です。修正する場合は、先に「確定を解除」してください。")
        else:
            if st.button("月間シフト表の直接入力内容を保存", use_container_width=True):
                saved = save_shift_matrix_from_editor(int(shift_year), int(shift_month), edited_matrix)
                st.success(f"月間シフト表を保存しました。登録件数：{saved}件")
                st.rerun()

    st.markdown("### 3. 職員別勤務回数上限")
    st.caption("職員ごとに、この月に入れる日勤・夜勤・合計勤務の上限を設定できます。AIシフト案はこの上限を超えない範囲で候補を出します。0回指定も有効です。")
    limits_df = get_staff_shift_limits()
    if limits_df.empty:
        st.info("職員マスタに職員が登録されていません。")
    else:
        edited_limits = st.data_editor(
            limits_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "職員名": st.column_config.TextColumn("職員名", disabled=True, width="medium"),
                "日勤上限": st.column_config.NumberColumn("日勤上限", min_value=0, max_value=31, step=1),
                "夜勤上限": st.column_config.NumberColumn("夜勤上限", min_value=0, max_value=31, step=1),
                "合計上限": st.column_config.NumberColumn("合計上限", min_value=0, max_value=31, step=1),
                "メモ": st.column_config.TextColumn("メモ", width="medium"),
            },
            key=f"shift_limit_editor_{int(shift_year)}_{int(shift_month)}",
        )
        if st.button("職員別勤務回数上限を保存", use_container_width=True):
            saved = save_staff_shift_limits_from_editor(edited_limits)
            st.success(f"職員別勤務回数上限を保存しました。登録件数：{saved}件")
            st.rerun()

    st.markdown("### 4. AIシフト案作成")
    st.caption("不足している日勤・夜勤について、希望休・有休・夜勤翌日の明け・明け翌日休み・職員別上限を超えない範囲で下書きを作ります。保存前に仮シフト表で確認できます。")

    # AIが実際に参照している上限を表示する。0回指定が反映されているか確認できます。
    with st.expander("AIが参照する職員別上限を確認"):
        st.dataframe(debug_shift_limit_summary_for_ai(), use_container_width=True, hide_index=True)

    # 古いAI案がブラウザ側に残っている場合は破棄する。
    if st.session_state.get("ai_shift_rule_version") != AI_SHIFT_RULE_VERSION:
        st.session_state.pop("ai_shift_draft", None)
        st.session_state["ai_shift_rule_version"] = AI_SHIFT_RULE_VERSION

    if st.button("不足分のAIシフト案を作成", use_container_width=True):
        st.session_state.pop("ai_shift_draft", None)
        st.session_state["ai_shift_rule_version"] = AI_SHIFT_RULE_VERSION
        st.session_state["ai_shift_draft"] = create_ai_shift_draft(
            shift_df,
            get_active_staff(),
            int(shift_year),
            int(shift_month),
        )

    draft = st.session_state.get("ai_shift_draft")
    if isinstance(draft, pd.DataFrame) and not draft.empty:
        draft = sanitize_ai_shift_draft_by_limits(draft, shift_df, int(shift_year), int(shift_month))
        st.session_state["ai_shift_draft"] = draft
        st.markdown("#### AIシフト案一覧（まだ保存されていません）")
        st.dataframe(draft, use_container_width=True, hide_index=True)

        warning_rows = draft[draft["勤務"].astype(str) == "注意"] if "勤務" in draft.columns else pd.DataFrame()
        if not warning_rows.empty:
            st.warning("明け翌日の休みなど、AIが自動で入れられなかった注意事項があります。保存前に確認してください。")
            for _, wr in warning_rows.iterrows():
                st.caption(f"⚠️ {wr.get('日付', '')}｜{wr.get('候補職員', '')}｜{wr.get('理由', '')}")

        preview_shift_df = apply_ai_shift_draft_to_df(shift_df, draft)
        st.markdown("#### AI案を反映した仮シフト表")
        st.caption("ここで内容を確認してから、下の保存ボタンでDBに反映します。")
        preview_matrix = create_shift_matrix(preview_shift_df, int(shift_year), int(shift_month))
        st.dataframe(preview_matrix, use_container_width=True, hide_index=True)

        preview_shortage = create_shift_shortage_table(preview_shift_df, int(shift_year), int(shift_month))
        preview_checks = create_shift_quality_check_table(preview_shift_df, int(shift_year), int(shift_month))
        preview_limit_checks = create_shift_limit_check_table(preview_matrix)

        p1, p2, p3 = st.columns(3)
        with p1:
            st.metric("仮不足日", len(preview_shortage[preview_shortage["状態"] != "OK"]))
        with p2:
            st.metric("仮確認事項", 0 if preview_checks.empty else len(preview_checks))
        with p3:
            st.metric("仮上限超過", 0 if preview_limit_checks.empty else len(preview_limit_checks))

        if not preview_shortage[preview_shortage["状態"] != "OK"].empty:
            st.warning("AI案反映後も人員不足の日があります。")
            st.dataframe(preview_shortage[preview_shortage["状態"] != "OK"], use_container_width=True, hide_index=True)
        if not preview_checks.empty:
            st.warning("AI案反映後の確認事項があります。")
            st.dataframe(preview_checks, use_container_width=True, hide_index=True)
        if not preview_limit_checks.empty:
            st.warning("AI案反映後、職員別上限を超えている箇所があります。")
            st.dataframe(preview_limit_checks, use_container_width=True, hide_index=True)

        c_save, c_clear = st.columns(2)
        with c_save:
            if st.button("確認したAIシフト案を保存する", use_container_width=True):
                saved = save_ai_shift_draft_rows(draft)
                st.success(f"AIシフト案を {saved} 件保存しました。")
                st.session_state.pop("ai_shift_draft", None)
                st.rerun()
        with c_clear:
            if st.button("AIシフト案を破棄する", use_container_width=True):
                st.session_state.pop("ai_shift_draft", None)
                st.warning("AIシフト案を破棄しました。")
                st.rerun()

    st.markdown("---")
    st.markdown("### 5. シフトチェック・集計")

    limit_checks = create_shift_limit_check_table(matrix)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        ng = shortage[shortage["状態"] != "OK"]
        st.metric("人員不足日", len(ng))
    with c2:
        high_checks = checks[checks["重要度"] == "高"] if checks is not None and not checks.empty else pd.DataFrame()
        st.metric("高重要度確認", len(high_checks))
    with c3:
        st.metric("上限超過", 0 if limit_checks.empty else len(limit_checks))
    with c4:
        st.metric("状態", "確定" if month_status.get("is_confirmed") else "作成中")

    if matrix.empty:
        st.info("この月のシフトはまだ登録されていません。")
    else:
        st.dataframe(matrix, use_container_width=True, hide_index=True)

    ng = shortage[shortage["状態"] != "OK"]
    if ng.empty:
        st.success("日勤2名・夜勤1名の基準を満たしています。")
    else:
        st.warning("日勤2名・夜勤1名の基準に満たない日があります。")
        st.dataframe(ng, use_container_width=True, hide_index=True)

    if checks is None or checks.empty:
        st.success("夜勤翌日勤務・明け翌日勤務・5連勤以上・希望休重複などの大きな確認事項はありません。")
    else:
        st.warning("シフト確認事項があります。赤・黄色相当のチェックとして確認してください。")
        st.dataframe(checks, use_container_width=True, hide_index=True)

    if limit_checks.empty:
        st.success("職員別の日勤・夜勤・合計勤務回数の上限内です。")
    else:
        st.warning("職員別の勤務回数上限を超えている箇所があります。")
        st.dataframe(limit_checks, use_container_width=True, hide_index=True)

    st.markdown("### 6. 確定・PDF・Excel出力")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if not month_status.get("is_confirmed"):
            if st.button("この月のシフトを確定する", use_container_width=True):
                if not checks.empty or not shortage[shortage["状態"] != "OK"].empty:
                    st.warning("確認事項または不足日があります。内容を確認したうえで、必要ならもう一度確定してください。")
                set_shift_month_status(int(shift_year), int(shift_month), True, current_login_user_for_shift())
                st.success("この月のシフトを確定しました。")
                st.rerun()
        else:
            if st.button("確定を解除する", use_container_width=True):
                set_shift_month_status(int(shift_year), int(shift_month), False, current_login_user_for_shift())
                st.warning("この月のシフト確定を解除しました。")
                st.rerun()

    with c2:
        if REPORTLAB_AVAILABLE:
            try:
                pdf_bytes = make_staff_shift_pdf(int(shift_year), int(shift_month))
                st.download_button(
                    "勤務表PDFをダウンロード",
                    data=pdf_bytes,
                    file_name=f"hidamari_shift_{int(shift_year)}_{int(shift_month):02d}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"シフトPDFを作成できませんでした：{e}")
    with c3:
        if OPENPYXL_AVAILABLE:
            if month_status.get("is_confirmed"):
                try:
                    excel_bytes = make_staff_shift_excel(int(shift_year), int(shift_month))
                    st.download_button(
                        "確定勤務表Excelをダウンロード",
                        data=excel_bytes,
                        file_name=f"hidamari_shift_{int(shift_year)}_{int(shift_month):02d}_confirmed.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"シフトExcelを作成できませんでした：{e}")
            else:
                st.download_button(
                    "確定勤務表Excelをダウンロード",
                    data=b"",
                    file_name=f"hidamari_shift_{int(shift_year)}_{int(shift_month):02d}_confirmed.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    disabled=True,
                    help="この月のシフトを確定するとExcel出力できます。",
                )
        else:
            st.warning("openpyxl が未導入のためExcel出力できません。")
    with c4:
        st.caption("確定前は作成中、確定後はPDF・Excelとして保存できます。Excelは編集・集計用、PDFは掲示・印刷用を想定しています。")

    st.markdown("---")
    st.markdown("### 7. 過去シフト検索・更新・削除")

    c1, c2, c3 = st.columns(3)
    with c1:
        search_start = st.date_input("検索開始", value=date(int(shift_year), int(shift_month), 1))
    with c2:
        search_end = st.date_input("検索終了", value=date(int(shift_year), int(shift_month), calendar.monthrange(int(shift_year), int(shift_month))[1]))
    with c3:
        shift_keyword = st.text_input("職員名・勤務区分・メモ検索")

    search_df = get_staff_shifts(search_start.strftime("%Y-%m-%d"), search_end.strftime("%Y-%m-%d"), shift_keyword)
    if search_df.empty:
        st.info("該当するシフトはありません。")
    else:
        st.dataframe(search_df[["id", "shift_date", "staff_name", "shift_kind", "start_time", "end_time", "next_day", "memo"]], use_container_width=True, hide_index=True)
        selected_shift_id = st.selectbox("更新・削除するシフトID", search_df["id"].tolist())
        target = search_df[search_df["id"] == selected_shift_id].iloc[0]

        with st.form("shift_update_form"):
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                u_date = st.date_input("日付", value=datetime.strptime(target["shift_date"], "%Y-%m-%d").date())
            with c2:
                staff_list_for_edit = staff_options
                current_staff = target["staff_name"] if target["staff_name"] in staff_options else ""
                u_staff = st.selectbox("職員", staff_list_for_edit, index=staff_list_for_edit.index(current_staff) if current_staff in staff_list_for_edit else 0)
            with c3:
                u_kind = st.selectbox("勤務区分", SHIFT_KINDS, index=SHIFT_KINDS.index(target["shift_kind"]) if target["shift_kind"] in SHIFT_KINDS else 0)
            with c4:
                u_next = st.checkbox("翌日終了", value=bool(target["next_day"]))
            u_start = st.text_input("開始時刻", value=target["start_time"] or "")
            u_end = st.text_input("終了時刻", value=target["end_time"] or "")
            u_memo = st.text_input("メモ", value=target["memo"] or "")
            c_update, c_delete = st.columns(2)
            with c_update:
                update_shift = st.form_submit_button("シフトを更新")
            with c_delete:
                delete_shift = st.form_submit_button("シフトを削除")
        if update_shift:
            if not u_staff:
                st.error("職員を選択してください。")
            else:
                execute("""
                    UPDATE staff_shifts
                    SET shift_date=?, staff_name=?, shift_kind=?, start_time=?, end_time=?, next_day=?, memo=?, updated_at=?
                    WHERE id=?
                """, (u_date.strftime("%Y-%m-%d"), u_staff, u_kind, u_start or None, u_end or None, 1 if u_next else 0, u_memo or None, now_text(), int(selected_shift_id)))
                st.success("シフトを更新しました。")
                st.rerun()
        if delete_shift:
            execute("DELETE FROM staff_shifts WHERE id=?", (int(selected_shift_id),))
            st.warning("シフトを削除しました。")
            st.rerun()

    st.markdown("---")
    st.markdown("### 8. カレンダー予定へのAI担当割当")

    month_start = f"{int(shift_year)}-{int(shift_month):02d}-01"
    month_end = f"{int(shift_year)}-{int(shift_month):02d}-{calendar.monthrange(int(shift_year), int(shift_month))[1]:02d}"
    events_df = fetch_df("""
        SELECT *
        FROM events
        WHERE event_date BETWEEN ? AND ?
        ORDER BY event_date, start_time, id
    """, (month_start, month_end))

    if events_df.empty:
        st.info("この月の予定はありません。")
        return

    options = {}
    for _, ev in events_df.iterrows():
        current_staff = ev["staff_name"] if ev["staff_name"] else "未担当"
        label = f"ID:{ev['id']}｜{ev['event_date']}｜{ev['start_time'] or ''}｜{ev['title']}｜担当:{current_staff}"
        options[label] = int(ev["id"])

    selected_event_label = st.selectbox("担当候補を見る予定", list(options.keys()))
    selected_event_id = options[selected_event_label]
    event_row = events_df[events_df["id"] == selected_event_id].iloc[0]

    candidates = get_shift_candidates_for_event(event_row, shift_df)
    if candidates.empty:
        st.warning("勤務表から担当候補が見つかりません。")
    else:
        st.dataframe(candidates, use_container_width=True, hide_index=True)
        c1, c2 = st.columns(2)
        with c1:
            ai_staff = st.selectbox("AI候補から選ぶ", candidates["職員名"].tolist())
            if st.button("AI候補を担当に反映", use_container_width=True):
                execute("UPDATE events SET staff_name=?, updated_at=? WHERE id=?", (ai_staff, now_text(), int(selected_event_id)))
                st.success(f"予定ID:{selected_event_id} の担当を {ai_staff} さんにしました。")
                st.rerun()
        with c2:
            manual_staff = st.selectbox("自分で担当を選ぶ", staff_options, key="manual_assign_staff")
            if st.button("自分で選んだ担当を反映", use_container_width=True):
                execute("UPDATE events SET staff_name=?, updated_at=? WHERE id=?", (manual_staff or None, now_text(), int(selected_event_id)))
                st.success("担当を更新しました。")
                st.rerun()

    st.markdown("#### 未担当予定の一括AI割当")
    preview = build_event_assignment_preview(events_df, shift_df, only_unassigned=True)
    if preview.empty:
        st.info("未担当予定はありません。")
    else:
        st.dataframe(preview, use_container_width=True, hide_index=True)
        assignable = preview[preview["AI候補"].fillna("") != ""]
        if not assignable.empty:
            if st.button("未担当予定へ第1候補を一括反映", use_container_width=True):
                params = [(r["AI候補"], now_text(), int(r["予定ID"])) for _, r in assignable.iterrows()]
                updated = execute_many("UPDATE events SET staff_name=?, updated_at=? WHERE id=?", params)
                st.success(f"{updated}件の予定へ担当候補を反映しました。")
                st.rerun()


def current_login_user_for_shift():
    """シフト確定者表示用。ログイン機構がない場合も止めない。"""
    for key in ["username", "user_id", "login_user", "user"]:
        try:
            value = st.session_state.get(key)
            if value:
                return str(value)
        except Exception:
            pass
    return "system"

def page_storage_check():
    st.subheader("保存状態チェック")
    st.caption("PostgreSQL・Supabase Storage・DB上のファイル紐づけを確認します。")

    # -----------------------------
    # 接続・設定状態
    # -----------------------------
    st.markdown("### 1. Storage設定の状態")

    storage_ok = storage_is_configured()
    bucket_ok, bucket_msg = check_storage_bucket_access()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("PostgreSQL", "接続中")
    with c2:
        st.metric("Storage設定", "OK" if storage_ok else "未設定")
    with c3:
        st.metric("Storageバケット", "OK" if bucket_ok else "要確認")
    with c4:
        st.metric("requests", "OK" if requests is not None else "未導入")

    st.write(f"**SUPABASE_URL**：`{get_supabase_url() or '未設定'}`")
    st.write(f"**SUPABASE_STORAGE_BUCKET**：`{get_supabase_storage_bucket() or '未設定'}`")
    st.write(f"**SUPABASE_SERVICE_ROLE_KEY**：`{'設定あり' if bool(get_supabase_storage_key()) else '未設定'}`")

    if bucket_ok:
        st.success(bucket_msg)
    else:
        st.warning(bucket_msg)

    # -----------------------------
    # 件数サマリー
    # -----------------------------
    st.markdown("---")
    st.markdown("### 2. 保存件数")

    event_count = count_query("SELECT COUNT(*) FROM events")
    photo_count = count_query("SELECT COUNT(*) FROM event_photos")
    file_count = count_query("SELECT COUNT(*) FROM event_files")
    storage_photo_count = count_query("SELECT COUNT(*) FROM event_photos WHERE file_path LIKE ?", ("storage://%",))
    storage_file_count = count_query("SELECT COUNT(*) FROM event_files WHERE file_path LIKE ?", ("storage://%",))
    local_photo_count = photo_count - storage_photo_count
    local_file_count = file_count - storage_file_count

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("予定件数", event_count)
    with c2:
        st.metric("写真メモ件数", photo_count)
    with c3:
        st.metric("添付ファイル件数", file_count)
    with c4:
        st.metric("Storage保存パス件数", storage_photo_count + storage_file_count)

    c5, c6 = st.columns(2)
    with c5:
        st.info(f"写真メモ：Storage {storage_photo_count}件 / 旧ローカル {local_photo_count}件")
    with c6:
        st.info(f"添付ファイル：Storage {storage_file_count}件 / 旧ローカル {local_file_count}件")

    # -----------------------------
    # 最新登録10件
    # -----------------------------
    st.markdown("---")
    st.markdown("### 3. 最新登録10件")

    latest_df = fetch_df("""
        SELECT
            e.id,
            e.event_date,
            e.category,
            e.title,
            e.user_name,
            e.staff_name,
            e.created_at,
            (SELECT COUNT(*) FROM event_photos p WHERE p.event_id = e.id) AS photo_count,
            (SELECT COUNT(*) FROM event_files f WHERE f.event_id = e.id) AS file_count
        FROM events e
        ORDER BY e.id DESC
        LIMIT 10
    """)
    if latest_df.empty:
        st.info("予定はまだ登録されていません。")
    else:
        st.dataframe(latest_df, use_container_width=True, hide_index=True)

    # -----------------------------
    # DBにはあるが読めないファイルの確認
    # -----------------------------
    st.markdown("---")
    st.markdown("### 4. DBにはあるがStorage・ローカルから読めないファイル")

    check_limit = st.number_input("確認件数上限", min_value=1, max_value=500, value=100, step=10)
    st.caption("Storage確認は通信が発生するため、件数が多い場合は少し時間がかかります。")

    if st.button("ファイル紐づけを確認する", use_container_width=True):
        targets = []

        photos = fetch_df("""
            SELECT
                '写真メモ' AS 種別,
                p.id AS id,
                p.event_id,
                p.file_name,
                p.file_path,
                p.created_at,
                e.event_date,
                e.title
            FROM event_photos p
            LEFT JOIN events e ON p.event_id = e.id
            ORDER BY p.id DESC
            LIMIT ?
        """, (int(check_limit),))

        files = fetch_df("""
            SELECT
                '添付ファイル' AS 種別,
                f.id AS id,
                f.event_id,
                f.file_name,
                f.file_path,
                f.created_at,
                e.event_date,
                e.title
            FROM event_files f
            LEFT JOIN events e ON f.event_id = e.id
            ORDER BY f.id DESC
            LIMIT ?
        """, (int(check_limit),))

        if not photos.empty:
            targets.extend(photos.to_dict("records"))
        if not files.empty:
            targets.extend(files.to_dict("records"))

        if not targets:
            st.info("確認対象の写真・添付ファイルはありません。")
        else:
            results = []
            progress = st.progress(0)
            for i, item in enumerate(targets, start=1):
                ok, message = storage_object_exists(item.get("file_path"))
                results.append({
                    "状態": "OK" if ok else "NG",
                    "種別": item.get("種別"),
                    "ID": item.get("id"),
                    "予定ID": item.get("event_id"),
                    "予定日": item.get("event_date"),
                    "予定タイトル": item.get("title"),
                    "ファイル名": item.get("file_name"),
                    "file_path": item.get("file_path"),
                    "確認結果": message,
                    "登録日時": item.get("created_at"),
                })
                progress.progress(i / len(targets))

            result_df = pd.DataFrame(results)
            ng_df = result_df[result_df["状態"] == "NG"].copy()

            st.metric("確認件数", len(result_df))
            st.metric("読めないファイル件数", len(ng_df))

            if ng_df.empty:
                st.success("確認した範囲では、DBに紐づいたファイルはすべて読み取り可能でした。")
            else:
                st.warning("DBには登録されていますが、Storage・ローカルから読めないファイルがあります。")
                st.dataframe(ng_df, use_container_width=True, hide_index=True)

            with st.expander("確認結果の全件を見る"):
                st.dataframe(result_df, use_container_width=True, hide_index=True)

    # -----------------------------
    # Storageパス一覧
    # -----------------------------
    st.markdown("---")
    st.markdown("### 5. Storage保存パス一覧")

    path_df = fetch_df("""
        SELECT '写真メモ' AS 種別, id, event_id, file_name, file_path, created_at
        FROM event_photos
        WHERE file_path LIKE ?
        UNION ALL
        SELECT '添付ファイル' AS 種別, id, event_id, file_name, file_path, created_at
        FROM event_files
        WHERE file_path LIKE ?
        ORDER BY created_at DESC
        LIMIT 50
    """, ("storage://%", "storage://%"))
    if path_df.empty:
        st.info("Storage保存パスはまだ登録されていません。")
    else:
        st.dataframe(path_df, use_container_width=True, hide_index=True)




def page_export():
    st.subheader("Excel・PDF出力")

    st.markdown("### Excel出力")
    events = fetch_df("SELECT * FROM events ORDER BY event_date, start_time, id")
    photos = fetch_df("SELECT * FROM event_photos ORDER BY event_id, id")
    files = fetch_df("SELECT * FROM event_files ORDER BY event_id, id")
    categories = fetch_df("SELECT * FROM categories ORDER BY sort_order, category_name")
    users = fetch_df("SELECT * FROM users ORDER BY user_name")
    staff = fetch_df("SELECT * FROM staff ORDER BY staff_name")

    output = Path("hidamari_calendar_export.xlsx")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        events.to_excel(writer, sheet_name="予定", index=False)
        photos.to_excel(writer, sheet_name="写真メモ", index=False)
        files.to_excel(writer, sheet_name="添付ファイル", index=False)
        categories.to_excel(writer, sheet_name="カテゴリマスタ", index=False)
        users.to_excel(writer, sheet_name="利用者マスタ", index=False)
        staff.to_excel(writer, sheet_name="職員マスタ", index=False)

    with open(output, "rb") as f:
        st.download_button(
            label="Excelをダウンロード",
            data=f,
            file_name="hidamari_calendar_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.caption("予定・利用者マスタ・職員マスタをまとめてExcel出力します。")

    st.markdown("---")
    st.markdown("### PDFカレンダー出力")
    st.caption("A4横の月間カレンダーPDFと、予定詳細一覧を出力します。掲示・申し送り・印刷用に使えます。")

    today = today_jst()
    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        pdf_year = st.number_input("PDF出力 年", min_value=2020, max_value=2100, value=today.year, step=1)
    with c2:
        pdf_month = st.number_input("PDF出力 月", min_value=1, max_value=12, value=today.month, step=1)
    with c3:
        include_detail = st.checkbox("予定詳細一覧も付ける", value=True)

    if not REPORTLAB_AVAILABLE:
        st.error("PDF出力には reportlab が必要です。requirements.txt に reportlab を追加してください。")
    else:
        try:
            pdf_bytes = make_calendar_pdf(int(pdf_year), int(pdf_month), include_detail=include_detail)
            st.download_button(
                label="PDFカレンダーをダウンロード",
                data=pdf_bytes,
                file_name=f"hidamari_calendar_{int(pdf_year)}_{int(pdf_month):02d}.pdf",
                mime="application/pdf",
            )
        except Exception as e:
            st.error(f"PDFを作成できませんでした：{e}")




def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="📅", layout="wide")
    init_db_once()
    add_css()

    st.title("📅 ひだまり帳 Ver1.4.7 PostgreSQL版")
    st.caption("紙の壁カレンダー感覚で、通院・面会・行事・注意事項を一枚で")

    menu = st.sidebar.radio(
        "メニュー",
        [
            "月間カレンダー",
            "今日は何ある",
            "予定登録",
            "予定検索・更新・削除",
            "予定データ取込",
            "写真メモ一覧",
            "Excel・書類ファイル一覧",
            "シフト管理・AI割当",
            "保存状態チェック",
            "予定カテゴリ設定",
            "利用者マスタ",
            "職員マスタ",
            "Excel・PDF出力",
        ],
    )

    if menu == "月間カレンダー":
        page_calendar()
    elif menu == "今日は何ある":
        page_today()
    elif menu == "予定登録":
        page_event_register()
    elif menu == "予定検索・更新・削除":
        page_event_manage()
    elif menu == "予定データ取込":
        page_schedule_import()
    elif menu == "写真メモ一覧":
        page_photo_notes()
    elif menu == "Excel・書類ファイル一覧":
        page_attached_files()
    elif menu == "シフト管理・AI割当":
        page_shift_manager()
    elif menu == "保存状態チェック":
        page_storage_check()
    elif menu == "予定カテゴリ設定":
        page_category_master()
    elif menu == "利用者マスタ":
        page_master_users()
    elif menu == "職員マスタ":
        page_master_staff()
    elif menu == "Excel・PDF出力":
        page_export()


if __name__ == "__main__":
    main()