import calendar
import sys
import types
import unittest
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

fake_streamlit = types.ModuleType("streamlit")
fake_streamlit.secrets = {}
fake_streamlit.cache_data = lambda func=None, **_kwargs: func if func else (lambda target: target)
sys.modules.setdefault("streamlit", fake_streamlit)

from report_service import make_staff_shift_excel


SUMMARY_COLUMNS = ["日勤", "管", "夜勤", "明", "休み", "希望休", "有休", "他", "合計", "最大連勤"]
WEEKDAY_LABELS = "月火水木金土日"


def _matrix_for_month(year, month):
    last_day = calendar.monthrange(year, month)[1]
    row = {"職員名": "テスト職員"}
    row.update({str(day): "" for day in range(1, last_day + 1)})
    row[str(last_day)] = "夜"
    row.update({column: 0 for column in SUMMARY_COLUMNS})
    row["夜勤"] = 1
    row["合計"] = 1
    columns = ["職員名"] + [str(day) for day in range(1, last_day + 1)] + SUMMARY_COLUMNS
    return pd.DataFrame([row], columns=columns)


def _make_workbook(year, month):
    matrix = _matrix_for_month(year, month)
    empty = pd.DataFrame()
    content = make_staff_shift_excel(
        year,
        month,
        lambda *_args, **_kwargs: empty,
        lambda *_args, **_kwargs: matrix,
        lambda *_args, **_kwargs: empty,
        lambda *_args, **_kwargs: empty,
        lambda *_args, **_kwargs: empty,
        lambda *_args, **_kwargs: {"is_confirmed": True, "confirmed_at": "2026-07-20"},
    )
    return load_workbook(BytesIO(content))


class StaffShiftExcelTest(unittest.TestCase):
    def test_all_date_headers_have_the_correct_weekday_and_data_alignment(self):
        for year, month in [(2024, 2), (2025, 2), (2026, 4), (2026, 7)]:
            with self.subTest(year=year, month=month):
                ws = _make_workbook(year, month)["勤務表"]
                last_day = calendar.monthrange(year, month)[1]

                for day in range(1, last_day + 1):
                    column = day + 1
                    self.assertEqual(ws.cell(5, column).value, str(day))
                    self.assertEqual(
                        ws.cell(6, column).value,
                        WEEKDAY_LABELS[date(year, month, day).weekday()],
                    )

                self.assertEqual(ws.cell(7, 1).value, "テスト職員")
                self.assertEqual(ws.cell(7, last_day + 1).value, "夜")
                self.assertEqual(ws.cell(7, last_day + 4).value, 1)  # 夜勤集計
                self.assertEqual(ws.cell(7, last_day + 10).value, 1)  # 合計
                self.assertEqual(ws.freeze_panes, "B7")
                self.assertEqual(ws.column_dimensions["B"].width, 4.0)

    def test_weekend_header_style_is_applied_to_date_and_weekday_rows(self):
        ws = _make_workbook(2024, 2)["勤務表"]

        for row in (5, 6):
            self.assertEqual(ws.cell(row, 4).fill.fgColor.rgb, "00DDEBF7")  # 2月3日（土）
            self.assertEqual(ws.cell(row, 5).fill.fgColor.rgb, "00FCE4D6")  # 2月4日（日）
            self.assertEqual(ws.cell(row, 4).alignment.horizontal, "center")
            self.assertEqual(ws.cell(row, 4).border.bottom.style, "thin")

    def test_non_date_headers_span_two_rows_and_filter_covers_shift_data(self):
        ws = _make_workbook(2024, 2)["勤務表"]

        self.assertIn("A5:A6", {str(cell_range) for cell_range in ws.merged_cells.ranges})
        self.assertIn("AE5:AE6", {str(cell_range) for cell_range in ws.merged_cells.ranges})
        self.assertEqual(ws.auto_filter.ref, "A5:AN7")
        self.assertEqual(ws.print_area, "")


if __name__ == "__main__":
    unittest.main()
