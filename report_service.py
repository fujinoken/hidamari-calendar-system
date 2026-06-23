# -*- coding: utf-8 -*-
"""PDF, Excel, and CSV export helpers for ひだまり帳."""
import calendar
import re
from datetime import date, datetime
from io import BytesIO

import pandas as pd

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OpenpyxlWorkbook = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    OPENPYXL_AVAILABLE = False

from config import (
    KOT_DAY_PATTERN_CODE_KEY,
    KOT_HOPE_REST_LEAVE_NAME_KEY,
    KOT_NIGHT_PATTERN_CODE_KEY,
    KOT_PAID_LEAVE_NAME_KEY,
    KOT_REST_LEAVE_NAME_KEY,
    PDF_FONT_GOTHIC,
    PDF_FONT_MINCHO,
)
from db import fetch_df, today_jst
from storage_service import get_secret_or_env

def init_pdf_fonts():
    """ReportLabの日本語CIDフォントを登録する。"""
    if not REPORTLAB_AVAILABLE:
        return False
    try:
        if PDF_FONT_GOTHIC not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT_GOTHIC))
        if PDF_FONT_MINCHO not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(UnicodeCIDFont(PDF_FONT_MINCHO))
        return True
    except Exception:
        return False


def pdf_text(value):
    """PDF描画用にNone/nanを空文字へ寄せる。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def text_width(text, font_name, font_size):
    """ReportLab上の文字幅を取得する。"""
    try:
        return pdfmetrics.stringWidth(str(text), font_name, font_size)
    except Exception:
        return len(str(text)) * font_size


def fit_pdf_text(text, max_width, font_name=PDF_FONT_GOTHIC, font_size=8, ellipsis="…"):
    """
    指定幅に収まるよう、1行表示用の文字列へ整える。
    改行・連続空白を1つの空白へ寄せ、長い場合は末尾を…で省略する。
    """
    s = re.sub(r"\s+", " ", pdf_text(text)).strip()
    if not s:
        return ""
    if text_width(s, font_name, font_size) <= max_width:
        return s

    # 省略記号だけでも幅を超える場合
    if text_width(ellipsis, font_name, font_size) > max_width:
        return ""

    low, high = 0, len(s)
    best = ""
    while low <= high:
        mid = (low + high) // 2
        candidate = s[:mid] + ellipsis
        if text_width(candidate, font_name, font_size) <= max_width:
            best = candidate
            low = mid + 1
        else:
            high = mid - 1
    return best or ellipsis


def wrap_pdf_text(text, max_width, font_name, font_size, max_lines=3):
    """
    日本語を含む文字列を、文字単位で指定幅に折り返す。
    長すぎる場合は最終行を…で省略する。
    """
    text = pdf_text(text).replace("\r", "\n")
    if not text:
        return []

    lines = []
    for raw in text.splitlines():
        raw = raw.strip()
        if not raw:
            continue
        current = ""
        for ch in raw:
            candidate = current + ch
            if text_width(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = ch
                if len(lines) >= max_lines:
                    break
        if current and len(lines) < max_lines:
            lines.append(current)
        if len(lines) >= max_lines:
            break

    if len(lines) > max_lines:
        lines = lines[:max_lines]

    if lines:
        original_joined = "".join([ln for ln in text.splitlines() if ln.strip()])
        wrapped_joined = "".join(lines)
        if len(wrapped_joined) < len(original_joined):
            last = lines[-1]
            while last and text_width(last + "…", font_name, font_size) > max_width:
                last = last[:-1]
            lines[-1] = (last + "…") if last else "…"

    return lines


def draw_wrapped_text(c, text, x, y, max_width, font_name, font_size, leading, max_lines=3, color=None):
    """折り返しテキストを描画し、描画後のy座標を返す。"""
    if color is not None:
        c.setFillColor(color)
    c.setFont(font_name, font_size)
    lines = wrap_pdf_text(text, max_width, font_name, font_size, max_lines=max_lines)
    for line in lines:
        c.drawString(x, y, line)
        y -= leading
    c.setFillColor(colors.black)
    return y


def draw_single_line(c, text, x, y, max_width, font_name=PDF_FONT_GOTHIC, font_size=8, color=None):
    """
    PDF上に1行だけ描画する。
    幅を超える文字列は自動で…省略するため、次の項目と重ならない。
    """
    if color is not None:
        c.setFillColor(color)
    c.setFont(font_name, font_size)
    c.drawString(x, y, fit_pdf_text(text, max_width, font_name, font_size))
    c.setFillColor(colors.black)


def calendar_events_df(year, month):
    """指定月の予定をDataFrameで取得する。"""
    start = f"{int(year)}-{int(month):02d}-01"
    last_day = calendar.monthrange(int(year), int(month))[1]
    end = f"{int(year)}-{int(month):02d}-{last_day:02d}"
    return fetch_df("""
        SELECT
            id, event_date, category, title, user_id, user_name, staff_name,
            start_time, end_time, memo, important, created_at, updated_at
        FROM events
        WHERE event_date BETWEEN ? AND ?
        ORDER BY event_date,
            CASE WHEN start_time IS NULL OR start_time = '' THEN 1 ELSE 0 END,
            start_time,
            category,
            id
    """, (start, end))


def make_event_summary_for_pdf(ev, compact=True):
    """カレンダー枠内に表示する予定の短い説明を作る。"""
    category_name = re.sub(r"\s+", " ", pdf_text(ev.get("category", ""))).strip()
    time_part = re.sub(r"\s+", " ", pdf_text(ev.get("start_time", ""))).strip()
    title = re.sub(r"\s+", " ", pdf_text(ev.get("title", ""))).strip()
    user_name = re.sub(r"\s+", " ", pdf_text(ev.get("user_name", ""))).strip()
    staff_name = re.sub(r"\s+", " ", pdf_text(ev.get("staff_name", ""))).strip()
    important = "★" if int(ev.get("important", 0) or 0) == 1 else ""

    if compact:
        # 月間カレンダー枠内は重なり防止を最優先し、情報量を絞る。
        # 利用者・担当・長い詳細は次ページの予定詳細一覧で確認する。
        parts = []
        if time_part:
            parts.append(time_part)

        label_core = f"【{category_name}】" if category_name else ""
        if title:
            label_core += title
        elif user_name:
            label_core += user_name

        if important:
            label_core = f"{important}{label_core}"

        if label_core:
            parts.append(label_core)

        return " ".join([p for p in parts if p]).strip()

    parts = []
    if time_part:
        parts.append(time_part)
    label_core = f"【{category_name}】"
    if important:
        label_core += important
    if title:
        label_core += title
    parts.append(label_core)
    if user_name:
        parts.append(f"利用者:{user_name}")
    if staff_name:
        parts.append(f"担当:{staff_name}")
    return " ".join([p for p in parts if p]).strip()


def draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=False):
    """予定詳細一覧ページのヘッダーを描画する。"""
    suffix = "（続き）" if continuation else ""
    c.setFont(PDF_FONT_GOTHIC, 16)
    c.setFillColor(colors.HexColor("#333333"))
    c.drawString(margin_x, height - 34, f"予定詳細一覧　{year}年 {month}月{suffix}")
    c.setFont(PDF_FONT_GOTHIC, 9)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawRightString(width - margin_x, height - 32, f"合計 {total_count} 件")
    c.setFillColor(colors.black)


def draw_pdf_detail_table_header(c, x, y, widths, row_h):
    """予定詳細一覧の列見出しを描画する。"""
    headers = ["時間", "分類", "予定", "利用者", "担当", "詳細・メモ"]
    c.setFillColor(colors.HexColor("#eee7dc"))
    c.rect(x, y - 4, sum(widths), row_h, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#333333"))
    c.setFont(PDF_FONT_GOTHIC, 8)
    cur_x = x
    for header, w in zip(headers, widths):
        c.drawString(cur_x + 4, y + 1, header)
        cur_x += w
    c.setFillColor(colors.black)
    return y - row_h


def make_calendar_pdf(year, month, include_detail=True):
    """
    A4横の月間カレンダーPDFを作る。
    1ページ目: 月間カレンダー
    2ページ目以降: 予定詳細一覧

    Ver1.3.3相当の修正:
    ・予定詳細一覧を1予定1行の表形式へ変更
    ・詳細・メモも1行に収め、長い場合は…で省略
    ・カレンダー枠内の予定文字が重ならないよう、表示行数・文字量・余白を再調整
    ・月間カレンダー枠内は要約表示に絞り、あふれる予定は「ほか◯件」で表示
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab がインストールされていません。requirements.txt に reportlab を追加してください。")

    init_pdf_fonts()

    year = int(year)
    month = int(month)
    df = calendar_events_df(year, month)

    events_by_day = {}
    if not df.empty:
        for _, row in df.iterrows():
            events_by_day.setdefault(pdf_text(row["event_date"]), []).append(row.to_dict())

    buffer = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    margin_x = 24
    margin_top = 28
    title_h = 40
    grid_x = margin_x
    grid_y_top = height - margin_top - title_h
    grid_w = width - margin_x * 2
    grid_h = height - margin_top - title_h - 24
    day_header_h = 22
    cell_w = grid_w / 7
    cell_h = (grid_h - day_header_h) / 6

    # タイトル
    c.setFont(PDF_FONT_GOTHIC, 18)
    c.setFillColor(colors.HexColor("#333333"))
    c.drawString(margin_x, height - 34, f"ひだまり帳 月間カレンダー　{year}年 {month}月")
    c.setFont(PDF_FONT_GOTHIC, 9)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawRightString(width - margin_x, height - 32, f"出力日: {today_jst().strftime('%Y-%m-%d')}")
    c.setFillColor(colors.black)

    # 曜日ヘッダ
    week_labels = ["日", "月", "火", "水", "木", "金", "土"]
    header_y = grid_y_top - day_header_h
    for i, label in enumerate(week_labels):
        x = grid_x + i * cell_w
        c.setFillColor(colors.HexColor("#f3eee6"))
        c.rect(x, header_y, cell_w, day_header_h, fill=1, stroke=1)
        c.setFillColor(colors.HexColor("#333333"))
        if i == 0:
            c.setFillColor(colors.HexColor("#c0392b"))
        elif i == 6:
            c.setFillColor(colors.HexColor("#1f4e79"))
        c.setFont(PDF_FONT_GOTHIC, 10)
        c.drawCentredString(x + cell_w / 2, header_y + 7, label)

    # 月間グリッド
    weeks = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
    while len(weeks) < 6:
        weeks.append([0, 0, 0, 0, 0, 0, 0])

    for r, week in enumerate(weeks[:6]):
        for col, day in enumerate(week):
            x = grid_x + col * cell_w
            y = header_y - (r + 1) * cell_h

            c.setFillColor(colors.HexColor("#fffdf8") if day else colors.HexColor("#fafafa"))
            c.rect(x, y, cell_w, cell_h, fill=1, stroke=1)

            if not day:
                continue

            # 日付を少し小さくし、予定表示エリアを広げる
            date_color = colors.HexColor("#333333")
            if col == 0:
                date_color = colors.HexColor("#c0392b")
            elif col == 6:
                date_color = colors.HexColor("#1f4e79")
            c.setFillColor(date_color)
            c.setFont(PDF_FONT_GOTHIC, 15)
            c.drawString(x + 6, y + cell_h - 18, str(day))

            key = f"{year}-{month:02d}-{day:02d}"
            evs = events_by_day.get(key, [])
            if not evs:
                continue

            # 枠内で文字が重ならないよう、予定表示エリアを固定管理する
            left_pad = 5
            right_pad = 5
            top_after_date = 28
            bottom_reserved = 10
            band_h = 10
            row_gap = 2
            row_pitch = band_h + row_gap
            event_area_top = y + cell_h - top_after_date
            event_area_bottom = y + bottom_reserved
            usable_h = max(0, event_area_top - event_area_bottom)
            max_lines_total = max(1, int(usable_h // row_pitch))
            max_lines_total = min(max_lines_total, 4)

            visible_events = evs[:max_lines_total]
            for idx, ev in enumerate(visible_events):
                summary = make_event_summary_for_pdf(ev, compact=True)
                important = int(ev.get("important", 0) or 0) == 1

                band_top = event_area_top - idx * row_pitch
                band_y = band_top - band_h
                band_w = cell_w - left_pad - right_pad

                c.setFillColor(colors.HexColor("#ffe9e0") if important else colors.HexColor("#f6efe6"))
                c.roundRect(x + left_pad, band_y, band_w, band_h, 2.5, fill=1, stroke=0)

                text_color = colors.HexColor("#8a2d18") if important else colors.HexColor("#3f3a35")
                draw_single_line(
                    c,
                    summary,
                    x + left_pad + 2,
                    band_y + 2.1,
                    band_w - 4,
                    PDF_FONT_GOTHIC,
                    6.0,
                    color=text_color,
                )

            remaining = len(evs) - len(visible_events)
            if remaining > 0:
                c.setFont(PDF_FONT_GOTHIC, 6.3)
                c.setFillColor(colors.HexColor("#555555"))
                c.drawString(x + 7, y + 5, f"ほか {remaining} 件")
                c.setFillColor(colors.black)

    # 凡例
    c.setFont(PDF_FONT_GOTHIC, 8)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawString(margin_x, 12, "※枠内は要約表示です。詳細は次ページ以降の一覧で確認できます。重要予定は薄赤で表示します。")
    c.setFillColor(colors.black)

    # 詳細一覧
    if include_detail:
        c.showPage()
        c.setPageSize(page_size)

        total_count = 0 if df.empty else len(df)
        draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=False)

        detail_x = margin_x
        detail_w = width - margin_x * 2
        row_h = 17
        date_h = 18
        bottom_y = 34

        # 横幅に収まる固定列幅。最後の「詳細・メモ」は残り幅を使う。
        col_widths = [58, 54, 160, 78, 78]
        memo_w = detail_w - sum(col_widths)
        col_widths.append(memo_w)

        y = height - 62
        y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)

        if df.empty:
            c.setFont(PDF_FONT_GOTHIC, 12)
            c.setFillColor(colors.HexColor("#333333"))
            c.drawString(detail_x, y - 8, "この月の予定はありません。")
        else:
            current_date = ""
            row_index = 0

            for _, row in df.iterrows():
                event_date = pdf_text(row["event_date"])

                # 日付見出し＋1行分が入らなければ改ページ
                if y < bottom_y + date_h + row_h + 8:
                    c.showPage()
                    c.setPageSize(page_size)
                    draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=True)
                    y = height - 62
                    y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)
                    current_date = ""

                if event_date != current_date:
                    current_date = event_date
                    try:
                        d = datetime.strptime(event_date, "%Y-%m-%d").date()
                        dow = ["月", "火", "水", "木", "金", "土", "日"][d.weekday()]
                        date_label = f"{d.month}/{d.day}（{dow}）"
                    except Exception:
                        date_label = event_date

                    c.setFillColor(colors.HexColor("#f7f1e8"))
                    c.rect(detail_x, y - 3, detail_w, date_h, fill=1, stroke=0)
                    c.setFillColor(colors.HexColor("#333333"))
                    c.setFont(PDF_FONT_GOTHIC, 9)
                    c.drawString(detail_x + 6, y + 2, date_label)
                    c.setFillColor(colors.black)
                    y -= date_h

                # 1行分が入らなければ改ページ
                if y < bottom_y + row_h:
                    c.showPage()
                    c.setPageSize(page_size)
                    draw_pdf_detail_page_header(c, width, height, margin_x, year, month, total_count, continuation=True)
                    y = height - 62
                    y = draw_pdf_detail_table_header(c, detail_x, y, col_widths, row_h)

                important = int(row.get("important", 0) or 0) == 1
                category_name = pdf_text(row.get("category", ""))

                start_time = pdf_text(row.get("start_time", ""))
                end_time = pdf_text(row.get("end_time", ""))
                if start_time and end_time:
                    time_text = f"{start_time}〜{end_time}"
                elif start_time:
                    time_text = start_time
                elif end_time:
                    time_text = f"〜{end_time}"
                else:
                    time_text = "時間未定"

                title_text = pdf_text(row.get("title", ""))
                user_text = pdf_text(row.get("user_name", ""))
                staff_text = pdf_text(row.get("staff_name", ""))
                memo_text = pdf_text(row.get("memo", ""))
                if important:
                    title_text = "重要 " + title_text

                # 行背景
                bg = "#fffdf8" if row_index % 2 == 0 else "#faf6ef"
                if important:
                    bg = "#fff0e8"
                c.setFillColor(colors.HexColor(bg))
                c.rect(detail_x, y - 3, detail_w, row_h, fill=1, stroke=0)

                # 下線
                c.setStrokeColor(colors.HexColor("#e2d8cc"))
                c.line(detail_x, y - 3, detail_x + detail_w, y - 3)
                c.setStrokeColor(colors.black)

                # 各列を1行で描画。幅を超えたら…で省略する。
                values = [time_text, category_name, title_text, user_text, staff_text, memo_text]
                cur_x = detail_x
                text_color = colors.HexColor("#8a2d18") if important else colors.HexColor("#222222")
                for value, w in zip(values, col_widths):
                    draw_single_line(
                        c, value, cur_x + 4, y + 1, max(w - 8, 10),
                        PDF_FONT_GOTHIC, 8, color=text_color
                    )
                    cur_x += w

                y -= row_h
                row_index += 1

    c.save()
    buffer.seek(0)
    return buffer.getvalue()





# -----------------------------
# シフト管理・AI担当割当


def make_staff_shift_pdf(
    year,
    month,
    get_staff_shifts_month,
    create_shift_matrix,
    create_shift_shortage_table,
    create_shift_quality_check_table,
    create_shift_limit_check_table,
    get_shift_month_status,
):
    """掲示・印刷しやすい横長のシフトPDFを作成する。"""
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab がインストールされていません。")

    init_pdf_fonts()

    df = get_staff_shifts_month(int(year), int(month), include_prev_day=True)
    matrix = create_shift_matrix(df, int(year), int(month))
    shortage = create_shift_shortage_table(df, int(year), int(month))
    checks = create_shift_quality_check_table(df, int(year), int(month))
    limit_checks = create_shift_limit_check_table(matrix)
    status = get_shift_month_status(int(year), int(month))

    buffer = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    margin = 14
    title_y = height - 24
    c.setFont(PDF_FONT_GOTHIC, 14.5)
    c.drawString(margin, title_y, f"従業員勤務表　{int(year)}年{int(month)}月")
    c.setFont(PDF_FONT_GOTHIC, 8)
    status_text = "確定" if status.get("is_confirmed") else "作成中"
    c.drawRightString(width - margin, title_y, f"{status_text}　出力日: {today_jst().strftime('%Y-%m-%d')}")

    last_day = calendar.monthrange(int(year), int(month))[1]
    table_x = margin
    table_y_top = height - 48
    staff_w = 58
    summary_labels = ["日", "管", "夜", "明", "休", "希", "有", "他", "計"]
    summary_keys = ["日勤", "管", "夜勤", "明", "休み", "希望休", "有休", "他", "合計"]
    summary_w = 18
    day_w = (width - margin * 2 - staff_w - summary_w * len(summary_labels)) / last_day
    row_h = 15.2

    def draw_shift_table_vertical_lines(y0, y1):
        """日付ごとの縦罫線を見やすく引く。週境界と集計欄は少し強調する。"""
        old_stroke = getattr(c, "_strokeColorObj", colors.black)
        old_width = getattr(c, "_lineWidth", 1)

        # 氏名欄と日付欄の境界
        x_line = table_x + staff_w
        c.setStrokeColor(colors.HexColor("#8f8f8f"))
        c.setLineWidth(0.55)
        c.line(x_line, y0, x_line, y1)

        # 日付欄の縦罫線
        for d in range(1, last_day + 1):
            x_line = table_x + staff_w + day_w * d
            if d % 7 == 0:
                c.setStrokeColor(colors.HexColor("#8f8f8f"))
                c.setLineWidth(0.45)
            else:
                c.setStrokeColor(colors.HexColor("#c9c9c9"))
                c.setLineWidth(0.25)
            c.line(x_line, y0, x_line, y1)

        # 集計欄の縦罫線
        summary_start = table_x + staff_w + day_w * last_day
        c.setStrokeColor(colors.HexColor("#8f8f8f"))
        c.setLineWidth(0.45)
        c.line(summary_start, y0, summary_start, y1)
        for i in range(1, len(summary_labels)):
            x_line = summary_start + summary_w * i
            c.line(x_line, y0, x_line, y1)

        c.setStrokeColor(old_stroke)
        c.setLineWidth(old_width)

    # 凡例
    c.setFont(PDF_FONT_GOTHIC, 7.2)
    c.drawString(margin, height - 38, "凡例：日=日勤 8:30〜17:30　管=管理業務 8:30〜17:30　夜=夜勤 16:30〜翌9:30　明=夜勤明け　希=希望休　有=有休　他=その他")

    # ヘッダ
    c.setFillColor(colors.HexColor("#f3eee6"))
    c.rect(table_x, table_y_top - row_h, width - margin * 2, row_h, fill=1, stroke=1)
    draw_shift_table_vertical_lines(table_y_top - row_h, table_y_top)
    c.setFillColor(colors.black)
    c.setFont(PDF_FONT_GOTHIC, 6.4)
    c.drawString(table_x + 3, table_y_top - 10, "氏名")
    x = table_x + staff_w
    for d in range(1, last_day + 1):
        c.drawCentredString(x + day_w / 2, table_y_top - 10, str(d))
        x += day_w
    for label in summary_labels:
        c.drawCentredString(x + summary_w / 2, table_y_top - 10, label)
        x += summary_w

    y = table_y_top - row_h
    if matrix.empty:
        c.setFont(PDF_FONT_GOTHIC, 10)
        c.drawString(margin, y - 24, "この月のシフトは登録されていません。")
    else:
        c.setFont(PDF_FONT_GOTHIC, 6.0)
        for _, row in matrix.iterrows():
            if y < 58:
                c.showPage()
                c.setPageSize(page_size)
                y = height - 30
            y -= row_h
            c.setFillColor(colors.white)
            c.rect(table_x, y, width - margin * 2, row_h, fill=1, stroke=1)
            draw_shift_table_vertical_lines(y, y + row_h)
            c.setFillColor(colors.black)
            c.drawString(table_x + 3, y + 4, str(row["職員名"])[:8])
            x = table_x + staff_w
            for d in range(1, last_day + 1):
                val = str(row[str(d)] or "")
                c.drawCentredString(x + day_w / 2, y + 4, val)
                x += day_w
            for key in summary_keys:
                c.drawCentredString(x + summary_w / 2, y + 4, str(row.get(key, "")))
                x += summary_w

    # 不足・警告を下部表示。長い場合は右端で省略して、PDF外にはみ出さないようにする。
    y -= 16
    c.setFont(PDF_FONT_GOTHIC, 7.2)
    warning_width = width - margin * 2

    def draw_warning_line(label, text, y_pos):
        c.setFillColor(colors.HexColor("#b00020"))
        draw_single_line(c, f"{label}: {text}", margin, y_pos, warning_width, PDF_FONT_GOTHIC, 7.2, color=colors.HexColor("#b00020"))
        return y_pos - 9.5

    ng = shortage[shortage["状態"] != "OK"]
    if not ng.empty:
        warning_text = " / ".join([f"{str(r['日付'])[-2:]}日 {r['不足']}" for _, r in ng.head(12).iterrows()])
        if len(ng) > 12:
            warning_text += f" / ほか{len(ng) - 12}件"
        y = draw_warning_line("不足日", warning_text, y)
    if checks is not None and not checks.empty:
        warning_text = " / ".join([f"{r['日付']} {r['職員名']} {r['種類']}" for _, r in checks.head(8).iterrows()])
        if len(checks) > 8:
            warning_text += f" / ほか{len(checks) - 8}件"
        y = draw_warning_line("確認事項", warning_text, y)
    if limit_checks is not None and not limit_checks.empty:
        warning_text = " / ".join([f"{r['職員名']} {r['種類']}" for _, r in limit_checks.head(8).iterrows()])
        if len(limit_checks) > 8:
            warning_text += f" / ほか{len(limit_checks) - 8}件"
        y = draw_warning_line("上限確認", warning_text, y)
    c.setFillColor(colors.black)

    c.save()
    buffer.seek(0)
    return buffer.getvalue()


def make_shift_calendar_pdf(
    year,
    month,
    shift_df,
    staff_list=None,
    selected_staff_names=None,
    finalized=False,
):
    """月間シフトカレンダーをA4横PDFで作成する。"""
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab がインストールされていません。")

    init_pdf_fonts()

    year = int(year)
    month = int(month)
    staff_list = [str(x) for x in (staff_list or []) if str(x or "").strip()]
    selected_staff_names = [str(x) for x in (selected_staff_names or []) if str(x or "").strip()]
    target_names = selected_staff_names or staff_list
    target_label = "全員" if not selected_staff_names else "、".join(selected_staff_names)

    df = shift_df.copy() if isinstance(shift_df, pd.DataFrame) else pd.DataFrame()
    if not df.empty:
        df = df.copy()
        df["shift_date"] = df["shift_date"].astype(str)
        df["staff_name"] = df["staff_name"].astype(str)
        df["shift_kind"] = df["shift_kind"].astype(str)
        month_prefix = f"{year}-{month:02d}-"
        df = df[df["shift_date"].str.startswith(month_prefix)]
        if target_names:
            df = df[df["staff_name"].isin(target_names)]

    buffer = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buffer, pagesize=page_size)
    width, height = page_size

    margin = 18
    title_y = height - 26
    status_text = "確定済み" if finalized else "未確定"
    c.setFont(PDF_FONT_GOTHIC, 15)
    c.drawString(margin, title_y, f"ひだまり帳 月間シフトカレンダー　{year}年{month}月")
    c.setFont(PDF_FONT_GOTHIC, 8)
    c.drawRightString(width - margin, title_y, f"出力日時: {today_jst().strftime('%Y-%m-%d %H:%M')}　確定状態: {status_text}")
    draw_single_line(c, f"出力対象: {target_label}", margin, title_y - 14, width - margin * 2, PDF_FONT_GOTHIC, 8)

    legend_y = title_y - 28
    c.setFont(PDF_FONT_GOTHIC, 7.2)
    c.drawString(margin, legend_y, "凡例：日=日勤　管=管理業務　夜=夜勤　明=夜勤明け　希=希望休　有=有休　休=休み　他=その他")

    cal = calendar.Calendar(firstweekday=6)
    weeks = cal.monthdatescalendar(year, month)
    table_x = margin
    table_top = legend_y - 12
    table_w = width - margin * 2
    table_h = table_top - 22
    header_h = 16
    cell_w = table_w / 7
    cell_h = (table_h - header_h) / len(weeks)
    weekdays = ["日", "月", "火", "水", "木", "金", "土"]

    def short_shift_label(value):
        mapping = {
            "日勤": "日",
            "管": "管",
            "管理業務": "管",
            "夜勤": "夜",
            "夜勤明け": "明",
            "休み": "休",
            "希望休": "希",
            "有休": "有",
            "その他": "他",
        }
        return mapping.get(str(value or "").strip(), str(value or "").strip())

    def grouped_lines_for_day(target_date):
        if df.empty:
            return []
        day_text = target_date.strftime("%Y-%m-%d")
        day_df = df[df["shift_date"] == day_text]
        if day_df.empty:
            return []
        order = ["日", "管", "夜", "明", "休", "希", "有", "他"]
        grouped = {key: [] for key in order}
        for _, row in day_df.iterrows():
            staff_name = pdf_text(row.get("staff_name"))
            label = short_shift_label(row.get("shift_kind"))
            if not staff_name or not label:
                continue
            grouped.setdefault(label, []).append(staff_name)

        lines = []
        for label in order:
            names = grouped.get(label, [])
            if names:
                lines.append(f"{label}: {'、'.join(names)}")
        for label in sorted(k for k in grouped.keys() if k not in order):
            names = grouped.get(label, [])
            if names:
                lines.append(f"{label}: {'、'.join(names)}")
        return lines

    c.setFillColor(colors.HexColor("#f3eee6"))
    c.rect(table_x, table_top - header_h, table_w, header_h, fill=1, stroke=1)
    c.setFillColor(colors.black)
    c.setFont(PDF_FONT_GOTHIC, 9)
    for idx, label in enumerate(weekdays):
        x = table_x + idx * cell_w
        if idx == 0:
            c.setFillColor(colors.HexColor("#b00020"))
        elif idx == 6:
            c.setFillColor(colors.HexColor("#1f4e79"))
        else:
            c.setFillColor(colors.black)
        c.drawCentredString(x + cell_w / 2, table_top - 11, label)
    c.setFillColor(colors.black)

    grid_top = table_top - header_h
    for week_idx, week in enumerate(weeks):
        y_top = grid_top - week_idx * cell_h
        for day_idx, target_date in enumerate(week):
            x = table_x + day_idx * cell_w
            y = y_top - cell_h
            in_month = target_date.month == month
            if not in_month:
                fill = colors.HexColor("#f5f5f5")
            elif day_idx == 0:
                fill = colors.HexColor("#fff1ee")
            elif day_idx == 6:
                fill = colors.HexColor("#eef5ff")
            else:
                fill = colors.white
            c.setFillColor(fill)
            c.rect(x, y, cell_w, cell_h, fill=1, stroke=1)
            if not in_month:
                continue

            if day_idx == 0:
                c.setFillColor(colors.HexColor("#b00020"))
            elif day_idx == 6:
                c.setFillColor(colors.HexColor("#1f4e79"))
            else:
                c.setFillColor(colors.black)
            c.setFont(PDF_FONT_GOTHIC, 8.5)
            c.drawString(x + 3, y_top - 11, f"{target_date.day}")

            c.setFillColor(colors.black)
            line_y = y_top - 23
            for line in grouped_lines_for_day(target_date)[:7]:
                draw_single_line(c, line, x + 3, line_y, cell_w - 6, PDF_FONT_GOTHIC, 6.4)
                line_y -= 8
                if line_y < y + 5:
                    break

    c.setFont(PDF_FONT_GOTHIC, 7)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawString(margin, 12, "未保存の画面編集はPDFに反映されません。必要に応じて先にシフト表を保存してください。")
    c.setFillColor(colors.black)

    c.save()
    buffer.seek(0)
    return buffer.getvalue()



def get_kot_pattern_code(shift_kind):
    """KING OF TIME用パターンコード。secrets/envで上書き可能。"""
    key_map = {
        "日勤": (KOT_DAY_PATTERN_CODE_KEY, "日勤"),
        "管": (KOT_DAY_PATTERN_CODE_KEY, "日勤"),
        "管理業務": (KOT_DAY_PATTERN_CODE_KEY, "日勤"),
        "夜勤": (KOT_NIGHT_PATTERN_CODE_KEY, "夜勤"),
    }
    key, default = key_map.get(str(shift_kind), ("", str(shift_kind or "")))
    if key:
        return get_secret_or_env(key, default=default)
    return default


def kot_time_value(time_text, next_day=False):
    """KING OF TIMEの対象日HH:mm形式へ変換する。"""
    value = str(time_text or "").strip()
    if not value:
        return ""
    prefix = "翌日" if next_day else "当日"
    if value.startswith(("当日", "翌日", "前日")):
        return value
    return f"{prefix}{value}"


def kot_break_minutes(shift_kind):
    """KING OF TIME CSV用の休憩予定時間。必要に応じて施設運用に合わせて変更する。"""
    if str(shift_kind) in ["日勤", "管", "管理業務"]:
        return 60
    if str(shift_kind) == "夜勤":
        return 120
    return ""


def kot_full_day_leave_name(shift_kind):
    """KING OF TIME CSVの全日休暇名。施設設定に合わせてsecrets/envで上書きできる。"""
    if str(shift_kind) == "休み":
        return get_secret_or_env(KOT_REST_LEAVE_NAME_KEY, default="公休")
    if str(shift_kind) == "有休":
        return get_secret_or_env(KOT_PAID_LEAVE_NAME_KEY, default="有休")
    if str(shift_kind) == "希望休":
        return get_secret_or_env(KOT_HOPE_REST_LEAVE_NAME_KEY, default="希望休")
    return ""


def make_king_of_time_shift_csv(
    year,
    month,
    get_staff_shifts,
    get_staff_code_map,
    normalize_staff_name,
    default_shift_times,
):
    """
    ひだまり帳の月間シフトをKING OF TIMEのスケジュールデータCSV向けに出力する。
    出力列：勤務日・従業員コード・パターンコード・出勤予定・退勤予定・休憩予定時間・全日休暇・備考
    """
    year = int(year)
    month = int(month)
    last_day = calendar.monthrange(year, month)[1]
    start = f"{year}-{month:02d}-01"
    end = f"{year}-{month:02d}-{last_day:02d}"
    df = get_staff_shifts(start, end)

    columns = ["勤務日", "従業員コード", "パターンコード", "出勤予定", "退勤予定", "休憩予定時間", "全日休暇", "備考"]
    if df is None or df.empty:
        empty_df = pd.DataFrame(columns=columns)
        return empty_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

    code_map = get_staff_code_map(active_only=False)
    rows = []
    df = df.sort_values(["shift_date", "staff_name", "shift_kind", "id"])
    for _, r in df.iterrows():
        shift_kind = str(r.get("shift_kind") or "").strip()
        staff_name = normalize_staff_name(r.get("staff_name", ""))
        if not staff_name:
            continue

        shift_date = str(r.get("shift_date") or "").strip()
        try:
            work_date = pd.to_datetime(shift_date).date().strftime("%Y%m%d")
        except Exception:
            work_date = shift_date.replace("-", "").replace("/", "")

        staff_code = str(code_map.get(staff_name, "") or "").strip()
        pattern_code = ""
        start_plan = ""
        end_plan = ""
        break_minutes = ""
        full_day_leave = ""
        memo_parts = []
        original_memo = str(r.get("memo") or "").strip()

        if shift_kind in ["日勤", "管", "管理業務", "夜勤"]:
            pattern_code = get_kot_pattern_code(shift_kind)
            stime = str(r.get("start_time") or "").strip()
            etime = str(r.get("end_time") or "").strip()
            if not stime or not etime:
                default_start, default_end, default_next_day = default_shift_times("日勤" if shift_kind in ["管", "管理業務"] else shift_kind)
                stime = stime or default_start
                etime = etime or default_end
                next_day = bool(default_next_day)
            else:
                next_day = bool(int(r.get("next_day") or 0))
            start_plan = kot_time_value(stime, next_day=False)
            end_plan = kot_time_value(etime, next_day=next_day)
            break_minutes = kot_break_minutes(shift_kind)
            if shift_kind in ["管", "管理業務"]:
                memo_parts.append("管理業務")
        elif shift_kind in ["休み", "有休", "希望休"]:
            full_day_leave = kot_full_day_leave_name(shift_kind)
            memo_parts.append(f"{shift_kind}（ひだまり帳）")
        elif shift_kind == "夜勤明け":
            memo_parts.append("夜勤明け（ひだまり帳）")
        elif shift_kind == "その他":
            memo_parts.append("その他（ひだまり帳）")
        else:
            memo_parts.append(shift_kind)

        if original_memo:
            memo_parts.append(original_memo)

        rows.append({
            "勤務日": work_date,
            "従業員コード": staff_code,
            "パターンコード": pattern_code,
            "出勤予定": start_plan,
            "退勤予定": end_plan,
            "休憩予定時間": break_minutes,
            "全日休暇": full_day_leave,
            "備考": " / ".join([x for x in memo_parts if x]),
        })

    out_df = pd.DataFrame(rows, columns=columns)
    return out_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def make_staff_shift_excel(
    year,
    month,
    get_staff_shifts_month,
    create_shift_matrix,
    create_shift_shortage_table,
    create_shift_quality_check_table,
    create_shift_limit_check_table,
    get_shift_month_status,
):
    """
    確定済みシフトをExcel（xlsx）で出力する。
    勤務表シートに月間表、別シートに不足・確認事項・上限確認を出す。
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl がインストールされていません。requirements.txt に openpyxl を追加してください。")

    year = int(year)
    month = int(month)
    df = get_staff_shifts_month(year, month, include_prev_day=True)
    matrix = create_shift_matrix(df, year, month)
    shortage = create_shift_shortage_table(df, year, month)
    checks = create_shift_quality_check_table(df, year, month)
    limit_checks = create_shift_limit_check_table(matrix)
    status = get_shift_month_status(year, month)
    last_day = calendar.monthrange(year, month)[1]

    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = "勤務表"

    title_fill = PatternFill("solid", fgColor="F3EEE6")
    header_fill = PatternFill("solid", fgColor="E7EEF8")
    weekend_sun_fill = PatternFill("solid", fgColor="FCE4D6")
    weekend_sat_fill = PatternFill("solid", fgColor="DDEBF7")
    summary_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, len(matrix.columns) if not matrix.empty else 8))
    ws.cell(1, 1).value = f"従業員勤務表　{year}年{month}月"
    ws.cell(1, 1).font = Font(bold=True, size=16)
    ws.cell(1, 1).fill = title_fill
    ws.cell(1, 1).alignment = left

    status_text = "確定" if status.get("is_confirmed") else "作成中"
    ws.cell(2, 1).value = "状態"
    ws.cell(2, 2).value = status_text
    ws.cell(2, 3).value = "確定日時"
    ws.cell(2, 4).value = status.get("confirmed_at", "") or ""
    ws.cell(2, 5).value = "出力日"
    ws.cell(2, 6).value = today_jst().strftime("%Y-%m-%d")

    ws.cell(3, 1).value = "凡例"
    ws.cell(3, 2).value = "日=日勤 8:30〜17:30 / 管=管理業務 8:30〜17:30 / 夜=夜勤 16:30〜翌9:30 / 明=夜勤明け / 希=希望休 / 有=有休 / 他=その他 / 休みは日別セルには表示しません"
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=max(8, len(matrix.columns) if not matrix.empty else 8))

    start_row = 5
    if matrix is None or matrix.empty:
        ws.cell(start_row, 1).value = "この月のシフトはまだ登録されていません。"
    else:
        columns = list(matrix.columns)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(start_row, col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill
            if str(col_name).isdigit():
                d = int(col_name)
                weekday = date(year, month, d).weekday()  # 月=0 日=6
                if weekday == 6:
                    cell.fill = weekend_sun_fill
                elif weekday == 5:
                    cell.fill = weekend_sat_fill
            elif col_name in ["日勤", "管", "夜勤", "明", "休み", "希望休", "有休", "他", "合計", "最大連勤"]:
                cell.fill = summary_fill

        for row_idx, (_, row) in enumerate(matrix.iterrows(), start=start_row + 1):
            for col_idx, col_name in enumerate(columns, start=1):
                value = row.get(col_name, "")
                if pd.isna(value):
                    value = ""
                cell = ws.cell(row_idx, col_idx)
                cell.value = value
                cell.alignment = center if col_name != "職員名" else left
                cell.border = border
                if str(col_name).isdigit():
                    d = int(col_name)
                    weekday = date(year, month, d).weekday()
                    if weekday == 6:
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    elif weekday == 5:
                        cell.fill = PatternFill("solid", fgColor="EAF3F8")
                elif col_name in ["日勤", "管", "夜勤", "明", "休み", "希望休", "有休", "他", "合計", "最大連勤"]:
                    cell.fill = PatternFill("solid", fgColor="F2F8EE")

        ws.freeze_panes = "B6"
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{start_row + len(matrix)}"

        # 列幅調整
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            if col_name == "職員名":
                ws.column_dimensions[letter].width = 14
            elif str(col_name).isdigit():
                ws.column_dimensions[letter].width = 4
            else:
                ws.column_dimensions[letter].width = 8

    for row in range(1, 4):
        for col in range(1, max(8, (len(matrix.columns) if not matrix.empty else 8)) + 1):
            ws.cell(row, col).alignment = left if col <= 2 else center

    def add_df_sheet(sheet_name, data_df, empty_message):
        safe_name = sheet_name[:31]
        s = wb.create_sheet(safe_name)
        if data_df is None or data_df.empty:
            s.cell(1, 1).value = empty_message
            s.cell(1, 1).font = Font(bold=True)
            s.column_dimensions["A"].width = 60
            return s
        cols = list(data_df.columns)
        for cidx, col in enumerate(cols, start=1):
            cell = s.cell(1, cidx)
            cell.value = col
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            s.column_dimensions[get_column_letter(cidx)].width = 18 if col not in ["内容", "理由"] else 55
        for ridx, (_, r) in enumerate(data_df.iterrows(), start=2):
            for cidx, col in enumerate(cols, start=1):
                value = r.get(col, "")
                if pd.isna(value):
                    value = ""
                cell = s.cell(ridx, cidx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left" if col in ["内容", "理由"] else "center", vertical="center", wrap_text=True)
                cell.border = border
        s.freeze_panes = "A2"
        s.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data_df)+1}"
        return s

    ng = shortage[shortage["状態"] != "OK"] if shortage is not None and not shortage.empty else pd.DataFrame()
    add_df_sheet("不足確認", ng, "日勤2名・夜勤1名の不足日はありません。")
    add_df_sheet("確認事項", checks, "夜勤翌日勤務・明け翌日勤務・5連勤以上・希望休重複などの確認事項はありません。")
    add_df_sheet("上限確認", limit_checks, "職員別の日勤・夜勤・合計勤務回数は上限内です。")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

